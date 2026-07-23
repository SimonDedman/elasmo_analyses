#!/usr/bin/env python3
"""Resolve a better title_language (and year) for library PDFs.

The non_extractable report's ``title_language`` was detected from page-1
OCR text, which for scanned/historical volumes is blank or garbage — so
languages are frequently wrong (e.g. a German volume mislabelled French).

This script re-derives language from cleaner signals, in priority order:
  1. the filename title fragment (always present, human-typed),
  2. the corpus title + journal name (when the file matches a
     viz_data.csv entry by first-author surname + year +/- 1),
which are far more reliable than OCR'd page-1 text. Year comes from the
filename (authoritative), enabling Fraktur routing for old German works.

Usage:
    python scripts/resolve_pdf_language.py \
        --report outputs/non_extractable_relang_large.xlsx \
        --out    outputs/non_extractable_relang_large_resolved.xlsx
"""

import argparse
import csv
import re
import unicodedata
from pathlib import Path

from openpyxl import load_workbook, Workbook

PROJECT = Path("/media/simon/data/Documents/Si Work/PostDoc Work/EEA/2025/Data Panel")
CORPUS = PROJECT / "outputs/viz_data.csv"

# Language markers: stopwords + orthographic cues. Scored against the
# available text (filename fragment + corpus title + journal). Labels match
# ocr_library.py LANG_MAP keys.
STOPWORDS = {
    "German": {"der", "die", "das", "und", "von", "aus", "zur", "zum", "über",
               "eine", "einer", "neue", "neuen", "fische", "haie", "haifische",
               "systematik", "verzeichniss", "untersuchung", "beitrag",
               "gattung", "arten", "zähne", "kreide", "meer", "nach", "bei"},
    "French": {"les", "des", "une", "dans", "pour", "sur", "par", "aux", "avec",
               "nouvelle", "poissons", "requins", "raies", "étude", "sur",
               "histoire", "naturelle", "côte", "mer", "genre", "espèces"},
    "Spanish": {"los", "las", "del", "por", "con", "una", "para", "sobre",
                "entre", "tiburón", "tiburones", "rayas", "peces", "estudio",
                "nueva", "nuevo", "sobre", "marinos", "terrenos", "fósiles"},
    "Portuguese": {"dos", "das", "uma", "com", "para", "sobre", "peixes",
                   "tubarão", "estudo", "nova", "novo", "análise"},
    "Italian": {"dei", "delle", "una", "sulla", "sui", "pesci", "squali",
                "studio", "nuova", "nuovo", "specie", "della"},
    "Dutch": {"het", "een", "van", "de", "en", "vissen", "haaien", "over",
              "nieuwe", "bijdrage", "onderzoek"},
    "Latin": {"de", "et", "sive", "seu", "naturae", "systema", "genera",
              "species", "piscium", "nova", "descriptio", "historia",
              "animalium", "regni", "methodus"},
}
ACCENTS = [
    (re.compile(r"[äöüß]"), "German"),
    (re.compile(r"[çèêëàâ]"), "French"),
    (re.compile(r"[ñ]"), "Spanish"),
    (re.compile(r"[ãõ]"), "Portuguese"),
]
# Journal-name cues (strong signal when a corpus match provides the journal).
JOURNAL_CUES = {
    "German": ("für", "zeitschrift", "archiv", "berichte", "paläont",
               "abhandlungen", "jahrbuch", "mitteilungen", "naturkunde"),
    "French": ("bulletin", "revue", "annales", "société", "mémoires",
               "muséum", "française"),
    "Spanish": ("boletín", "revista", "anales", "sociedad", "española"),
    "Portuguese": ("boletim", "revista", "arquivos", "sociedade"),
    "Italian": ("bollettino", "rivista", "atti", "società", "italiana"),
}


def norm(s: str) -> str:
    return unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode().lower()


def detect_language(fragment: str, corpus_title: str = "", journal: str = "") -> str:
    """Return best language label, or '' if no clear signal."""
    text = " ".join([fragment, corpus_title]).lower()
    toks = set(re.findall(r"[a-zà-ÿ]+", text))
    scores = {lang: len(toks & words) for lang, words in STOPWORDS.items()}
    # accent cues (on raw, unnormalised text)
    for rx, lang in ACCENTS:
        if rx.search(text):
            scores[lang] = scores.get(lang, 0) + 2
    # journal cues (strong)
    jl = journal.lower()
    for lang, cues in JOURNAL_CUES.items():
        if any(c in jl for c in cues):
            scores[lang] = scores.get(lang, 0) + 3
    best = max(scores, key=scores.get)
    return best if scores[best] >= 2 else ""


def load_corpus_index():
    idx = {}
    if not CORPUS.exists():
        return idx
    with open(CORPUS, encoding="utf-8", errors="replace") as f:
        for d in csv.DictReader(f):
            y = str(d.get("year", "")).split(".")[0]
            authors = d.get("authors", "") or ""
            first = re.split(r"[;,]", authors)[0].strip()
            surn = norm(first.split()[-1]) if first.split() else ""
            if surn and y.isdigit():
                rec = (d.get("title", ""), d.get("journal", ""))
                for yy in (int(y) - 1, int(y), int(y) + 1):
                    idx.setdefault((surn, yy), []).append((frozenset(norm(d.get("title", "")).split()), rec))
    return idx


def parse_filename(fname: str):
    m = re.match(r"(.+?)(?:\.etal)?\.(\d{4})\.(.+?)(?:\.pdf)?$", fname)
    if not m:
        return "", None, ""
    author = m.group(1).split()[-1] if m.group(1).split() else m.group(1)
    return norm(author), int(m.group(2)), m.group(3)


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--report", required=True)
    ap.add_argument("--out", required=True)
    args = ap.parse_args()

    idx = load_corpus_index()
    wb = load_workbook(args.report, read_only=True, data_only=True)
    ws = wb.active
    out = Workbook()
    ows = out.active
    ows.append(["year", "filename", "title_language", "resolved_year", "lang_source", "open_file"])

    n = changed = matched = 0
    for year, fname, old_lang, link in ws.iter_rows(min_row=2, values_only=True):
        if not fname:
            continue
        n += 1
        surn, fyear, frag = parse_filename(str(fname))
        corpus_title = journal = ""
        source = "filename"
        cands = idx.get((surn, fyear), []) if fyear else []
        if cands and frag:
            fragset = set(norm(frag).split())
            best = max(cands, key=lambda c: len(fragset & c[0]))
            if len(fragset & best[0]) >= 2:  # title-word overlap confirms match
                corpus_title, journal = best[1]
                source = "corpus"
                matched += 1
        new_lang = detect_language(frag, corpus_title, journal)
        final_lang = new_lang or str(old_lang)
        if new_lang and new_lang != str(old_lang):
            changed += 1
        ows.append([year, fname, final_lang, fyear or year,
                    source if new_lang else "kept", link])

    out.save(args.out)
    print(f"rows: {n}   corpus-matched: {matched}   language changed: {changed}")
    print(f"wrote {args.out}")


if __name__ == "__main__":
    main()
