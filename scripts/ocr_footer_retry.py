#!/usr/bin/env python3
"""Aggressive-preprocessing OCR pass for old scans where --force-ocr
produced only sparse body text.

Targets the 7 "scanned footer only, rest of text remains image" rows
from ocr_retry_report.xlsx. Uses:
  - --force-ocr: rasterise everything, redo OCR
  - --deskew + --rotate-pages: fix scan skew/rotation
  - --clean: unpaper preprocessing to remove artifacts before OCR
  - --oversample 600: higher DPI for tesseract
  - --tesseract-oem 1: LSTM-only engine (better for clean text)

Writes outputs/ocr_footer_retry_report.xlsx.
"""

import os
import subprocess
from pathlib import Path
from urllib.parse import quote

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

PROJECT = Path("/media/simon/data/Documents/Si Work/PostDoc Work/EEA/2025/Data Panel")
PDF_ROOT = Path("/media/simon/data/Documents/Si Work/Papers & Books/SharkPapers")
IN_XLSX = PROJECT / "outputs/ocr_retry_report.xlsx"
OUT_XLSX = PROJECT / "outputs/ocr_footer_retry_report.xlsx"
LOG_FILE = PROJECT / "logs/ocr_footer_retry_log.txt"


def whole_doc_alpha(pdf: Path, timeout: int = 60) -> int:
    try:
        r = subprocess.run(
            ["pdftotext", str(pdf), "-"],
            capture_output=True, timeout=timeout,
        )
        return sum(1 for c in r.stdout.decode("utf-8", errors="replace") if c.isalpha())
    except Exception:
        return 0


def aggressive_ocr(pdf: Path) -> tuple[bool, str]:
    tmp = pdf.with_suffix(".aggressive.tmp.pdf")
    try:
        r = subprocess.run(
            ["ocrmypdf",
             "--force-ocr",
             "--deskew",
             "--rotate-pages",
             "--clean",
             "--oversample", "600",
             "--tesseract-oem", "1",
             "--output-type", "pdf",
             "--language", "eng",
             "--quiet",
             str(pdf), str(tmp)],
            capture_output=True, timeout=1800,
        )
        if r.returncode != 0 or not tmp.exists() or tmp.stat().st_size == 0:
            if tmp.exists():
                tmp.unlink()
            err = r.stderr.decode("utf-8", errors="replace")[:200]
            return False, f"aggressive ocr failed: {err}"
        os.replace(tmp, pdf)
        return True, f"aggressive OCR OK ({pdf.stat().st_size:,} bytes)"
    except subprocess.TimeoutExpired:
        if tmp.exists():
            tmp.unlink()
        return False, "timeout (>30 min)"
    except Exception as e:
        if tmp.exists():
            tmp.unlink()
        return False, f"exception: {e}"


def main():
    # Select rows where note matches "scanned footer only"
    wb_in = load_workbook(IN_XLSX, data_only=True)
    ws_in = wb_in.active
    targets = []
    for row in ws_in.iter_rows(min_row=2, values_only=True):
        year, fname, lang, prev, notes, action, b, a, final, outcome, *_ = (
            row + (None,) * 11
        )[:11]
        if not notes:
            continue
        n = str(notes).lower()
        if "scanned footer" in n or ("footer scan issue" in n and "brute force" in n):
            targets.append({
                "year": str(year), "filename": str(fname),
                "note": str(notes), "path": PDF_ROOT / str(year) / str(fname),
            })
    wb_in.close()
    print(f"Loaded {len(targets)} footer-only targets")

    results = []
    for i, t in enumerate(targets, 1):
        p = t["path"]
        before = whole_doc_alpha(p) if p.exists() else -1
        if not p.exists():
            outcome = "FILE_MISSING"
            after = -1
        else:
            ok, msg = aggressive_ocr(p)
            after = whole_doc_alpha(p)
            outcome = ("AGGRESSIVE_OK: " if ok else "AGGRESSIVE_FAIL: ") + msg
        delta = after - before if (before >= 0 and after >= 0) else -1
        results.append({**t, "before_alpha": before, "after_alpha": after,
                        "delta": delta, "outcome": outcome})
        status = "GAIN" if delta > 100 else "SAME" if delta == 0 else "LOSS" if delta < 0 else "SMALL_GAIN"
        print(f"  [{i}/{len(targets)}] {status} ({before}→{after}, Δ={delta:+d})  "
              f"{t['filename'][:60]}")

    # Summary
    total_before = sum(r["before_alpha"] for r in results if r["before_alpha"] >= 0)
    total_after = sum(r["after_alpha"] for r in results if r["after_alpha"] >= 0)
    print(f"\nSummary:")
    print(f"  Total alpha chars before: {total_before:,}")
    print(f"  Total alpha chars after:  {total_after:,}")
    print(f"  Net gain: {total_after - total_before:+,}")

    LOG_FILE.parent.mkdir(exist_ok=True)
    with open(LOG_FILE, "w") as f:
        for r in results:
            f.write(f"{r['filename']}  before={r['before_alpha']} "
                    f"after={r['after_alpha']} delta={r['delta']:+d}  "
                    f"— {r['outcome']}\n")

    # XLSX
    wb = Workbook()
    ws = wb.active
    ws.title = "Footer Retry Report"
    headers = ["year", "filename", "note", "before_alpha", "after_alpha",
               "delta", "outcome", "open_file"]
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")
    link_font = Font(color="0000FF", underline="single")
    for i, r in enumerate(results, 2):
        ws.cell(row=i, column=1, value=r["year"])
        ws.cell(row=i, column=2, value=r["filename"])
        ws.cell(row=i, column=3, value=r["note"])
        ws.cell(row=i, column=4, value=r["before_alpha"])
        ws.cell(row=i, column=5, value=r["after_alpha"])
        ws.cell(row=i, column=6, value=r["delta"])
        ws.cell(row=i, column=7, value=r["outcome"])
        uri = "file://" + quote(str(r["path"]), safe="/:")
        c = ws.cell(row=i, column=8, value="Open")
        c.hyperlink = uri
        c.font = link_font
    for col_idx, h in enumerate(headers, 1):
        max_len = len(h)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=2,
                                max_row=ws.max_row, values_only=True):
            val = str(row[0]) if row[0] is not None else ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 70)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    wb.save(OUT_XLSX)
    print(f"\nReport: {OUT_XLSX}")
    print(f"Log:    {LOG_FILE}")


if __name__ == "__main__":
    main()
