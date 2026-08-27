"""SAFE-EDIT guard for the shared conference coverage matrix.

The xlsx is shared with colleagues via NAS -> Google Drive, so cells may be
edited by other people. Before ANY programmatic edit, run this to:
  1. diff the current file against the last committed (agent-generated) version
     -> surfaces cells humans changed;
  2. validate every cell against the schema (status vocabulary + "Location;
     Status [— note]" format) -> surfaces typos / off-schema entries.

Usage: ./venv/bin/python scripts/conf_abstracts/check_coverage_matrix.py
Exit 0 = clean & unchanged; 1 = human edits and/or schema issues to review.

Workflow rule (see memory feedback_coverage_matrix_shared_edits):
  - ALWAYS run this before editing the matrix.
  - REPORT any human-changed cells to Simon; do NOT overwrite them without
    his explicit confirmation.
  - Fix/flag schema violations (typos) but confirm before changing others' cells.
"""
import subprocess
import sys
import tempfile
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parents[2]
# The NAS->Google-Drive sync operates on the outputs/ copy (colleague edits land
# there); database/ is the local working copy. Check both and flag divergence.
REL = "outputs/conference_coverage_matrix.xlsx"
REL_DB = "database/conference_coverage_matrix.xlsx"
CUR = REPO / REL

STATUSES = {"Missing", "Hardcopy", "Programme", "Digital", "OCR",
            "Schedule", "Ingested", "Pending", "NA"}


def baseline_copy():
    """Extract the last committed version of the xlsx to a temp file."""
    try:
        blob = subprocess.run(["git", "show", f"HEAD:{REL}"], cwd=REPO,
                              capture_output=True).stdout
        if not blob:
            return None
        tmp = Path(tempfile.mkdtemp()) / "baseline.xlsx"
        tmp.write_bytes(blob)
        return tmp
    except Exception:
        return None


def cells(path):
    wb = openpyxl.load_workbook(path)
    ws = wb["Coverage"] if "Coverage" in wb.sheetnames else wb.active
    out = {}
    for row in ws.iter_rows():
        for c in row:
            if c.value not in (None, ""):
                out[c.coordinate] = str(c.value)
    return out


def parse_status(text):
    """Return the status token from a 'Location; Status — note' cell, or None."""
    if ";" not in text:
        return None
    tail = text.split(";", 1)[1].strip()
    return tail.split("—")[0].strip().split()[0] if tail else None


def main():
    if not CUR.exists():
        print(f"MISSING: {CUR}")
        sys.exit(1)
    cur = cells(CUR)
    issues = 0

    # 1. diff vs last committed (human edits arrive via the Google Drive sync)
    base_path = baseline_copy()
    if base_path:
        base = cells(base_path)
        changed = {k: (base.get(k), cur[k]) for k in cur
                   if base.get(k) != cur[k]}
        removed = {k: base[k] for k in base if k not in cur}
        if changed or removed:
            issues += 1
            print("=== HUMAN EDITS since last agent version (DO NOT overwrite "
                  "without Simon's OK) ===")
            for k, (old, new) in sorted(changed.items()):
                print(f"  {k}: {old!r} -> {new!r}")
            for k, old in sorted(removed.items()):
                print(f"  {k}: {old!r} -> (cleared)")
        else:
            print("No changes vs last committed version.")
    else:
        print("(no git baseline found — cannot diff)")

    # 1b. the two local copies must agree (a divergence = a stale sync)
    db = REPO / REL_DB
    if db.exists():
        c2 = cells(db)
        if c2 != cur:
            issues += 1
            ndiff = len(set(cur) ^ set(c2)) + sum(1 for k in set(cur) & set(c2) if cur[k] != c2[k])
            print(f"\n=== DIVERGENCE: outputs/ and database/ copies differ ({ndiff} cells) "
                  "— likely a stale NAS/Drive sync; reconcile before editing ===")

    # 2. schema validation (skip header row 1 and the Year column A)
    print("\n=== schema check (Coverage data cells) ===")
    bad = 0
    for coord, text in sorted(cur.items()):
        col = "".join(ch for ch in coord if ch.isalpha())
        rownum = int("".join(ch for ch in coord if ch.isdigit()))
        if rownum == 1 or col == "A" or text in ("—",):
            continue
        st = parse_status(text)
        if st is None:
            print(f"  {coord}: no ';Status' -> {text!r}")
            bad += 1
        elif st not in STATUSES:
            print(f"  {coord}: unknown status {st!r} -> {text!r}")
            bad += 1
    if bad:
        issues += 1
        print(f"  {bad} cell(s) off-schema (typo or new status?).")
    else:
        print("  all data cells match the schema.")

    sys.exit(1 if issues else 0)


if __name__ == "__main__":
    main()


def legend_report(cur_path, ref_wb=None):
    """Surface human edits to the Legend & Notes wording. The builder now
    preserves these (data/matrix_legend.json), but they must still be visible so
    nobody silently reverts them in code."""
    from openpyxl import load_workbook
    import json
    store = REPO / "data" / "matrix_legend.json"
    try:
        ws = load_workbook(cur_path)["Legend & Notes"]
    except Exception as e:
        print(f"  legend unreadable: {e}")
        return
    sheet = {r[0]: r[1] for r in ws.iter_rows(min_row=2, max_row=9, max_col=2, values_only=True) if r[0]}
    if not store.exists():
        print("  no legend store yet; run the builder once to capture current wording")
        return
    saved = json.loads(store.read_text(encoding="utf-8")).get("meanings", {})
    drift = [k for k, v in sheet.items() if saved.get(k) != v]
    print("\n=== Legend & Notes ===")
    if drift:
        print("  edited in the sheet since the last build (will be preserved on next build):")
        for k in drift:
            print(f"    {k}: {sheet[k]}")
    else:
        print("  legend wording matches the preserved store.")


if __name__ == "__main__" and "--legend" in __import__("sys").argv:
    legend_report(REPO / "outputs" / "conference_coverage_matrix.xlsx")
