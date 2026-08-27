"""Build the conference-abstract coverage matrix (xlsx).

Sheets (order): Legend & Notes | Coverage | Dashboard.
- Coverage: year x series, colour-coded red->green by status. ASIH/JMIH/Other
  collapsed to one 'ASIH/JMIH' column (they share one source book); AES kept
  separate (carries the elasmo count). Cells = "Location; Status". 'No
  conference' cells are left blank/unshaded.
- Dashboard: per-society totals bar chart + abstracts-over-time line chart.

Sources: the conference_abstracts DB; Cat Gordon (EEA locations + status);
Brit Finucci (OCS pending); Carylanne (1992-2024 hardcopy); AESconfLocations
(ASIH/JMIH host cities 1916-2025); known SI years.
"""
import csv
import sqlite3
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

REPO = Path(__file__).resolve().parents[2]
DB = REPO / "database" / "conference_abstracts.db"
OUT = REPO / "outputs" / "conference_coverage_matrix.xlsx"
ASIH_CSV = REPO / "database" / "asih_meetings.csv"

# status order (red -> green). 'Programme' distinguishes a schedule/grid PDF we
# hold (NO abstract bodies, abstract book still needed) from 'Digital' = a
# parseable abstract-book PDF we hold and can ingest.
# Colour = distance from a complete, searchable set of abstracts, worst first.
# The ramp runs red -> orange -> amber -> yellow -> green in this order:
#   Missing < Schedule < Hardcopy < OCR < Programme < Pending < Digital < Ingested
# Rationale (Simon, 2026-08-27): "Digital" means the book is in hand and only
# needs processing, so it is nearly done and reads light green; "OCR" means the
# abstracts are in but from a degraded scan we still want re-sourced, so it sits
# back beside "Hardcopy" (Simon, 2026-08-27: functionally the same problem);
# "Schedule" is titles without abstract text, useful but
# not abstracts, so it sits just above Missing; "Hardcopy" and "Programme" are
# the same practical problem (we need the abstract book) and sit adjacent.
STATUS_ORDER = ["Missing", "Schedule", "Hardcopy", "OCR", "Programme",
                "Pending", "Digital", "Ingested"]
FILL = {
    "Missing": "E06666",    # red: nothing exists anywhere
    "Schedule": "ED9C6B",   # red-orange: titles/authors only, no abstract text
    "Hardcopy": "F6B26B",   # dark orange: paper book known, nothing digital
    "OCR": "F9CB9C",        # light orange: abstracts in but from a degraded scan
    "Programme": "FFD966",  # amber: digital programme held, book still needed
    "Pending": "FFE599",    # yellow: a named contact has it or is looking
    "Digital": "B6D7A8",    # light green: book in hand, extraction pending
    "Ingested": "6AA84F",   # green: done
    "NA": None,
}
# (location, status, action-note). status conveys what we HOLD; the note says
# what's still NEEDED so the sheet is self-documenting.
# EEA abstract books extracted via Fable (2026-08-14): 731 abstracts across 11
# meetings into conference_abstracts_fable.db. "Ingested" notes carry the count.
# Host cities for the meetings we have no book for. 2003 and 2005-2009 from the
# EEA's own meetings page (eulasmo.org/scientific-meetings, read 2026-08-27);
# meeting numbers there confirm an unbroken annual series from 1997 (1st), which
# matches the 2002 Cardiff booklet calling itself the 6th. 1998 Lisbon is from
# APECE's account of hosting that year. 1997, 1999-2001 still unknown.
EEA_EARLY = {
    2003: ("San Marino", "7th EEA; Ali Hood searching her paperwork (2026-08-27)"),
    2005: ("Monaco", "9th EEA; Ali Hood searching her paperwork (2026-08-27)"),
    2006: ("Hamburg", "10th EEA; Ali Hood searching her paperwork (2026-08-27)"),
    2007: ("Brest", "11th EEA; Ali Hood searching her paperwork (2026-08-27)"),
    2008: ("Lisbon", "12th EEA; Ali Hood searching her paperwork (2026-08-27)"),
    2009: ("Palma de Majorca", "13th EEA; Ali Hood searching her paperwork (2026-08-27)"),
}
EEA_EARLY_LOC = {1998: "Lisbon"}

# Host group per year (Cat Gordon, 2026-07-28), drives the per-group asks.
EEA_HOST = {
    2010: "IEG (Ireland)", 2011: "DEG (Germany)", 2012: "GRIS (Italy)",
    2013: "Shark Trust (UK)", 2014: "NEV (Netherlands)", 2015: "APECE (Portugal)",
    2016: "Shark Trust (UK)", 2017: "NEV (Netherlands)", 2018: "APECE (Portugal)",
    2019: "GRIS (Italy)", 2021: "NEV (Netherlands)",
    2022: "Shark Trust / Submon / Lamna", 2023: "Shark Trust (UK)",
    2024: "iSea (Greece)", 2025: "NEV (Netherlands)", 2026: "Shark Trust (UK)",
}
EEA = {
    2002: ("Cardiff", "Digital", "abstract booklet (Word) found by Ali Hood 2026-08-27; queued for extraction"),
    2004: ("London", "Ingested", "55 abstracts (Fable)"),
    2010: ("Galway", "Hardcopy", "Cat holds a hardcopy but has no scanning capacity; digital copy needed from IEG (Ireland)"),
    2011: ("Berlin", "Ingested", "58 abstracts (Fable)"),
    2012: ("Milan", "Hardcopy", "Cat holds a hardcopy but has no scanning capacity; digital copy needed from GRIS (Italy)"),
    2013: ("Plymouth", "Ingested", "93 abstracts (Fable)"),
    2014: ("Leeuwarden", "Ingested", "61 abstracts (Fable)"),
    2015: ("Peniche", "Digital", "19th EEA Book of Abstracts (99pp, born-digital) held by Simon all along; queued for extraction"),
    2016: ("Bristol", "Ingested", "93 abstracts (Fable)"),
    2017: ("Amsterdam", "Schedule", "62 talks ingested, NO abstract bodies; Cat holds a hardcopy but has no scanning capacity; abstract book needed from NEV"),
    2018: ("Peniche", "Ingested", "75 abstracts (Fable)"),
    2019: ("Rende", "Ingested", "136 abstracts (Fable)"),
    2020: ("?", "Missing", "unknown whether a meeting was held: Cat thinks it was online for covid but was on maternity leave and is unsure; organiser unknown; abstracts unknown"),
    2021: ("Leiden", "Programme", "programme only; abstract book needed from NEV"),
    2022: ("Valencia (=SI2022)", "Digital", "full SI2022 abstract book received 2026-08-27; queued for extraction"),
    2023: ("Brighton", "Ingested", "oral 64 + poster 34 = 98 abstracts (Fable)"),
    2024: ("Thessaloniki", "Digital", "clean abstract book received from Cat 2026-08-27; queued for extraction"),
    2025: ("Rotterdam", "Programme", "agenda only; Cat has no abstract book and is unsure one was produced; NEV/Irene to confirm whether it exists"),
    2026: ("online", "Pending", "will be digital (Shark Trust hosting)"),
}
SI = {2010: ("Cairns", "Missing", "find source"),
      2014: ("Durban", "Missing", "find source"),
      2018: ("Joao Pessoa", "Ingested", ""),
      2022: ("Valencia", "Digital", "full abstract book received 2026-08-27; queued (schedule already ingested)"),
      2026: ("Colombo", "Ingested", "")}
# JMIH/ASIH years where we hold a PDF that is NOT an ingestable abstract book:
# a grid programme (2017/2019 — detail doesn't extract; get the abstract book)
# or a schedule-only programme (2026 — has a text layer, no abstract bodies).
JMIH_PROGRAMME = {
    2017: "grid programme only — abstract book needed (Carylanne)",
    2019: "grid programme only — abstract book needed (Carylanne)",
    2026: "schedule-only programme (OCR text layer OK, not yet parsed) — abstract book needed",
}
# OCS meetings we can evidence from public sources (read 2026-08-27); Brit
# Finucci is confirming the full series and locations.
OCS = {
    2012: ("Adelaide", "joint with ASFB; Brit Finucci confirming"),
    2018: ("North Stradbroke Is.", "Moreton Bay Research Station; Brit Finucci confirming"),
    2024: ("Geelong", "Brit Finucci confirming"),
}
SERIES = ["AES", "ASIH", "HL", "SSAR", "NIA", "EEA", "SI"]


def load_asih_locations():
    locs = {}
    with open(ASIH_CSV) as fh:
        for row in csv.DictReader(fh):
            try:
                locs[int(row["year"])] = row["location"].strip()
            except (ValueError, KeyError):
                pass
    return locs


JMIH_LOC = load_asih_locations()
JMIH_LOC.setdefault(2026, "New Orleans, LA")


def db_meeting_status():
    con = sqlite3.connect(str(DB))
    out = {}
    # LEFT JOIN so a held-but-unparsed book (e.g. a degraded scan that yielded
    # 0 records) still surfaces; body count makes the status content-aware
    # (a 520-page "programme" whose records carry bodies IS an abstract book).
    for yr, meeting, doc, isocr, n, el, bodies in con.execute(
        """select m.year,m.meeting,m.doc_type,m.is_ocr,count(a.abstract_id) n,
                  sum(a.is_elasmo) el,
                  sum(case when length(a.abstract_text) > 50 then 1 else 0 end) bodies
           from meetings m left join abstracts a on a.meeting_id=m.meeting_id
           where m.meeting in ('JMIH','ASIH') group by m.meeting_id"""):
        d = out.setdefault(yr, dict(abstract=False, schedule=False, ocr=False,
                                    ocr_failed=False, elasmo=0))
        has_bodies = (bodies or 0) >= max(2, 0.5 * n)
        if n > 1 and (doc == "abstract_book" or has_bodies):
            d.update(abstract=True, ocr=bool(isocr))
            d["elasmo"] += el or 0
        elif doc == "abstract_book" and isocr:
            d["ocr_failed"] = True  # scan held; OCR recovered nothing usable
        elif doc == "program_book" and n > 20:
            d["schedule"] = True
            d["elasmo"] += el or 0
    con.close()
    return out


def db_aes_web():
    """AES abstracts harvested from elasmo.org (meeting='AES'): year -> count."""
    con = sqlite3.connect(str(DB))
    out = {yr: n for yr, n in con.execute(
        "select m.year, count(*) from meetings m join abstracts a using(meeting_id) "
        "where m.meeting='AES' group by m.year")}
    con.close()
    return out


def db_year_society():
    con = sqlite3.connect(str(DB))
    cnt = defaultdict(int)
    for yr, soc, meeting, n in con.execute(
        """select m.year,a.society,m.meeting,count(*) from abstracts a
           join meetings m on a.meeting_id=m.meeting_id
           where m.meeting in ('JMIH','ASIH','SI','EEA','AES') group by m.year,a.society,m.meeting"""):
        if meeting in ("SI", "EEA", "AES"):
            cnt[(yr, meeting)] += n
        else:
            for s in (soc or "").split("|"):
                if s in SERIES:
                    cnt[(yr, s)] += n
    con.close()
    return cnt


def meeting_cell(year, db):
    """Return (location, status, note, elasmo_count)."""
    loc = JMIH_LOC.get(year, "?")
    d = db.get(year)
    if d and d["abstract"]:
        if d["ocr"]:
            return loc, "OCR", "degraded scan (needs_review) — flatbed re-scan planned", d["elasmo"]
        return loc, "Ingested", "", d["elasmo"]
    if d and d["ocr_failed"]:
        return loc, "OCR", "degraded phone scan — 0 abstracts recovered — flatbed re-scan needed (Carylanne)", 0
    if d and d["schedule"]:
        return loc, "Schedule", "programme ingested (no abstract bodies) — abstract book needed", d["elasmo"]
    if year in JMIH_PROGRAMME:
        return loc, "Programme", JMIH_PROGRAMME[year], 0
    if 1992 <= year <= 2024:
        return loc, "Hardcopy", "Carylanne has hardcopy (1992-2024) — get abstract book", 0
    return loc, "Missing", "", 0


def build():
    db = db_meeting_status()
    yr_soc = db_year_society()
    aes_web = db_aes_web()
    wb = Workbook()
    wb.remove(wb.active)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ---- Legend & Notes (first, per Simon's tab swap) ----
    leg = wb.create_sheet("Legend & Notes")
    leg.append(["Status", "Meaning", "Colour"])
    for c in leg[1]:
        c.font = Font(bold=True)
    # listed worst-to-best so the colour ramp reads top to bottom
    MEANING = {
        "Missing": "No known source anywhere.",
        "Schedule": "Programme ingested: titles, authors, presentation type. NO abstract text. Useful, but it isn't abstracts.",
        "Hardcopy": "A physical book is known to exist (Carylanne, Cat, Ali) but nothing digital. Needs scanning, or a digital copy from the host.",
        "Programme": "A digital programme is held, but not the abstract book. Same practical need as Hardcopy: get the book.",
        "Pending": "A named contact has it or is looking for it, not yet received (EEA: Cat and Ali; OCS: Brit).",
        "OCR": "Abstracts ingested, but from a degraded scan and flagged needs_review. Still worth re-sourcing a clean copy.",
        "Digital": "Abstract book in hand and ingestable. Only extraction remains.",
        "Ingested": "Full abstracts ingested into the database.",
    }
    order = [(st, MEANING[st]) for st in STATUS_ORDER]
    for i, (st, mean) in enumerate(order, 2):
        leg.cell(row=i, column=1, value=st)
        leg.cell(row=i, column=2, value=mean)
        leg.cell(row=i, column=3).fill = PatternFill("solid", fgColor=FILL[st])
    for note in [
        "", "NOTES:",
        "- 'ASIH/JMIH' = the American joint meeting (ASIH pre-1997, JMIH from 1997). ASIH/JMIH/HL/SSAR/NIA share one source book, collapsed to this column to remove duplicates.",
        "- 'AES' (American Elasmobranch Society, founded 1983) kept separate — cell shows elasmo-abstract count where ingested. Pre-1983 = no conference (blank).",
        "- Host cities 1916-2025 from github.com/SimonDedman/AESconfLocations; 2026 = New Orleans.",
        "- 'OCR' = degraded 1997-2004 JMIH phone-photo scans (all needs_review); Carylanne's flatbed re-scans would upgrade the non-AES content.",
        "- AES 1985-2005: full abstracts harvested from elasmo.org/meetings/abstracts/abst<YYYY>/ (1,305 abstracts, 2026-08-25); JMIH-book copies of the same AES talks were removed. No AES source found for 1983-84 or 2006+ online.",
        "- ASIH/JMIH pre-1992: Carylanne's archive starts 1992 → host city shown but 'Missing'.",
        "- EEA from Cat Gordon (Shark Trust). 2004-2019 + 2023 ingested via Fable and merged 2026-08-25. Received 2026-08-27: 2002 Cardiff booklet, a clean 2024 Thessaloniki book, and the full SI2022 Valencia abstract book (all queued for extraction). 2025 Rotterdam abstract book is still a corrupt file. Cat holds hardcopies of 2010/2012/2015/2017 but has no scanning capacity, so those are being sought digitally from the host groups; Ali Hood is searching her paperwork for 2003-2009. EEA began in 1997 (2002 Cardiff was the 6th).",
        "- OCS (Oceania Chondrichthyan Soc, ~2011+, biennial): Brit Finucci collating, pending council — years/locations TBC.",
        "- Blank/unshaded cell = no conference that year for that series.",
        "- Per-society counts & trends: see the Dashboard tab.",
    ]:
        leg.append([note])
    leg.column_dimensions["A"].width = 12
    leg.column_dimensions["B"].width = 95

    # ---- Coverage ----
    ws = wb.create_sheet("Coverage")
    cols = ["Year", "ASIH/JMIH", "AES", "EEA", "OCS", "SI"]
    ws.append(cols)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="434343")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    def put(r, col, text, status):
        cell = ws.cell(row=r, column=col, value=text)
        if FILL.get(status):
            cell.fill = PatternFill("solid", fgColor=FILL[status])
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = border
        cell.font = Font(size=9)

    def txt(loc, st, note):
        return f"{loc}; {st}" + (f" — {note}" if note else "")

    r = 2
    for year in range(1916, 2027):
        ws.cell(row=r, column=1, value=year).font = Font(bold=True)
        ws.cell(row=r, column=1).border = border
        loc, st, note, el = meeting_cell(year, db)
        put(r, 2, txt(loc, st, note), st)
        # AES (founded 1983)
        if year in aes_web:
            # AES abstracts harvested from elasmo.org supersede the JMIH-book status
            put(r, 3, txt(f"{loc} ({aes_web[year]})", "Ingested",
                          "AES abstracts from elasmo.org (full bodies)"), "Ingested")
        elif year >= 1983:
            aloc = f"{loc}" + (f" ({el})" if el else "")
            put(r, 3, txt(aloc, st, note), st)
        else:
            put(r, 3, "", "NA")
        # EEA
        if year in EEA:
            l, s, n = EEA[year]; put(r, 4, txt(l, s, n), s)
        elif year in EEA_EARLY:
            loc_e, note_e = EEA_EARLY[year]
            put(r, 4, txt(loc_e, "Pending", note_e), "Pending")
        elif 1997 <= year <= 2001:
            loc_e = EEA_EARLY_LOC.get(year, "?")
            nth = {1997: "1st", 1998: "2nd", 1999: "3rd", 2000: "4th", 2001: "5th"}[year]
            put(r, 4, txt(loc_e, "Missing",
                          f"{nth} EEA; no known abstract source"
                          + ("" if year in EEA_EARLY_LOC else "; host city also unknown")), "Missing")
        else:
            put(r, 4, "", "NA")
        # OCS (biennial ~2012+)
        if year in OCS:
            oloc, onote = OCS[year]
            put(r, 5, txt(oloc, "Pending", onote), "Pending")
        elif year >= 2012 and year % 2 == 0:
            put(r, 5, "?; Pending — Brit Finucci collating; year/location to confirm", "Pending")
        else:
            put(r, 5, "", "NA")
        # SI (quadrennial)
        if year in SI:
            l, s, n = SI[year]; put(r, 6, txt(l, s, n), s)
        else:
            put(r, 6, "", "NA")
        r += 1
    ws.column_dimensions["A"].width = 6
    for col in "BCDEF":
        ws.column_dimensions[col].width = 44
    ws.freeze_panes = "B2"

    # ---- Dashboard ----
    dash = wb.create_sheet("Dashboard")
    dash["A1"] = "Conference abstract coverage — dashboard"
    dash["A1"].font = Font(bold=True, size=13)
    # data table: year x series counts
    hdr_row = 3
    dash.cell(row=hdr_row, column=1, value="Year").font = Font(bold=True)
    for j, s in enumerate(SERIES, 2):
        dash.cell(row=hdr_row, column=j, value=s).font = Font(bold=True)
    years = list(range(1992, 2027))
    for i, y in enumerate(years, hdr_row + 1):
        dash.cell(row=i, column=1, value=y)
        for j, s in enumerate(SERIES, 2):
            dash.cell(row=i, column=j, value=yr_soc.get((y, s), 0))
    last = hdr_row + len(years)
    # totals row
    tot_row = last + 1
    dash.cell(row=tot_row, column=1, value="TOTAL").font = Font(bold=True)
    for j, s in enumerate(SERIES, 2):
        dash.cell(row=tot_row, column=j,
                  value=sum(yr_soc.get((y, s), 0) for y in years)).font = Font(bold=True)

    # Bar chart: totals per society
    bar = BarChart()
    bar.title = "Total abstracts ingested per society/series"
    bar.type = "col"
    bar.y_axis.title = "Abstracts"
    data = Reference(dash, min_col=2, max_col=len(SERIES) + 1, min_row=tot_row, max_row=tot_row)
    cats = Reference(dash, min_col=2, max_col=len(SERIES) + 1, min_row=hdr_row, max_row=hdr_row)
    bar.add_data(data, from_rows=True, titles_from_data=False)
    bar.set_categories(cats)
    bar.legend = None
    bar.height, bar.width = 8, 16
    dash.add_chart(bar, "I3")

    # Line chart: over time, one line per society
    line = LineChart()
    line.title = "Abstracts over time, by society/series"
    line.y_axis.title = "Abstracts"
    line.x_axis.title = "Year"
    ldata = Reference(dash, min_col=2, max_col=len(SERIES) + 1,
                      min_row=hdr_row, max_row=last)
    lcats = Reference(dash, min_col=1, min_row=hdr_row + 1, max_row=last)
    line.add_data(ldata, titles_from_data=True)
    line.set_categories(lcats)
    line.height, line.width = 10, 20
    dash.add_chart(line, "I20")
    dash.column_dimensions["A"].width = 7

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"wrote {OUT}  (sheets: {wb.sheetnames})")


if __name__ == "__main__":
    build()
