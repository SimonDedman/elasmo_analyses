"""Export the abstracts DB to parquet (elasmo subset, corpus-aligned), JSON
(full, nested authors), and a formatted xlsx review sheet."""
import json

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font


def _authors_by_abstract(con):
    rows = con.execute(
        "SELECT abstract_id, full_name, position, is_presenter, affiliation "
        "FROM authors ORDER BY abstract_id, position").fetchall()
    out = {}
    for aid, name, pos, pres, aff in rows:
        out.setdefault(aid, []).append(
            dict(full_name=name, position=pos, is_presenter=bool(pres),
                 affiliation=aff))
    return out


def to_parquet_elasmo(con, path) -> int:
    """Elasmo-only rows, columns aligned to literature_review_enriched.parquet
    (title, year, authors, abstract, plus provenance). Returns row count."""
    df = pd.read_sql_query(
        """SELECT a.abstract_id, m.meeting, m.year, a.title,
                  a.abstract_text AS abstract, a.keywords, a.presentation_type,
                  a.society, a.society_basis, a.elasmo_basis, a.award,
                  a.session_name, a.confidence, a.needs_review, m.source_pdf
           FROM abstracts a JOIN meetings m ON a.meeting_id=m.meeting_id
           WHERE a.is_elasmo=1""", con)
    auth = _authors_by_abstract(con)
    df["authors"] = df["abstract_id"].map(
        lambda i: "; ".join(a["full_name"] for a in auth.get(i, [])))
    df.to_parquet(path, index=False)
    return len(df)


def to_json(con, path):
    """Full set, one object per abstract with nested authors."""
    df = pd.read_sql_query(
        """SELECT a.*, m.meeting, m.year, m.source_pdf, m.doc_type
           FROM abstracts a JOIN meetings m ON a.meeting_id=m.meeting_id""", con)
    auth = _authors_by_abstract(con)
    records = []
    for _, r in df.iterrows():
        rec = r.to_dict()
        rec["authors"] = auth.get(r["abstract_id"], [])
        records.append(rec)
    with open(path, "w") as fh:
        json.dump(records, fh, indent=1, default=str)


def to_xlsx(con, path):
    """Formatted workbook (frozen/bold/autofiltered) for sharing/review."""
    df = pd.read_sql_query(
        """SELECT m.meeting, m.year, a.program_number, a.title,
                  a.presentation_type, a.society, a.society_basis, a.is_elasmo,
                  a.elasmo_basis, a.award, a.length_words, a.confidence,
                  a.needs_review, m.source_pdf
           FROM abstracts a JOIN meetings m ON a.meeting_id=m.meeting_id
           ORDER BY m.year, m.meeting, a.program_number""", con)
    df.to_excel(path, index=False)
    wb = load_workbook(path)
    ws = wb.active
    ws.freeze_panes = "A2"
    for c in ws[1]:
        c.font = Font(bold=True)
    ws.auto_filter.ref = ws.dimensions
    for col in ws.columns:
        w = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(w + 2, 10), 70)
    wb.save(path)
