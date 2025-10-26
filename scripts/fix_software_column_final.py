#!/usr/bin/env python3
"""
Final fix for software and data_science columns:
- data_science: TRUE/FALSE (not 1/0)
- software: TEXT (software names), not boolean
"""

import pandas as pd
from pathlib import Path
from openpyxl import load_workbook

# Paths
excel_path = Path("data/Techniques DB for Panel Review.xlsx")

# Software mapping for known techniques
SOFTWARE_MAPPING = {
    'maxent': 'MaxEnt software',
    'marxan': 'Marxan',
    'zonation': 'Zonation',
    'isospace': 'IsoSpace',
    'siar': 'SIAR, MixSIAR',
    'mixsiar': 'MixSIAR',
    'brt': 'gbm, dismo',
    'boosted regression': 'gbm, dismo',
    'nmds': 'vegan (R)',
    'permanova': 'vegan (R)',
    'boris': 'BORIS',
    'r package': 'R',
    'python': 'Python',
}


def get_software_for_technique(tech_name, description, synonyms):
    """Determine software for a technique based on name/description."""
    text = f"{tech_name} {description} {synonyms}".lower() if pd.notna(tech_name) and pd.notna(description) else ""

    software_list = []
    for key, software in SOFTWARE_MAPPING.items():
        if key in text:
            if software not in software_list:
                software_list.append(software)

    return ', '.join(software_list) if software_list else None


def main():
    print("=" * 80)
    print("FINAL FIX: SOFTWARE AND DATA_SCIENCE COLUMNS")
    print("=" * 80)

    # Load Excel
    print(f"\nLoading: {excel_path}")
    wb = load_workbook(excel_path)
    ws = wb['Full_List']

    # Get column mapping
    headers = [cell.value for cell in ws[1]]
    col_mapping = {header: idx + 1 for idx, header in enumerate(headers)}

    # Create backup
    print(f"\n📦 Creating backup...")
    backup_path = excel_path.with_suffix('.backup5.xlsx')
    import shutil
    shutil.copy2(excel_path, backup_path)
    print(f"✓ Backup created: {backup_path}")

    # Fix data_science column (convert to TRUE/FALSE)
    if 'data_science' in col_mapping:
        print(f"\nFixing 'data_science' column:")
        col_idx = col_mapping['data_science']
        true_count = 0
        false_count = 0

        for row_idx in range(2, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value

            # Convert to boolean
            if value == 1 or value == 1.0 or value == '1' or value is True:
                ws.cell(row=row_idx, column=col_idx, value=True)
                true_count += 1
            else:
                ws.cell(row=row_idx, column=col_idx, value=False)
                false_count += 1

        print(f"  TRUE: {true_count}, FALSE: {false_count}")

    # Fix software column (convert 0/1 back to actual software names or None)
    if 'software' in col_mapping:
        print(f"\nFixing 'software' column (restoring text values):")
        software_col = col_mapping['software']
        tech_col = col_mapping['technique_name']
        desc_col = col_mapping.get('description')
        syn_col = col_mapping.get('synonyms')

        populated = 0
        cleared = 0

        for row_idx in range(2, ws.max_row + 1):
            current_value = ws.cell(row=row_idx, column=software_col).value

            # Get technique info
            tech_name = ws.cell(row=row_idx, column=tech_col).value
            description = ws.cell(row=row_idx, column=desc_col).value if desc_col else ""
            synonyms = ws.cell(row=row_idx, column=syn_col).value if syn_col else ""

            # Determine correct software value
            software = get_software_for_technique(tech_name, description, synonyms)

            if software:
                ws.cell(row=row_idx, column=software_col, value=software)
                populated += 1
                if current_value in [0, 1]:
                    print(f"  Row {row_idx}: {tech_name:40s} -> {software}")
            else:
                ws.cell(row=row_idx, column=software_col, value=None)
                cleared += 1

        print(f"\n  Populated: {populated} techniques with software names")
        print(f"  Cleared: {cleared} techniques (no software)")

    # Save
    wb.save(excel_path)
    print(f"\n✓ Saved successfully")

    # Verify
    print("\n" + "=" * 80)
    print("VERIFICATION")
    print("=" * 80)

    df_verify = pd.read_excel(excel_path, sheet_name='Full_List')

    print("\ndata_science column:")
    true_count = (df_verify['data_science'] == True).sum()
    false_count = (df_verify['data_science'] == False).sum()
    print(f"  TRUE: {true_count}, FALSE: {false_count}")

    print("\nsoftware column:")
    has_software = df_verify['software'].notna().sum()
    no_software = df_verify['software'].isna().sum()
    print(f"  Has software name: {has_software}")
    print(f"  No software: {no_software}")

    if has_software > 0:
        print(f"\n  Techniques with software (first 20):")
        software_rows = df_verify[df_verify['software'].notna()].head(20)
        for idx, row in software_rows.iterrows():
            print(f"    {row['technique_name']:40s}: {row['software']}")

    print("\n" + "=" * 80)
    print("COMPLETE")
    print("=" * 80)
    print("\n✓ All fixes applied!")


if __name__ == "__main__":
    main()
