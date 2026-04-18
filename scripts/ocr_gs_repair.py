#!/usr/bin/env python3
"""Ghostscript-repair pass for PDFs that pikepdf couldn't fix.

Reads `outputs/ocr_retry_report.xlsx`, picks rows where final_status is
STILL_NON_EXTRACTABLE and action was 'repair+ocr' (i.e. pikepdf repair
failed). Runs `gs -o tmp.pdf -sDEVICE=pdfwrite ...` which re-serialises
the PDF through Ghostscript, rebuilding xref and page tree. Then retries
ocrmypdf (skip-text with force-ocr fallback).

Writes `outputs/ocr_gs_retry_report.xlsx`.
"""

import os
import subprocess
from pathlib import Path
from urllib.parse import quote

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

PROJECT = Path("/media/simon/data/Documents/Si Work/PostDoc Work/EEA/2025/Data Panel")
PDF_ROOT = Path("/media/simon/data/Documents/Si Work/Papers & Books/SharkPapers")
IN_XLSX = PROJECT / "outputs/ocr_retry_report.xlsx"
OUT_XLSX = PROJECT / "outputs/ocr_gs_retry_report.xlsx"
LOG_FILE = PROJECT / "logs/ocr_gs_retry_log.txt"

MIN_ALPHA_WHOLE = 200


def whole_doc_alpha(pdf: Path, timeout: int = 120) -> int:
    try:
        r = subprocess.run(
            ["pdftotext", str(pdf), "-"],
            capture_output=True, timeout=timeout,
        )
        return sum(1 for c in r.stdout.decode("utf-8", errors="replace") if c.isalpha())
    except Exception:
        return 0


def gs_repair(pdf: Path) -> tuple[bool, str]:
    """Re-serialise PDF through Ghostscript, rebuilding structure."""
    tmp = pdf.with_suffix(".gsrepair.tmp.pdf")
    try:
        r = subprocess.run(
            ["gs", "-o", str(tmp),
             "-sDEVICE=pdfwrite",
             "-dPDFSETTINGS=/default",
             "-dBATCH", "-dNOPAUSE", "-dQUIET",
             "-dPDFSTOPONERROR=false",
             str(pdf)],
            capture_output=True, timeout=600,
        )
        if not tmp.exists() or tmp.stat().st_size == 0:
            if tmp.exists():
                tmp.unlink()
            err = r.stderr.decode("utf-8", errors="replace")[:180]
            return False, f"gs produced empty output: {err}"
        # Sanity: GS output should be at least 1KB for any real PDF
        if tmp.stat().st_size < 1024:
            tmp.unlink()
            return False, f"gs output suspiciously small ({tmp.stat().st_size} bytes)"
        os.replace(tmp, pdf)
        return True, f"gs repair OK ({pdf.stat().st_size:,} bytes)"
    except subprocess.TimeoutExpired:
        if tmp.exists():
            tmp.unlink()
        return False, "gs timeout"
    except Exception as e:
        if tmp.exists():
            tmp.unlink()
        return False, f"gs exception: {e}"


def ocr_after_repair(pdf: Path) -> tuple[bool, str]:
    tmp = pdf.with_suffix(".ocr.tmp.pdf")
    for flag in ["--skip-text", "--force-ocr"]:
        try:
            r = subprocess.run(
                ["ocrmypdf", flag, "--output-type", "pdf",
                 "--language", "eng", "--quiet",
                 str(pdf), str(tmp)],
                capture_output=True, timeout=900,
            )
            if r.returncode == 0 and tmp.exists() and tmp.stat().st_size > 0:
                os.replace(tmp, pdf)
                return True, f"ocr OK ({flag})"
            if tmp.exists():
                tmp.unlink()
        except Exception as e:
            if tmp.exists():
                tmp.unlink()
            return False, f"ocr exception: {e}"
    return False, "ocrmypdf failed both --skip-text and --force-ocr"


def main():
    wb_in = load_workbook(IN_XLSX, data_only=True)
    ws_in = wb_in.active
    targets = []
    for row in ws_in.iter_rows(min_row=2, values_only=True):
        # year, filename, language, prev_status, notes, action, before, after, final, outcome, open_file
        year, fname, lang, prev, notes, action, before, after, final, outcome, *_ = (
            row + (None,) * 11
        )[:11]
        if final != "STILL_NON_EXTRACTABLE":
            continue
        if action != "repair+ocr":
            continue
        targets.append({
            "year": str(year), "filename": str(fname),
            "notes": str(notes or ""), "prev_outcome": str(outcome or ""),
            "path": PDF_ROOT / str(year) / str(fname),
        })
    wb_in.close()
    print(f"Loaded {len(targets)} repair+ocr targets")

    results = []
    for i, t in enumerate(targets, 1):
        p = t["path"]
        before = whole_doc_alpha(p, timeout=60) if p.exists() else -1
        if not p.exists():
            outcome = "FILE_MISSING"
            after = -1
        else:
            ok_r, msg_r = gs_repair(p)
            if not ok_r:
                outcome = f"GS_REPAIR_FAIL: {msg_r}"
                after = whole_doc_alpha(p, timeout=60)
            else:
                ok_o, msg_o = ocr_after_repair(p)
                after = whole_doc_alpha(p, timeout=120)
                outcome = f"GS_{'OK' if ok_r else 'FAIL'} → OCR_{'OK' if ok_o else 'FAIL'}: {msg_o}"

        final = ("FLIPPED" if before < MIN_ALPHA_WHOLE <= after
                 else "STILL_NON_EXTRACTABLE" if after < MIN_ALPHA_WHOLE
                 else "ALREADY_EXTRACTABLE")
        results.append({**t, "before_alpha": before, "after_alpha": after,
                        "outcome": outcome, "final_status": final})
        print(f"  [{i}/{len(targets)}] {final}  ({before}→{after})  "
              f"{t['filename'][:70]}")

    # Summary
    counts = {}
    for r in results:
        counts[r["final_status"]] = counts.get(r["final_status"], 0) + 1
    print("\nSummary:")
    for k, v in sorted(counts.items()):
        print(f"  {k}: {v}")

    LOG_FILE.parent.mkdir(exist_ok=True)
    with open(LOG_FILE, "w") as f:
        for r in results:
            f.write(f"{r['final_status']}: {r['filename']} "
                    f"before={r['before_alpha']} after={r['after_alpha']} — "
                    f"{r['outcome']}\n")

    wb = Workbook()
    ws = wb.active
    ws.title = "Ghostscript Retry"
    headers = ["year", "filename", "notes", "before_alpha", "after_alpha",
               "final_status", "outcome", "open_file"]
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")
    link_font = Font(color="0000FF", underline="single")
    for i, r in enumerate(results, 2):
        ws.cell(row=i, column=1, value=r["year"])
        ws.cell(row=i, column=2, value=r["filename"])
        ws.cell(row=i, column=3, value=r["notes"])
        ws.cell(row=i, column=4, value=r["before_alpha"])
        ws.cell(row=i, column=5, value=r["after_alpha"])
        ws.cell(row=i, column=6, value=r["final_status"])
        ws.cell(row=i, column=7, value=r["outcome"])
        uri = "file://" + quote(str(r["path"]), safe="/:")
        c = ws.cell(row=i, column=8, value="Open")
        c.hyperlink = uri
        c.font = link_font
    for col_idx, h in enumerate(headers, 1):
        max_len = len(h)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=2,
                                max_row=ws.max_row, values_only=True):
            val = str(row[0]) if row[0] is not None else ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 70)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    wb.save(OUT_XLSX)
    print(f"\nReport: {OUT_XLSX}")
    print(f"Log:    {LOG_FILE}")


if __name__ == "__main__":
    main()
