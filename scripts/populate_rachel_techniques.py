#!/usr/bin/env python3
"""
Populate empty cells in Rachel's technique additions to Techniques DB for Panel Review.xlsx

Fills in missing fields like:
- category_name (derived from discipline)
- is_parent (default FALSE)
- data_source (default 'literature')
- search_query_new (derived from technique_name and synonyms)
- data_collection_technology, statistical_model, etc. based on technique type
"""

import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
import re

# Paths
excel_path = Path("data/Techniques DB for Panel Review.xlsx")
output_path = Path("data/Techniques DB for Panel Review.xlsx")

# Discipline to category mapping
DISCIPLINE_TO_CATEGORY = {
    'BIO': 'Biological',
    'BEH': 'Behavioral',
    'CON': 'Conservation',
    'DATA': 'Data Science',
    'FISH': 'Fisheries',
    'GEN': 'Genetics',
    'MOV': 'Movement',
    'TRO': 'Trophic'
}

# Technique classification patterns
STATISTICAL_MODELS = {
    'glm', 'gam', 'gamm', 'glmm', 'mixed model', 'bayesian', 'regression',
    'permanova', 'anova', 'ancova', 'manova', 'model', 'analysis of variance'
}

DATA_COLLECTION_TECH = {
    'telemetry', 'tracking', 'tag', 'satellite', 'acoustic', 'archival',
    'camera', 'drone', 'uav', 'video', 'sensor', 'logger',
    'metabarcoding', 'sequencing', 'edna', 'erna'
}

ANALYTICAL_ALGORITHMS = {
    'maxent', 'marxan', 'brt', 'random forest', 'neural network', 'svm',
    'clustering', 'pca', 'nmds', 'mds', 'ordination', 'network analysis'
}

INFERENCE_FRAMEWORKS = {
    'bayesian', 'mcmc', 'maximum likelihood', 'bootstrap', 'permutation',
    'resampling', 'cross-validation'
}

SOFTWARE_TOOLS = {
    'maxent': 'MaxEnt software',
    'marxan': 'Marxan',
    'isospace': 'IsoSpace',
    'siar': 'SIAR, MixSIAR',
    'brt': 'gbm, dismo',
    'nmds': 'vegan (R)',
    'permanova': 'vegan (R)',
}


def classify_technique(tech_name, description, synonyms):
    """
    Classify technique into categories.

    Returns dict with:
    - data_collection_technology
    - statistical_model
    - analytical_algorithm
    - inference_framework
    - software
    """
    text = f"{tech_name} {description} {synonyms}".lower()

    result = {}

    # Data collection technology
    if any(term in text for term in DATA_COLLECTION_TECH):
        result['data_collection_technology'] = True
    else:
        result['data_collection_technology'] = False

    # Statistical model
    if any(term in text for term in STATISTICAL_MODELS):
        result['statistical_model'] = True
    else:
        result['statistical_model'] = False

    # Analytical algorithm
    if any(term in text for term in ANALYTICAL_ALGORITHMS):
        result['analytical_algorithm'] = True
    else:
        result['analytical_algorithm'] = False

    # Inference framework
    if any(term in text for term in INFERENCE_FRAMEWORKS):
        result['inference_framework'] = True
    else:
        result['inference_framework'] = False

    # Software (extract specific tool names)
    software_list = []
    for tool_key, tool_name in SOFTWARE_TOOLS.items():
        if tool_key in text:
            software_list.append(tool_name)

    result['software'] = ', '.join(software_list) if software_list else None

    return result


def create_search_query(tech_name, synonyms):
    """
    Create search query from technique name and synonyms.
    """
    queries = [tech_name]

    if pd.notna(synonyms) and synonyms != 'EMPTY':
        # Split synonyms by comma
        syn_list = [s.strip() for s in str(synonyms).split(',')]
        queries.extend(syn_list)

    # Clean up and format
    queries = [q.strip() for q in queries if q.strip()]

    # Return unique queries joined by OR
    return ' OR '.join(f'"{q}"' for q in queries[:3])  # Limit to 3 terms


def populate_empty_cells(df):
    """
    Fill in empty cells for incomplete rows.
    """
    print("Populating empty cells...")

    rows_modified = 0

    for idx, row in df.iterrows():
        modified = False

        # Skip if technique_name is empty
        if pd.isna(row['technique_name']):
            continue

        tech_name = str(row['technique_name'])
        discipline = str(row['discipline']) if pd.notna(row['discipline']) else ''
        description = str(row['description']) if pd.notna(row['description']) else ''
        synonyms = str(row['synonyms']) if pd.notna(row['synonyms']) else ''

        # 1. Fill category_name from discipline
        if pd.isna(row['category_name']) and discipline in DISCIPLINE_TO_CATEGORY:
            df.at[idx, 'category_name'] = DISCIPLINE_TO_CATEGORY[discipline]
            modified = True

        # 2. Fill is_parent (default FALSE)
        if pd.isna(row['is_parent']):
            df.at[idx, 'is_parent'] = False
            modified = True

        # 3. Fill data_source (default 'literature')
        if pd.isna(row['data_source']):
            df.at[idx, 'data_source'] = 'literature'
            modified = True

        # 4. Fill search_query_new
        if pd.isna(row['search_query_new']) or row['search_query_new'] == '':
            search_q = create_search_query(tech_name, synonyms)
            df.at[idx, 'search_query_new'] = search_q
            modified = True

        # 5. Classify technique and fill categorical columns
        classification = classify_technique(tech_name, description, synonyms)

        if pd.isna(row['data_collection_technology']):
            df.at[idx, 'data_collection_technology'] = classification['data_collection_technology']
            modified = True

        if pd.isna(row['statistical_model']):
            df.at[idx, 'statistical_model'] = classification['statistical_model']
            modified = True

        if pd.isna(row['analytical_algorithm']):
            df.at[idx, 'analytical_algorithm'] = classification['analytical_algorithm']
            modified = True

        if pd.isna(row['inference_framework']):
            df.at[idx, 'inference_framework'] = classification['inference_framework']
            modified = True

        if pd.isna(row['software']) and classification['software']:
            df.at[idx, 'software'] = classification['software']
            modified = True

        # 6. Fill conceptual_field from discipline
        if pd.isna(row['conceptual_field']) and discipline:
            df.at[idx, 'conceptual_field'] = DISCIPLINE_TO_CATEGORY.get(discipline, discipline)
            modified = True

        if modified:
            rows_modified += 1
            print(f"  ✓ Populated: {tech_name}")

    print(f"\n✓ Modified {rows_modified} rows")
    return df


def main():
    print("=" * 80)
    print("POPULATING RACHEL'S TECHNIQUE ADDITIONS")
    print("=" * 80)

    # Load Excel file
    print(f"\nLoading: {excel_path}")
    df = pd.read_excel(excel_path, sheet_name='Full_List')
    print(f"Loaded {len(df):,} techniques")

    # Count empty cells before
    print("\nEmpty cells before:")
    for col in ['category_name', 'is_parent', 'data_source', 'search_query_new',
                'data_collection_technology', 'statistical_model',
                'analytical_algorithm', 'inference_framework', 'software', 'conceptual_field']:
        empty_count = df[col].isna().sum()
        if empty_count > 0:
            print(f"  {col:30s}: {empty_count:>4} empty")

    # Populate empty cells
    df = populate_empty_cells(df)

    # Count empty cells after
    print("\nEmpty cells after:")
    for col in ['category_name', 'is_parent', 'data_source', 'search_query_new',
                'data_collection_technology', 'statistical_model',
                'analytical_algorithm', 'inference_framework', 'software', 'conceptual_field']:
        empty_count = df[col].isna().sum()
        print(f"  {col:30s}: {empty_count:>4} empty")

    # Create backup
    print(f"\n📦 Creating backup...")
    backup_path = output_path.with_suffix('.backup2.xlsx')
    import shutil
    shutil.copy2(output_path, backup_path)
    print(f"✓ Backup created: {backup_path}")

    # Save using openpyxl to preserve formatting
    print(f"\n💾 Saving to: {output_path}")

    # Load workbook
    wb = load_workbook(excel_path)
    ws = wb['Full_List']

    # Get column mapping
    headers = [cell.value for cell in ws[1]]
    col_mapping = {header: idx + 1 for idx, header in enumerate(headers)}

    # Update cells
    for idx, row in df.iterrows():
        excel_row = idx + 2  # Excel is 1-indexed, skip header

        # Update each column
        for col_name in df.columns:
            if col_name in col_mapping:
                col_idx = col_mapping[col_name]
                value = row[col_name]

                # Convert pandas NA/NaN to None for Excel
                if pd.isna(value):
                    value = None

                ws.cell(row=excel_row, column=col_idx, value=value)

    # Save
    wb.save(output_path)
    print(f"✓ Saved successfully")

    print("\n" + "=" * 80)
    print("COMPLETE")
    print("=" * 80)


if __name__ == "__main__":
    main()
