#!/usr/bin/env python3
"""Scan SharkPapers PDFs for non-extractable text and produce XLSX report."""

import os
import re
import subprocess
import unicodedata
from collections import Counter
from multiprocessing import Pool, cpu_count
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

PDF_ROOT = Path("/media/simon/data/Documents/Si Work/Papers & Books/SharkPapers")
OUTPUT = Path("/media/simon/data/Documents/Si Work/PostDoc Work/EEA/2025/Data Panel/outputs/non_extractable_pdfs.xlsx")
MIN_ALPHA = 50  # minimum alphabetic chars on page 1

# ── language detection ────────────────────────────────────────────────

LANG_WORDS = {
    "English": {
        "the", "and", "for", "with", "from", "new", "species", "shark",
        "sharks", "ray", "rays", "marine", "fish", "ecology", "biology",
        "conservation", "distribution", "population", "ocean", "study",
        "analysis", "data", "research", "review", "effects", "habitat",
        "fisheries", "assessment", "management", "south", "north", "west",
        "east", "coastal", "deep", "sea", "reef", "reproductive",
        "feeding", "movement", "tagging", "genetic", "morphology",
    },
    "French": {
        "les", "des", "une", "dans", "pour", "sur", "par", "est", "aux",
        "nouvelle", "poissons", "requins", "raies", "mer", "entre",
        "etude", "biologie", "pêche", "côte",
    },
    "German": {
        "der", "die", "das", "und", "ein", "eine", "von", "aus", "zur",
        "neue", "fische", "haie", "über", "untersuchung", "meer",
    },
    "Spanish": {
        "los", "las", "del", "por", "con", "una", "entre", "sobre",
        "tiburon", "tiburones", "rayas", "peces", "mar", "estudio",
        "nueva", "nuevo",
    },
    "Portuguese": {
        "dos", "das", "uma", "com", "para", "sobre", "entre", "pelo",
        "tubarao", "peixes", "mar", "estudo", "nova", "novo",
    },
}

# Accent / script patterns
ACCENT_LANG = [
    (re.compile(r"[çéèêë]", re.I), "French"),
    (re.compile(r"[äöüß]", re.I), "German"),
    (re.compile(r"[ñ]", re.I), "Spanish"),
    (re.compile(r"[ãõ]", re.I), "Portuguese"),
    (re.compile(r"[šžčřůďťň]", re.I), "Czech/Slovak/Croatian"),
]

SCRIPT_RANGES = [
    ("Cyrillic", "Russian"),
    ("CJK", "Japanese/Chinese"),
    ("Hangul", "Korean"),
    ("Arabic", "Arabic"),
    ("Devanagari", "Hindi"),
    ("Thai", "Thai"),
]


def _detect_script(text: str) -> str | None:
    """Detect non-Latin scripts from Unicode categories."""
    for ch in text:
        name = unicodedata.name(ch, "")
        for script_kw, lang in SCRIPT_RANGES:
            if script_kw.upper() in name.upper():
                return lang
    return None


def guess_language(filename: str) -> str:
    """Guess language from the title portion of a filename like Author.etal.Year.Title.Words.pdf."""
    stem = Path(filename).stem
    # Strip author/year prefix: patterns like "Smith.etal.2020." or "Smith.Jones.2020."
    # Find year (4 digits) and take everything after it
    m = re.search(r"\.\d{4}[a-z]?\.(.+)", stem)
    if m:
        title_part = m.group(1)
    else:
        title_part = stem

    # Replace dots/underscores with spaces, lowercase
    words = set(re.split(r"[.\-_\s]+", title_part.lower()))

    # Check non-Latin scripts first
    script_lang = _detect_script(title_part)
    if script_lang:
        return script_lang

    # Word-based scoring
    scores: Counter = Counter()
    for lang, keywords in LANG_WORDS.items():
        hits = words & keywords
        scores[lang] = len(hits)

    if scores.most_common(1)[0][1] > 0:
        # If there's a tie including English, prefer English (most filenames partially English)
        top_score = scores.most_common(1)[0][1]
        top_langs = [l for l, s in scores.items() if s == top_score]
        if len(top_langs) == 1:
            return top_langs[0]
        # Tiebreak: prefer non-English if English is tied (common words)
        non_eng = [l for l in top_langs if l != "English"]
        if non_eng:
            return non_eng[0]
        return top_langs[0]

    # Accent fallback
    for pattern, lang in ACCENT_LANG:
        if pattern.search(title_part):
            return lang

    return "Unknown"


def check_pdf(pdf_path: str) -> tuple[str, bool] | None:
    """Return (path, is_non_extractable). None on error."""
    try:
        result = subprocess.run(
            ["pdftotext", "-l", "1", pdf_path, "-"],
            capture_output=True, timeout=30,
        )
        text = result.stdout.decode("utf-8", errors="replace")
        alpha_count = sum(1 for c in text if c.isalpha())
        return (pdf_path, alpha_count < MIN_ALPHA)
    except Exception:
        return (pdf_path, True)  # treat errors as non-extractable


def collect_pdfs() -> list[str]:
    """Collect all PDF paths."""
    pdfs = []
    for root, _, files in os.walk(PDF_ROOT):
        for f in files:
            if f.lower().endswith(".pdf"):
                pdfs.append(os.path.join(root, f))
    return sorted(pdfs)


def main():
    pdfs = collect_pdfs()
    print(f"Found {len(pdfs)} PDFs. Scanning with {cpu_count()} workers...")

    with Pool(processes=cpu_count()) as pool:
        results = pool.map(check_pdf, pdfs, chunksize=50)

    # Filter non-extractable
    non_ext = []
    for r in results:
        if r and r[1]:
            p = Path(r[0])
            year = p.parent.name
            fname = p.name
            lang = guess_language(fname)
            non_ext.append((year, fname, lang, r[0]))

    non_ext.sort(key=lambda x: (x[0], x[1]))
    print(f"\nNon-extractable PDFs: {len(non_ext)} / {len(pdfs)}")

    # Language breakdown
    lang_counts = Counter(row[2] for row in non_ext)
    print("\nLanguage breakdown:")
    for lang, count in lang_counts.most_common():
        print(f"  {lang}: {count}")

    # ── Write XLSX ────────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Non-Extractable PDFs"

    headers = ["year", "filename", "title_language", "open_file"]
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    link_font = Font(color="0000FF", underline="single")

    for i, (year, fname, lang, fullpath) in enumerate(non_ext, 2):
        ws.cell(row=i, column=1, value=year)
        ws.cell(row=i, column=2, value=fname)
        ws.cell(row=i, column=3, value=lang)

        # file:// hyperlink
        from urllib.parse import quote
        file_uri = "file://" + quote(fullpath, safe="/:")
        cell = ws.cell(row=i, column=4, value="Open")
        cell.hyperlink = file_uri
        cell.font = link_font

    # Auto-fit column widths (max 60)
    for col_idx in range(1, len(headers) + 1):
        max_len = len(headers[col_idx - 1])
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=2,
                                max_row=ws.max_row, values_only=True):
            val = str(row[0]) if row[0] else ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

    # Freeze top row, autofilter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

    wb.save(OUTPUT)
    print(f"\nSaved: {OUTPUT}")


if __name__ == "__main__":
    main()
