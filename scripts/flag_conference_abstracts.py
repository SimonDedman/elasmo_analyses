#!/usr/bin/env python3
"""
flag_conference_abstracts.py

Triage the no-DOI download queue: a large share (~19%) are conference abstracts /
programme booklets / poster sessions — NOT full papers, so the "full text" often
never existed and they should not be chased as downloads.

Classifies each no-DOI queue paper into:
  - HIGH   : journal is an unambiguous abstract/conference source -> auto-flag
  - MEDIUM : title/journal has proceedings/abstract keywords      -> review
  - none

Writes a review workbook (outputs/no_doi_abstract_triage.xlsx, frozen+bold+autofilter)
and, with --apply, adds a `triage: conference_abstract` field to the HIGH-confidence
entries in docs/papers_data.json (reversible; excludes them from download targets and
the backlog count). Default is DRY (review only).

Usage:
  python3 scripts/flag_conference_abstracts.py            # dry run: writes review xlsx
  python3 scripts/flag_conference_abstracts.py --apply    # + flags HIGH in papers_data.json
"""
import argparse, json, re
from pathlib import Path
BASE = Path(__file__).parent.parent
QUEUE = BASE / "docs/papers_data.json"

# Journals that are unambiguously abstract/conference sources (substring, case-insensitive)
HIGH_JOURNALS = [
    "american elasmobranch society", "conference abstract", "(abstract)",
    "book of abstracts", "résumés", "resumes des communications", "programme booklet",
    "program and poster", "programm and poster", "poster abstracts", "shark international",
    "world congress of herpetology", "encuentro colombiano", "annual meeting",
    "annual symposium", "proceedings of the european elasmobranch",
    "abstracts of the", "meeting abstracts", "oceania chondrichthyan",
]
# Keyword hints (title OR journal) -> MEDIUM (needs review; could be a real proceedings paper)
MED_PAT = re.compile(r"\b(proceedings|symposium|conference|abstract|meeting|colloque|congress|workshop)\b", re.I)

def classify(p):
    j = (p.get("journal_clean") or p.get("journal") or "").lower()
    t = (p.get("title") or "").lower()
    if any(h in j for h in HIGH_JOURNALS):
        return "HIGH"
    if MED_PAT.search(j) or MED_PAT.search(t):
        return "MEDIUM"
    return "none"

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--apply", action="store_true", help="flag HIGH-confidence entries in papers_data.json")
    args = ap.parse_args()
    q = json.load(open(QUEUE))
    nod = [p for p in q if not str(p.get("doi", "")).strip()]
    rows = []
    for p in nod:
        c = classify(p)
        if c != "none":
            rows.append((c, str(p.get("literature_id")), p.get("year"),
                         (p.get("journal_clean") or p.get("journal") or "")[:60],
                         (p.get("title") or "")[:80]))
    rows.sort(key=lambda r: (r[0] != "HIGH", r[3].lower()))
    high = [r for r in rows if r[0] == "HIGH"]
    med = [r for r in rows if r[0] == "MEDIUM"]
    print(f"no-DOI papers: {len(nod):,}")
    print(f"  HIGH-confidence conference abstracts (auto-flag): {len(high):,}")
    print(f"  MEDIUM (review): {len(med):,}")

    # review workbook
    try:
        import openpyxl
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "abstract_triage"
        hdr = ["confidence", "literature_id", "year", "journal", "title", "keep? (y/n)"]
        _illegal = re.compile(r"[\000-\010]|[\013-\014]|[\016-\037]")
        def clean(v): return _illegal.sub("", v) if isinstance(v, str) else v
        ws.append(hdr)
        for r in rows: ws.append([clean(v) for v in r] + [""])
        for c in range(1, len(hdr) + 1): ws.cell(1, c).font = Font(bold=True)
        ws.freeze_panes = "A2"; ws.auto_filter.ref = f"A1:{get_column_letter(len(hdr))}{len(rows)+1}"
        for i, w in enumerate([12, 14, 6, 52, 60, 12], 1): ws.column_dimensions[get_column_letter(i)].width = w
        out = BASE / "outputs/no_doi_abstract_triage.xlsx"; wb.save(out)
        print(f"  review workbook -> {out}")
    except ImportError:
        print("  (openpyxl missing — skipped xlsx)")

    if args.apply:
        high_ids = {r[1] for r in high}
        n = 0
        for p in q:
            if str(p.get("literature_id")) in high_ids and not str(p.get("doi", "")).strip():
                p["triage"] = "conference_abstract"; n += 1
        json.dump(q, open(QUEUE, "w"), indent=1, ensure_ascii=False)
        print(f"  APPLIED: flagged {n:,} entries as triage=conference_abstract in papers_data.json")
        print(f"  (reversible; filter them out of download targets. Backlog now excludes these.)")
    else:
        print("  DRY RUN — re-run with --apply to flag HIGH-confidence entries in papers_data.json")

if __name__ == "__main__":
    main()
