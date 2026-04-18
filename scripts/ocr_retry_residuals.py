#!/usr/bin/env python3
"""Retry OCR on the 68 residuals from the first library OCR pass.

Reads `outputs/ocr_failures_and_residuals.xlsx`. For each row:

  - OCR_SUCCEEDED_NO_USABLE_TEXT: re-run ocrmypdf with --force-ocr
    (rasterises every page, replacing any partial text layer — e.g.
    journal footers — with a full OCR pass).
  - OCR_FAILED, file size > 20 KB: try pikepdf repair (open + re-save
    unlinearised) to fix malformed structure, then retry ocrmypdf.
  - OCR_FAILED, file size <= 20 KB: flag as CORRUPTED_DOWNLOAD — likely
    an HTML error page saved with a .pdf extension.

Writes `outputs/ocr_retry_report.xlsx` with per-file outcomes.
"""

import shutil
import subprocess
import sys
import tempfile
from pathlib import Path
from urllib.parse import quote

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

PROJECT = Path("/media/simon/data/Documents/Si Work/PostDoc Work/EEA/2025/Data Panel")
PDF_ROOT = Path("/media/simon/data/Documents/Si Work/Papers & Books/SharkPapers")
IN_XLSX = PROJECT / "outputs/ocr_failures_and_residuals.xlsx"
OUT_XLSX = PROJECT / "outputs/ocr_retry_report.xlsx"
LOG_FILE = PROJECT / "logs/ocr_retry_log.txt"

TINY_FILE_BYTES = 20_000  # below this → almost certainly a broken download
MIN_ALPHA_WHOLE = 200      # whole-doc alpha count for "extractable"


def whole_doc_alpha(pdf: Path, timeout: int = 120) -> int:
    try:
        r = subprocess.run(
            ["pdftotext", str(pdf), "-"],
            capture_output=True, timeout=timeout,
        )
        return sum(1 for c in r.stdout.decode("utf-8", errors="replace") if c.isalpha())
    except Exception:
        return 0


def force_ocr(pdf: Path) -> tuple[bool, str]:
    """Re-OCR every page with --force-ocr, in-place via temp + atomic rename."""
    tmp = pdf.with_suffix(".forceocr.tmp.pdf")
    try:
        r = subprocess.run(
            ["ocrmypdf", "--force-ocr", "--output-type", "pdf",
             "--language", "eng", "--quiet",
             str(pdf), str(tmp)],
            capture_output=True, timeout=900,
        )
        if r.returncode != 0 or not tmp.exists() or tmp.stat().st_size == 0:
            if tmp.exists():
                tmp.unlink()
            err = r.stderr.decode("utf-8", errors="replace")[:180]
            return False, f"ocrmypdf --force-ocr failed: {err}"
        import os
        os.replace(tmp, pdf)
        return True, "force-ocr OK"
    except subprocess.TimeoutExpired:
        if tmp.exists():
            tmp.unlink()
        return False, "timeout"
    except Exception as e:
        if tmp.exists():
            tmp.unlink()
        return False, f"exception: {e}"


def repair_pdf(pdf: Path) -> tuple[bool, str]:
    """Try to repair a malformed PDF by round-tripping through pikepdf.

    pikepdf's save() will rewrite the file with a clean xref table and
    valid object stream, often fixing structural issues that break
    ocrmypdf's image-identification step.
    """
    import pikepdf
    tmp = pdf.with_suffix(".repair.tmp.pdf")
    try:
        with pikepdf.open(pdf, allow_overwriting_input=False) as src:
            src.save(tmp, linearize=False)
        if not tmp.exists() or tmp.stat().st_size == 0:
            return False, "pikepdf produced empty file"
        import os
        os.replace(tmp, pdf)
        return True, f"pikepdf repair OK ({pdf.stat().st_size:,} bytes)"
    except Exception as e:
        if tmp.exists():
            tmp.unlink()
        return False, f"pikepdf repair failed: {e}"


def ocr_normal(pdf: Path) -> tuple[bool, str]:
    """Normal --skip-text OCR pass (used after repair)."""
    tmp = pdf.with_suffix(".ocr.tmp.pdf")
    try:
        r = subprocess.run(
            ["ocrmypdf", "--skip-text", "--output-type", "pdf",
             "--language", "eng", "--quiet",
             str(pdf), str(tmp)],
            capture_output=True, timeout=900,
        )
        if r.returncode != 0 or not tmp.exists() or tmp.stat().st_size == 0:
            if tmp.exists():
                tmp.unlink()
            # Try --force-ocr as fallback
            r = subprocess.run(
                ["ocrmypdf", "--force-ocr", "--output-type", "pdf",
                 "--language", "eng", "--quiet",
                 str(pdf), str(tmp)],
                capture_output=True, timeout=900,
            )
            if r.returncode != 0 or not tmp.exists() or tmp.stat().st_size == 0:
                if tmp.exists():
                    tmp.unlink()
                err = r.stderr.decode("utf-8", errors="replace")[:180]
                return False, f"ocrmypdf failed after repair: {err}"
        import os
        os.replace(tmp, pdf)
        return True, "ocrmypdf OK after repair"
    except subprocess.TimeoutExpired:
        if tmp.exists():
            tmp.unlink()
        return False, "ocrmypdf timeout"
    except Exception as e:
        if tmp.exists():
            tmp.unlink()
        return False, f"exception: {e}"


def main():
    wb_in = load_workbook(IN_XLSX, data_only=True)
    ws_in = wb_in.active
    rows = []
    for row in ws_in.iter_rows(min_row=2, values_only=True):
        # year, filename, language, status, failure_reason, open_file, notes
        year, fname, lang, status, reason, _link, notes = (row + (None,) * 7)[:7]
        if not year or not fname:
            continue
        p = PDF_ROOT / str(year) / str(fname)
        rows.append({
            "year": str(year), "filename": str(fname), "language": str(lang or ""),
            "prev_status": str(status or ""), "prev_reason": str(reason or ""),
            "notes": str(notes or ""), "path": p,
        })
    wb_in.close()
    print(f"Loaded {len(rows)} residuals from {IN_XLSX.name}")

    results = []
    for i, r in enumerate(rows, 1):
        p = r["path"]
        before_alpha = whole_doc_alpha(p, timeout=60) if p.exists() else -1
        action = ""
        outcome = ""
        ok = False

        if not p.exists():
            action = "missing"
            outcome = "FILE_MISSING"
        elif p.stat().st_size < TINY_FILE_BYTES:
            action = "flag tiny"
            outcome = f"CORRUPTED_DOWNLOAD ({p.stat().st_size} bytes)"
        elif r["prev_status"] == "OCR_SUCCEEDED_NO_USABLE_TEXT":
            action = "--force-ocr"
            ok, msg = force_ocr(p)
            outcome = f"FORCE_OCR_{'OK' if ok else 'FAIL'}: {msg}"
        elif r["prev_status"] == "OCR_FAILED":
            action = "repair+ocr"
            ok_r, msg_r = repair_pdf(p)
            if not ok_r:
                outcome = f"REPAIR_FAIL: {msg_r}"
            else:
                ok_o, msg_o = ocr_normal(p)
                ok = ok_o
                outcome = f"REPAIR_OK → OCR_{'OK' if ok_o else 'FAIL'}: {msg_o}"
        else:
            action = "skip"
            outcome = f"SKIP (unknown status {r['prev_status']})"

        after_alpha = whole_doc_alpha(p, timeout=120) if p.exists() else -1
        flipped = before_alpha < MIN_ALPHA_WHOLE <= after_alpha

        status = ("FLIPPED" if flipped
                  else "STILL_NON_EXTRACTABLE" if after_alpha < MIN_ALPHA_WHOLE
                  else "ALREADY_EXTRACTABLE")

        results.append({**r, "action": action, "outcome": outcome,
                        "before_alpha": before_alpha, "after_alpha": after_alpha,
                        "final_status": status})
        print(f"  [{i}/{len(rows)}] {status}  ({before_alpha}→{after_alpha})  "
              f"{r['filename'][:70]}")

    # Summary
    counts = {}
    for r in results:
        counts[r["final_status"]] = counts.get(r["final_status"], 0) + 1
    print("\nSummary:")
    for k, v in sorted(counts.items()):
        print(f"  {k}: {v}")

    # Write log
    LOG_FILE.parent.mkdir(exist_ok=True)
    with open(LOG_FILE, "w") as f:
        for r in results:
            f.write(f"{r['final_status']}: {r['filename']} [{r['action']}] "
                    f"before={r['before_alpha']} after={r['after_alpha']} — "
                    f"{r['outcome']}\n")

    # Write XLSX
    wb = Workbook()
    ws = wb.active
    ws.title = "OCR Retry Report"
    headers = ["year", "filename", "language", "prev_status", "notes",
               "action", "before_alpha", "after_alpha", "final_status",
               "outcome", "open_file"]
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")

    link_font = Font(color="0000FF", underline="single")
    for i, r in enumerate(results, 2):
        ws.cell(row=i, column=1, value=r["year"])
        ws.cell(row=i, column=2, value=r["filename"])
        ws.cell(row=i, column=3, value=r["language"])
        ws.cell(row=i, column=4, value=r["prev_status"])
        ws.cell(row=i, column=5, value=r["notes"])
        ws.cell(row=i, column=6, value=r["action"])
        ws.cell(row=i, column=7, value=r["before_alpha"])
        ws.cell(row=i, column=8, value=r["after_alpha"])
        ws.cell(row=i, column=9, value=r["final_status"])
        ws.cell(row=i, column=10, value=r["outcome"])
        uri = "file://" + quote(str(r["path"]), safe="/:")
        c = ws.cell(row=i, column=11, value="Open")
        c.hyperlink = uri
        c.font = link_font

    for col_idx, h in enumerate(headers, 1):
        max_len = len(h)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=2,
                                max_row=ws.max_row, values_only=True):
            val = str(row[0]) if row[0] is not None else ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 70)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    wb.save(OUT_XLSX)
    print(f"\nReport: {OUT_XLSX}")
    print(f"Log:    {LOG_FILE}")


if __name__ == "__main__":
    main()
