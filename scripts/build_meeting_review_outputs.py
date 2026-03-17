#!/usr/bin/env python3
"""
build_meeting_review_outputs.py

Generates three outputs for the March 2026 team meeting review:

1. Excel workbook: schema_extraction_evidence filtered by team coauthors,
   with one tab per coauthor, papers grouped by paper ID within each tab.

2. Missing-papers report: year column plot + journal breakdown for papers
   remaining to download (~15,000 after Jürgen ingestion).

3. Per-journal HTML pages for collaborative paper acquisition, with clickable
   links that coauthors can open from GitHub to work through downloads.

Usage:
    python scripts/build_meeting_review_outputs.py

Author: Simon Dedman
Date: 2026-03-15
"""

import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime
import html
import re
import json
from openpyxl.styles import PatternFill

# ============================================================================
# CONFIGURATION
# ============================================================================

BASE_DIR = Path(__file__).parent.parent
LIT_REVIEW = BASE_DIR / "outputs" / "literature_review.parquet"
LIT_ENRICHED = BASE_DIR / "outputs" / "literature_review_enriched.parquet"
EVIDENCE_CSV = BASE_DIR / "outputs" / "schema_extraction_evidence.csv"
MISSING_BY_JOURNAL = BASE_DIR / "outputs" / "missing_papers_by_journal.csv"
MISSING_ANALYSIS = BASE_DIR / "outputs" / "missing_papers_analysis.csv"
OUTPUT_DIR = BASE_DIR / "outputs" / "meeting_review"
HTML_DIR = OUTPUT_DIR / "journal_download_pages"

# Team members: display name -> search patterns for author field
TEAM_MEMBERS = {
    "Simon Dedman": [r"Dedman"],
    "David Ruiz-Garcia": [r"Ruiz-Garc"],
    "Guuske Tiktak": [r"Tiktak"],
    "Elena Fernandez-Corredor": [r"Corredor"],
    "David Shiffman": [r"Shiffman"],
}


def find_coauthor_papers(lr, name, patterns):
    """Find papers where any pattern matches the authors field."""
    authors_col = lr["authors"].fillna("")
    mask = pd.Series(False, index=lr.index)
    for pat in patterns:
        mask = mask | authors_col.str.contains(pat, case=False, na=False, regex=True)
    return lr[mask].copy()


# ============================================================================
# 1. FILTERED EXCEL: Evidence by coauthor
# ============================================================================

def build_coauthor_evidence_excel():
    """Create Excel workbook with evidence filtered by team coauthor papers."""
    print("\n=== Building coauthor evidence Excel ===")

    lr = pd.read_parquet(LIT_REVIEW, columns=["literature_id", "authors", "title", "year", "journal"])
    evidence = pd.read_csv(EVIDENCE_CSV)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    outfile = OUTPUT_DIR / "schema_extraction_evidence_by_coauthor.xlsx"

    # Normalise literature_id to numeric in both datasets
    lr["literature_id"] = pd.to_numeric(lr["literature_id"], errors="coerce")
    evidence["literature_id"] = pd.to_numeric(evidence["literature_id"], errors="coerce")
    evidence_ids = set(evidence["literature_id"].dropna().astype(int))

    with pd.ExcelWriter(outfile, engine="openpyxl") as writer:
        # Summary tab
        summary_rows = []
        all_ids = set()

        for name, patterns in TEAM_MEMBERS.items():
            papers = find_coauthor_papers(lr, name, patterns)
            paper_ids = set(papers["literature_id"].dropna().astype(int))
            summary_rows.append({
                "Coauthor": name,
                "Papers in DB": len(papers),
                "Paper IDs with evidence": len(paper_ids & evidence_ids),
            })
            all_ids.update(paper_ids)

        summary_rows.append({
            "Coauthor": "ALL TEAM (union)",
            "Papers in DB": len(all_ids),
            "Paper IDs with evidence": len(all_ids & evidence_ids),
        })
        summary_df = pd.DataFrame(summary_rows)
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        print(f"  Summary: {len(summary_rows)} rows")
        sheets_written = True  # Summary always writes

        # Per-coauthor tabs
        for name, patterns in TEAM_MEMBERS.items():
            papers = find_coauthor_papers(lr, name, patterns)
            paper_ids = set(papers["literature_id"].dropna().astype(int))

            if not paper_ids:
                print(f"  {name}: 0 papers, skipping tab")
                continue

            # Filter evidence to this coauthor's papers (match on numeric IDs)
            ev_filtered = evidence[
                evidence["literature_id"].isin(paper_ids) |
                evidence["literature_id"].isin([float(x) for x in paper_ids])
            ].copy()

            if ev_filtered.empty:
                print(f"  {name}: {len(paper_ids)} papers but 0 evidence rows, skipping")
                continue

            # Merge paper metadata
            ev_merged = ev_filtered.merge(
                papers[["literature_id", "authors", "year", "journal"]].drop_duplicates("literature_id"),
                on="literature_id",
                how="left",
                suffixes=("", "_paper"),
            )

            # Clean illegal characters for Excel (control chars except \t \n \r)
            for col in ev_merged.select_dtypes(include=["object", "string"]).columns:
                ev_merged[col] = ev_merged[col].apply(
                    lambda x: re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', ' ', str(x)) if pd.notna(x) else x
                )

            # Sort by paper ID then column name
            ev_merged = ev_merged.sort_values(["literature_id", "column"])

            # Reorder columns for readability
            col_order = [
                "literature_id", "year", "authors", "journal", "title",
                "column", "binary", "total_freq", "term_count", "threshold",
                "matched_terms", "matched_anchors", "context",
            ]
            col_order = [c for c in col_order if c in ev_merged.columns]
            ev_merged = ev_merged[col_order]

            # Truncate sheet name to 31 chars (Excel limit)
            sheet_name = name[:31]
            ev_merged.to_excel(writer, sheet_name=sheet_name, index=False)

            # Apply alternating light-grey shading per paper (not per row)
            ws = writer.sheets[sheet_name]
            grey_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            shade = False
            prev_paper_id = None
            for row_idx in range(2, ws.max_row + 1):  # skip header row
                paper_id = ws.cell(row=row_idx, column=1).value  # literature_id is col 1
                if paper_id != prev_paper_id:
                    shade = not shade
                    prev_paper_id = paper_id
                if shade:
                    for col_idx in range(1, ws.max_column + 1):
                        ws.cell(row=row_idx, column=col_idx).fill = grey_fill

            print(f"  {name}: {len(paper_ids)} papers, {len(ev_merged)} evidence rows")

    print(f"  Saved: {outfile}")
    return outfile


# ============================================================================
# 2. MISSING PAPERS REPORT (Year plot + journal breakdown)
# ============================================================================

def build_missing_papers_report():
    """Build HTML report of papers remaining to download, by year and journal."""
    print("\n=== Building missing papers report ===")

    # Load the full literature review to identify papers without PDFs
    lr = pd.read_parquet(LIT_REVIEW, columns=[
        "literature_id", "title", "authors", "year", "journal", "doi", "pdf_url"
    ])

    # Also load missing_papers_by_journal for the Jürgen/new papers
    try:
        missing_journal = pd.read_csv(MISSING_BY_JOURNAL)
        print(f"  missing_papers_by_journal.csv: {len(missing_journal)} rows")
        has_missing_journal = True
    except Exception as e:
        print(f"  Could not load missing_papers_by_journal.csv: {e}")
        has_missing_journal = False
        missing_journal = pd.DataFrame()

    # Count papers by download status
    # Papers without literature_id in evidence = likely no PDF
    # For now, use the missing_papers_by_journal as the "remaining" set
    # plus papers in the main DB that we know are missing

    # Build year distribution from the missing journal data
    if has_missing_journal and "year" in missing_journal.columns:
        missing_journal["year"] = pd.to_numeric(missing_journal["year"], errors="coerce")
        by_year = missing_journal.groupby("year").size().reset_index(name="count")
        by_year = by_year.dropna(subset=["year"])
        by_year["year"] = by_year["year"].astype(int)
    else:
        # Fallback: use all papers in main DB, show full distribution
        lr["year"] = pd.to_numeric(lr["year"], errors="coerce")
        by_year = lr.groupby("year").size().reset_index(name="count")
        by_year = by_year.dropna(subset=["year"])
        by_year["year"] = by_year["year"].astype(int)

    # Journal distribution from missing papers
    if has_missing_journal and "journal" in missing_journal.columns:
        by_journal = (
            missing_journal["journal"]
            .fillna("Unknown")
            .value_counts()
            .reset_index()
        )
        by_journal.columns = ["journal", "count"]
    else:
        by_journal = pd.DataFrame(columns=["journal", "count"])

    # Also build overall DB stats
    total_papers = len(lr)
    total_missing = len(missing_journal) if has_missing_journal else 0
    papers_with_doi = lr["doi"].notna().sum()

    # Generate HTML report
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    outfile = OUTPUT_DIR / "missing_papers_report.html"

    # Prepare chart data
    year_labels = by_year["year"].tolist()
    year_counts = by_year["count"].tolist()

    # Top 50 journals for chart, rest in table
    top_journals = by_journal.head(50)
    journal_labels = top_journals["journal"].tolist()
    journal_counts = top_journals["count"].tolist()

    report_html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Missing Papers Report - EEA 2025 Data Panel</title>
    <script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0"></script>
    <style>
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            max-width: 1400px;
            margin: 0 auto;
            padding: 20px;
            background: #f5f5f5;
            color: #333;
        }}
        h1 {{ color: #2c3e50; border-bottom: 3px solid #3498db; padding-bottom: 10px; }}
        h2 {{ color: #34495e; margin-top: 30px; }}
        .stats-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 15px;
            margin: 20px 0;
        }}
        .stat-card {{
            background: white;
            padding: 20px;
            border-radius: 8px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
            text-align: center;
        }}
        .stat-card .number {{
            font-size: 2em;
            font-weight: bold;
            color: #3498db;
        }}
        .stat-card .label {{
            color: #666;
            margin-top: 5px;
        }}
        .chart-container {{
            background: white;
            padding: 20px;
            border-radius: 8px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
            margin: 20px 0;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            background: white;
            border-radius: 8px;
            overflow: hidden;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }}
        th, td {{
            padding: 10px 15px;
            text-align: left;
            border-bottom: 1px solid #eee;
        }}
        th {{
            background: #3498db;
            color: white;
            position: sticky;
            top: 0;
        }}
        tr:hover {{ background: #f0f8ff; }}
        .filter-input {{
            width: 100%;
            padding: 10px;
            margin: 10px 0;
            border: 2px solid #ddd;
            border-radius: 5px;
            font-size: 14px;
        }}
        .timestamp {{ color: #999; font-size: 0.9em; }}
    </style>
</head>
<body>
    <h1>Missing Papers Report</h1>
    <p class="timestamp">Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}</p>

    <div class="stats-grid">
        <div class="stat-card">
            <div class="number">{total_papers:,}</div>
            <div class="label">Total Papers in DB</div>
        </div>
        <div class="stat-card">
            <div class="number">{total_missing:,}</div>
            <div class="label">Papers in Missing List</div>
        </div>
        <div class="stat-card">
            <div class="number">{papers_with_doi:,}</div>
            <div class="label">Papers with DOI</div>
        </div>
        <div class="stat-card">
            <div class="number">{len(by_journal):,}</div>
            <div class="label">Unique Journals (missing)</div>
        </div>
    </div>

    <h2>Missing Papers by Year</h2>
    <div class="chart-container">
        <canvas id="yearChart" height="100"></canvas>
    </div>

    <h2>Missing Papers by Journal (Top 50)</h2>
    <div class="chart-container" style="height: 800px;">
        <canvas id="journalChart"></canvas>
    </div>

    <h2>Full Journal Breakdown ({len(by_journal)} journals)</h2>
    <input type="text" class="filter-input" id="journalFilter"
           placeholder="Filter journals..." oninput="filterTable()">
    <table id="journalTable">
        <thead>
            <tr><th>#</th><th>Journal</th><th>Papers Missing</th><th>% of Total</th></tr>
        </thead>
        <tbody>
"""
    for i, row in by_journal.iterrows():
        pct = (row["count"] / total_missing * 100) if total_missing > 0 else 0
        report_html += f'            <tr><td>{i+1}</td><td>{html.escape(str(row["journal"]))}</td><td>{row["count"]}</td><td>{pct:.1f}%</td></tr>\n'

    report_html += f"""        </tbody>
    </table>

    <script>
        // Year chart
        new Chart(document.getElementById('yearChart'), {{
            type: 'bar',
            data: {{
                labels: {json.dumps(year_labels)},
                datasets: [{{
                    label: 'Missing Papers',
                    data: {json.dumps(year_counts)},
                    backgroundColor: '#3498db',
                    borderColor: '#2980b9',
                    borderWidth: 1
                }}]
            }},
            options: {{
                responsive: true,
                plugins: {{
                    title: {{ display: true, text: 'Missing Papers by Publication Year', font: {{ size: 16 }} }}
                }},
                scales: {{
                    y: {{ beginAtZero: true, title: {{ display: true, text: 'Number of Papers' }} }},
                    x: {{ title: {{ display: true, text: 'Year' }} }}
                }}
            }}
        }});

        // Journal chart (horizontal bar)
        new Chart(document.getElementById('journalChart'), {{
            type: 'bar',
            data: {{
                labels: {json.dumps(journal_labels[:50])},
                datasets: [{{
                    label: 'Missing Papers',
                    data: {json.dumps(journal_counts[:50])},
                    backgroundColor: '#e74c3c',
                    borderColor: '#c0392b',
                    borderWidth: 1
                }}]
            }},
            options: {{
                indexAxis: 'y',
                responsive: true,
                maintainAspectRatio: false,
                plugins: {{
                    title: {{ display: true, text: 'Top 50 Journals with Missing Papers', font: {{ size: 16 }} }},
                    legend: {{ display: false }}
                }},
                scales: {{
                    x: {{ beginAtZero: true, title: {{ display: true, text: 'Number of Papers' }} }}
                }}
            }}
        }});

        // Table filter
        function filterTable() {{
            const filter = document.getElementById('journalFilter').value.toLowerCase();
            const rows = document.querySelectorAll('#journalTable tbody tr');
            rows.forEach(row => {{
                const journal = row.cells[1].textContent.toLowerCase();
                row.style.display = journal.includes(filter) ? '' : 'none';
            }});
        }}
    </script>
</body>
</html>"""

    with open(outfile, "w", encoding="utf-8") as f:
        f.write(report_html)

    print(f"  Saved: {outfile}")
    print(f"  Year range: {min(year_labels) if year_labels else 'N/A'} - {max(year_labels) if year_labels else 'N/A'}")
    print(f"  Top 5 journals: {by_journal.head(5).to_string()}")

    # Also save CSV for reference
    by_journal.to_csv(OUTPUT_DIR / "missing_papers_journal_summary.csv", index=False)
    by_year.to_csv(OUTPUT_DIR / "missing_papers_year_summary.csv", index=False)

    return outfile, by_journal


# ============================================================================
# 3. PER-JOURNAL HTML PAGES (for collaborative download)
# ============================================================================

def build_journal_download_pages(by_journal_df=None):
    """Build HTML pages per journal for collaborative paper acquisition."""
    print("\n=== Building per-journal HTML download pages ===")

    # Load missing papers data
    try:
        missing = pd.read_csv(MISSING_BY_JOURNAL)
    except Exception as e:
        print(f"  Error loading missing papers: {e}")
        return

    HTML_DIR.mkdir(parents=True, exist_ok=True)

    # Clean journal names
    missing["journal_clean"] = missing["journal"].fillna("Unknown")

    # Group by journal
    journals = missing.groupby("journal_clean")
    journal_count = len(journals)
    print(f"  Generating pages for {journal_count} journals...")

    # Build index page
    index_rows = []

    for journal_name, group in journals:
        safe_name = re.sub(r'[^\w\-]', '_', str(journal_name))[:80]
        filename = f"download_{safe_name}.html"
        n_papers = len(group)

        index_rows.append({
            "journal": journal_name,
            "count": n_papers,
            "filename": filename,
        })

        # Build individual journal page
        page_html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Download Papers - {html.escape(str(journal_name))}</title>
    <style>
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            max-width: 1200px; margin: 0 auto; padding: 20px; background: #f5f5f5;
        }}
        h1 {{ color: #2c3e50; border-bottom: 3px solid #e74c3c; padding-bottom: 10px; font-size: 1.4em; }}
        .stats {{ background: #fff; padding: 15px; border-radius: 8px; margin: 15px 0;
                  box-shadow: 0 2px 4px rgba(0,0,0,0.1); }}
        .paper {{
            background: white; padding: 15px; margin: 8px 0; border-radius: 8px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1); border-left: 4px solid #3498db;
            transition: border-color 0.3s;
        }}
        .paper.downloaded {{ border-left-color: #27ae60; opacity: 0.6; }}
        .paper .title {{ font-weight: bold; color: #2c3e50; margin-bottom: 5px; }}
        .paper .meta {{ color: #666; font-size: 0.9em; }}
        .paper .doi {{ margin-top: 5px; }}
        .paper .doi a {{ color: #3498db; text-decoration: none; }}
        .paper .doi a:hover {{ text-decoration: underline; }}
        .btn {{
            display: inline-block; padding: 5px 12px; margin: 3px;
            border: none; border-radius: 4px; cursor: pointer; font-size: 0.85em;
        }}
        .btn-done {{ background: #27ae60; color: white; }}
        .btn-skip {{ background: #f39c12; color: white; }}
        .btn-reset {{ background: #e74c3c; color: white; }}
        .progress {{ background: #ecf0f1; border-radius: 10px; height: 20px; margin: 10px 0; overflow: hidden; }}
        .progress-bar {{ background: #27ae60; height: 100%; transition: width 0.3s; border-radius: 10px; }}
        .back-link {{ margin: 10px 0; }}
        .back-link a {{ color: #3498db; text-decoration: none; }}
    </style>
</head>
<body>
    <div class="back-link"><a href="index.html">&larr; Back to Journal Index</a></div>
    <h1>{html.escape(str(journal_name))}</h1>
    <div class="stats">
        <strong>{n_papers}</strong> papers to acquire |
        <span id="doneCount">0</span> downloaded |
        <span id="skipCount">0</span> skipped
        <div class="progress"><div class="progress-bar" id="progressBar" style="width: 0%"></div></div>
        <button class="btn btn-reset" onclick="resetAll()">Reset All</button>
    </div>
"""
        for idx, (_, paper) in enumerate(group.iterrows()):
            authors = html.escape(str(paper.get("authors", "Unknown"))[:200])
            title_text = html.escape(str(paper.get("citation", paper.get("full_text", "No title")))[:300])
            year = paper.get("year", "")
            doi = paper.get("doi", "")
            doi_link = ""
            if pd.notna(doi) and str(doi).strip():
                doi_clean = str(doi).strip()
                if not doi_clean.startswith("http"):
                    doi_clean = f"https://doi.org/{doi_clean}"
                doi_link = f'<div class="doi"><a href="{html.escape(doi_clean)}" target="_blank">{html.escape(doi_clean)}</a></div>'

            # Search link for papers without DOI
            search_text = str(paper.get("citation", paper.get("full_text", "")))[:100]
            scholar_link = f'https://scholar.google.com/scholar?q={html.escape(search_text.replace(" ", "+"))}'

            page_html += f"""
    <div class="paper" id="paper-{idx}" data-status="pending">
        <div class="title">{title_text}</div>
        <div class="meta">{authors} ({year})</div>
        {doi_link}
        <div style="margin-top: 8px;">
            <a href="{scholar_link}" target="_blank" class="btn" style="background:#9b59b6;color:white;">Google Scholar</a>
            <button class="btn btn-done" onclick="markDone({idx})">Downloaded</button>
            <button class="btn btn-skip" onclick="markSkip({idx})">Skip/Unavailable</button>
        </div>
    </div>
"""

        page_html += f"""
    <script>
        const STORAGE_KEY = 'downloads_{safe_name}';
        const TOTAL = {n_papers};

        function getState() {{
            return JSON.parse(localStorage.getItem(STORAGE_KEY) || '{{}}');
        }}

        function saveState(state) {{
            localStorage.setItem(STORAGE_KEY, JSON.stringify(state));
            updateUI(state);
        }}

        function markDone(idx) {{
            const state = getState();
            state[idx] = 'done';
            saveState(state);
        }}

        function markSkip(idx) {{
            const state = getState();
            state[idx] = 'skip';
            saveState(state);
        }}

        function resetAll() {{
            if (confirm('Reset all progress for this journal?')) {{
                localStorage.removeItem(STORAGE_KEY);
                updateUI({{}});
            }}
        }}

        function updateUI(state) {{
            let done = 0, skip = 0;
            for (const [idx, status] of Object.entries(state)) {{
                const el = document.getElementById('paper-' + idx);
                if (el) {{
                    el.className = 'paper' + (status === 'done' ? ' downloaded' : '');
                    el.dataset.status = status;
                }}
                if (status === 'done') done++;
                if (status === 'skip') skip++;
            }}
            document.getElementById('doneCount').textContent = done;
            document.getElementById('skipCount').textContent = skip;
            const pct = ((done + skip) / TOTAL * 100).toFixed(1);
            document.getElementById('progressBar').style.width = pct + '%';
        }}

        // Init
        updateUI(getState());
    </script>
</body>
</html>"""

        with open(HTML_DIR / filename, "w", encoding="utf-8") as f:
            f.write(page_html)

    # Build index page
    index_rows.sort(key=lambda x: -x["count"])
    index_html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Journal Download Index - EEA 2025 Data Panel</title>
    <style>
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            max-width: 1200px; margin: 0 auto; padding: 20px; background: #f5f5f5;
        }}
        h1 {{ color: #2c3e50; border-bottom: 3px solid #3498db; padding-bottom: 10px; }}
        .stats {{ background: #fff; padding: 15px; border-radius: 8px; margin: 15px 0;
                  box-shadow: 0 2px 4px rgba(0,0,0,0.1); }}
        table {{
            width: 100%; border-collapse: collapse; background: white;
            border-radius: 8px; overflow: hidden; box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }}
        th, td {{ padding: 10px 15px; text-align: left; border-bottom: 1px solid #eee; }}
        th {{ background: #3498db; color: white; position: sticky; top: 0; cursor: pointer; }}
        th:hover {{ background: #2980b9; }}
        tr:hover {{ background: #f0f8ff; }}
        a {{ color: #3498db; text-decoration: none; }}
        a:hover {{ text-decoration: underline; }}
        .filter-input {{
            width: 100%; padding: 10px; margin: 10px 0;
            border: 2px solid #ddd; border-radius: 5px; font-size: 14px;
        }}
        .assign-btn {{
            padding: 3px 8px; border: 1px solid #ddd; border-radius: 3px;
            cursor: pointer; font-size: 0.8em; margin: 0 2px;
        }}
        .assign-btn.active {{ background: #3498db; color: white; border-color: #3498db; }}
        .timestamp {{ color: #999; font-size: 0.9em; }}
    </style>
</head>
<body>
    <h1>Journal Download Index</h1>
    <p class="timestamp">Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')} |
       {len(missing):,} papers across {journal_count} journals</p>

    <div class="stats">
        <strong>Instructions:</strong> Click a journal to open its download page.
        Each page has Google Scholar links + DOI links. Mark papers as downloaded or skipped.
        Progress is saved in your browser's localStorage.
        Assign journals to team members using the buttons below.
    </div>

    <input type="text" class="filter-input" id="journalFilter"
           placeholder="Filter journals..." oninput="filterTable()">

    <table id="journalTable">
        <thead>
            <tr>
                <th onclick="sortTable(0)">#</th>
                <th onclick="sortTable(1)">Journal</th>
                <th onclick="sortTable(2)">Papers</th>
                <th>Assigned To</th>
            </tr>
        </thead>
        <tbody>
"""
    team_initials = ["SD", "DRG", "GT", "EFC", "DS"]
    team_names_short = ["Simon", "David RG", "Guuske", "Elena", "Shiffman"]

    for i, row in enumerate(index_rows):
        assign_buttons = " ".join(
            f'<button class="assign-btn" data-journal="{i}" data-person="{init}" onclick="toggleAssign(this)">{init}</button>'
            for init, nm in zip(team_initials, team_names_short)
        )
        index_html += f'            <tr><td>{i+1}</td><td><a href="{row["filename"]}">{html.escape(str(row["journal"]))}</a></td><td>{row["count"]}</td><td>{assign_buttons}</td></tr>\n'

    index_html += f"""        </tbody>
    </table>

    <script>
        const ASSIGN_KEY = 'journal_assignments';

        function getAssignments() {{
            return JSON.parse(localStorage.getItem(ASSIGN_KEY) || '{{}}');
        }}

        function toggleAssign(btn) {{
            const journal = btn.dataset.journal;
            const person = btn.dataset.person;
            const assignments = getAssignments();
            const key = journal + '_' + person;
            if (assignments[key]) {{
                delete assignments[key];
                btn.classList.remove('active');
            }} else {{
                assignments[key] = true;
                btn.classList.add('active');
            }}
            localStorage.setItem(ASSIGN_KEY, JSON.stringify(assignments));
        }}

        function filterTable() {{
            const filter = document.getElementById('journalFilter').value.toLowerCase();
            const rows = document.querySelectorAll('#journalTable tbody tr');
            rows.forEach(row => {{
                const journal = row.cells[1].textContent.toLowerCase();
                row.style.display = journal.includes(filter) ? '' : 'none';
            }});
        }}

        function sortTable(col) {{
            const table = document.getElementById('journalTable');
            const tbody = table.querySelector('tbody');
            const rows = Array.from(tbody.querySelectorAll('tr'));
            const isNum = col === 0 || col === 2;
            rows.sort((a, b) => {{
                let va = a.cells[col].textContent;
                let vb = b.cells[col].textContent;
                if (isNum) return parseInt(va) - parseInt(vb);
                return va.localeCompare(vb);
            }});
            rows.forEach(r => tbody.appendChild(r));
        }}

        // Restore assignments
        (function() {{
            const assignments = getAssignments();
            document.querySelectorAll('.assign-btn').forEach(btn => {{
                const key = btn.dataset.journal + '_' + btn.dataset.person;
                if (assignments[key]) btn.classList.add('active');
            }});
        }})();
    </script>
</body>
</html>"""

    with open(HTML_DIR / "index.html", "w", encoding="utf-8") as f:
        f.write(index_html)

    print(f"  Saved: {HTML_DIR / 'index.html'}")
    print(f"  Generated {journal_count} journal pages in {HTML_DIR}")
    return HTML_DIR


# ============================================================================
# MAIN
# ============================================================================

if __name__ == "__main__":
    print("=" * 60)
    print("EEA 2025 Data Panel - Meeting Review Outputs")
    print("=" * 60)

    # 1. Coauthor evidence Excel
    excel_file = build_coauthor_evidence_excel()

    # 2. Missing papers report
    report_file, by_journal = build_missing_papers_report()

    # 3. Per-journal HTML pages
    html_dir = build_journal_download_pages(by_journal)

    print("\n" + "=" * 60)
    print("DONE! Outputs saved to:")
    print(f"  1. {excel_file}")
    print(f"  2. {report_file}")
    print(f"  3. {html_dir}")
    print("=" * 60)
