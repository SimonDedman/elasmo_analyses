#!/usr/bin/env python3
"""Generate per-coauthor evidence XLSX from schema extraction results.

Creates an Excel workbook with:
- Summary tab (coauthor name, paper count, evidence row count, glossary)
- One tab per team member with their papers' evidence rows
- Acronym Matches tab for validating case-sensitive acronym detection

Usage:
    python scripts/generate_coauthor_evidence_xlsx.py
"""

from __future__ import annotations

import re
import sqlite3
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
BASE = Path("/media/simon/data/Documents/Si Work/PostDoc Work/EEA/2025/Data Panel")
EVIDENCE_CSV = BASE / "outputs" / "schema_extraction_evidence.csv"
AUTHORS_CSV = BASE / "outputs" / "openalex_paper_authors.csv"
ALTMETRIC_CSV = BASE / "outputs" / "altmetric_scores.csv"
UNPAYWALL_CSV = BASE / "outputs" / "unpaywall_oa_by_doi.csv"
PARQUET = BASE / "outputs" / "literature_review_enriched.parquet"
DB = BASE / "database" / "technique_taxonomy.db"
OUTPUT = BASE / "outputs" / "meeting_review" / "schema_extraction_evidence_by_coauthor.xlsx"

# ---------------------------------------------------------------------------
# Team members: display name -> list of surname variants (lowercase, normalised)
# ---------------------------------------------------------------------------
TEAM: dict[str, list[str]] = {
    "Simon Dedman": ["dedman"],
    "David Ruiz-Garcia": ["ruiz-garcia", "ruiz-garcía"],
    "Guuske Tiktak": ["tiktak"],
    "Elena Fernandez-Corredor": ["fernandez-corredor", "fernández-corredor"],
    "David Shiffman": ["shiffman"],
    "Chris Mull": ["mull"],
    "Alex McInturf": ["mcinturf"],
    "Sophia Pelletier": ["pelletier"],
    "Deven Guerrero": ["guerrero"],
    "Ryan McMullen": ["mcmullen"],
}

# Acronyms to check in matched_terms
ACRONYMS = [
    "RAMP", "CPUE", "MPA", "IUU", "OMZ", "PCB", "PFAS", "ALAN", "EMF",
    "BRD", "TED", "EPM", "PRM", "ALDFG", "PAL", "DOA", "AVM", "SSB",
    "WTP", "SPR", "MSY", "CITES", "IUCN", "SNP", "BRUVs", "eDNA",
    "UAV", "ROV",
]

# Evidence tab columns (order matters)
EVIDENCE_COLS = [
    "literature_id", "year", "authors", "journal", "title",
    "column", "matched_terms", "section", "threshold", "total_freq",
    "raw_freq", "term_count", "matched_anchors", "binary", "context",
]
REVIEWER_COLS = ["Notes"]

# Glossary rows for Summary tab
GLOSSARY_FIELDS = [
    ("literature_id", "Unique paper identifier (Shark-References database)"),
    ("year", "Publication year"),
    ("authors", "Author list"),
    ("journal", "Journal name"),
    ("title", "Paper title"),
    ("column", "Schema column name being evaluated (e.g. eco_marine, imp_abundance)"),
    ("matched_terms", 'Terms that matched with frequencies, e.g. "marine(5); ocean(18)"'),
    ("section", "Paper section where primary match occurred (e.g. METHODS, RESULTS)"),
    ("threshold", "Minimum total_freq required for binary=1"),
    ("total_freq", "Frequency count after section weighting and proximity checks"),
    ("raw_freq", "Raw frequency count before section weighting (for comparison)"),
    ("term_count", "Number of distinct terms that matched"),
    ("matched_anchors", "Anchor/context terms that fired (impact columns only; empty = no anchors required or none fired)"),
    ("binary", "Final classification: 1 = column applies to this paper, 0 = does not"),
    ("context", "~160-character text snippet around first match"),
    ("Notes", "Reviewer notes and comments"),
]

TECHNIQUE_CODES = [
    ("KFT", "Keyword frequency threshold — total mentions >= N"),
    ("ANC", "Context anchors — anchor term must co-occur"),
    ("KPC", "Keyword proximity check — elasmobranch term within ±1 sentence"),
    ("SW", "Section-weighted scoring — section-specific weight multipliers"),
    ("AC", "Case-sensitive acronym matching"),
    ("SS", "Section stripping — references/acknowledgements removed"),
    ("KM", "Key match — species binomial or technique name lookup"),
    ("RE", "Regex extraction — structured value extracted from text"),
    ("SA", "Sentiment analysis — directional impact classification"),
    ("DRV", "Derived — value inferred from other fields"),
    ("API", "API/GEO lookup — value from external data source"),
]

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
DASH_CHARS = "\u2010\u2011\u2012\u2013\u2014\u2212"
_dash_table = str.maketrans({ch: "-" for ch in DASH_CHARS})

# Regex to strip illegal XML characters that openpyxl rejects
_ILLEGAL_XML_RE = re.compile(
    r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]"
)


def normalise_name(s: str | float) -> str:
    """Lowercase and replace all Unicode dashes with ASCII hyphen."""
    if pd.isna(s):
        return ""
    return str(s).lower().translate(_dash_table)


def lit_id_to_int(val: str | float) -> int:
    """Convert literature_id like '31532.0' to int 31532."""
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return -1


def style_header(ws: any) -> None:
    """Apply bold + fill to header row and freeze panes."""
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    ws.freeze_panes = "A2"


def auto_width(ws: any, max_width: int = 60) -> None:
    """Set column widths based on header length (capped)."""
    for col in ws.columns:
        header_val = str(col[0].value) if col[0].value else ""
        width = min(len(header_val) + 4, max_width)
        ws.column_dimensions[col[0].column_letter].width = width


def sanitise_value(val: any) -> any:
    """Remove illegal XML characters from string values."""
    if isinstance(val, str):
        return _ILLEGAL_XML_RE.sub("", val)
    return val


def write_dataframe(ws: any, df: pd.DataFrame) -> None:
    """Write a DataFrame to a worksheet including headers."""
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
        for c_idx, val in enumerate(row, start=1):
            val = sanitise_value(val)
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if r_idx > 1 and isinstance(val, str) and len(val) > 80:
                cell.alignment = Alignment(wrap_text=True)


def make_extra_row(
    lit_id: int,
    year: any,
    authors: str,
    journal: str,
    title: str,
    column: str,
    matched_terms: str,
    section: str,
    threshold: any,
    total_freq: any,
    raw_freq: any,
    term_count: any,
    matched_anchors: str,
    binary: int,
    context: str,
) -> dict:
    """Build a dict matching EVIDENCE_COLS order for extra evidence rows."""
    return {
        "literature_id": lit_id,
        "year": year,
        "authors": authors,
        "journal": journal,
        "title": title,
        "column": column,
        "matched_terms": matched_terms,
        "section": section,
        "threshold": threshold,
        "total_freq": total_freq,
        "raw_freq": raw_freq,
        "term_count": term_count,
        "matched_anchors": matched_anchors,
        "binary": binary,
        "context": context,
    }


def sp_col_to_binomial(col: str) -> str:
    """Convert sp_genus_species column name to 'Genus species' binomial."""
    # e.g. sp_carcharodon_carcharias -> Carcharodon carcharias
    parts = col[3:].split("_")  # strip "sp_"
    if len(parts) >= 2:
        return parts[0].capitalize() + " " + " ".join(parts[1:])
    return col[3:].replace("_", " ").capitalize()


def a_col_to_label(col: str) -> str:
    """Convert a_technique_name column name to human-readable label."""
    return col[2:].replace("_", " ").capitalize()


def is_truthy(val: any) -> bool:
    """Return True if val represents a positive/present value."""
    if pd.isna(val):
        return False
    if isinstance(val, bool):
        return val
    if isinstance(val, (int, float)):
        return val != 0
    s = str(val).strip().lower()
    return s not in ("", "false", "0", "none", "nan", "not stated")


# ---------------------------------------------------------------------------
# Build extra evidence rows for a set of paper IDs
# ---------------------------------------------------------------------------
def build_extra_evidence(
    paper_ids: set[int],
    meta_lookup: dict[int, dict],
    parquet_extra: pd.DataFrame,
    openalex_authors: pd.DataFrame,
    altmetric: pd.DataFrame,
    unpaywall: pd.DataFrame,
    geo_df: pd.DataFrame,
    sp_cols: list[str],
    a_cols: list[str],
) -> list[dict]:
    """Build extra evidence rows for techniques not in schema extraction CSV."""
    rows: list[dict] = []

    # Filter parquet_extra to this coauthor's papers
    sub = parquet_extra[parquet_extra["lit_id_int"].isin(paper_ids)].copy()

    for _, prow in sub.iterrows():
        lid = int(prow["lit_id_int"])
        m = meta_lookup.get(lid, {})
        year = m.get("year", "")
        authors = m.get("authors", "")
        journal = m.get("journal", "")
        title = m.get("title", "")

        def base_row(col, matched, section, binary=1, threshold="", total_freq="", raw_freq="", term_count="", context=""):
            return make_extra_row(
                lit_id=lid, year=year, authors=authors, journal=journal, title=title,
                column=col, matched_terms=str(matched), section=section,
                threshold=threshold, total_freq=total_freq, raw_freq=raw_freq,
                term_count=term_count, matched_anchors="", binary=binary, context=context,
            )

        # ---- 5a: Species (KM) ----
        for sc in sp_cols:
            if is_truthy(prow.get(sc)):
                rows.append(base_row(
                    col=sc,
                    matched=sp_col_to_binomial(sc),
                    section="N/A (binomial name match)",
                    threshold=1, total_freq=1, raw_freq=1, term_count=1,
                ))

        # ---- 5b: Technique (KM) ----
        for ac in a_cols:
            if is_truthy(prow.get(ac)):
                rows.append(base_row(
                    col=ac,
                    matched=a_col_to_label(ac),
                    section="N/A (technique name match)",
                    threshold=1, total_freq=1, raw_freq=1, term_count=1,
                ))

        # ---- 5c: Depth extraction (RE) ----
        dr = prow.get("depth_range")
        if not pd.isna(dr) and str(dr).strip() not in ("", "nan", "None"):
            rows.append(base_row(
                col="depth_range",
                matched=str(dr),
                section="N/A (regex extraction)",
                threshold=1, total_freq=1, raw_freq=1, term_count=1,
            ))

        # ---- 5d: Gear target species (RE) ----
        gts = prow.get("gear_target_species")
        if not pd.isna(gts) and str(gts).strip() not in ("", "nan", "None"):
            rows.append(base_row(
                col="gear_target_species",
                matched=str(gts),
                section="N/A (regex extraction)",
                threshold=1, total_freq=1, raw_freq=1, term_count=1,
            ))

        # ---- 5e: Impact direction (SA) ----
        imp_dir = prow.get("imp_direction")
        if not pd.isna(imp_dir) and str(imp_dir).strip().lower() not in ("", "nan", "none", "not stated"):
            rows.append(base_row(
                col="imp_direction",
                matched=str(imp_dir),
                section="SA (sentiment)",
                threshold=1, total_freq=1, raw_freq=1, term_count=1,
            ))

        # ---- 5f: Derived columns (DRV) ----
        for dcol in ("eco_1_guess", "eco_2_guess", "eco_3_guess", "imp_is_bycatch", "imp_quantified"):
            dval = prow.get(dcol)
            if is_truthy(dval):
                rows.append(base_row(
                    col=dcol,
                    matched=str(dval),
                    section="DRV (derived)",
                    threshold=1, total_freq=1, raw_freq=1, term_count=1,
                ))

    # ---- 5g: API/GEO lookups ----
    paper_ids_list = list(paper_ids)

    # OpenAlex: first author per paper
    if not openalex_authors.empty:
        oa_sub = openalex_authors[openalex_authors["lit_id_int"].isin(paper_ids)].copy()
        # Take first author (author_position == 'first' if available, else position 0)
        if "author_position" in oa_sub.columns:
            first_authors = oa_sub[oa_sub["author_position"] == "first"].drop_duplicates(subset="lit_id_int")
            # Fall back to first row per paper if no 'first' position
            fallback_mask = ~oa_sub["lit_id_int"].isin(first_authors["lit_id_int"])
            fallback = oa_sub[fallback_mask].drop_duplicates(subset="lit_id_int")
            first_authors = pd.concat([first_authors, fallback], ignore_index=True)
        else:
            first_authors = oa_sub.drop_duplicates(subset="lit_id_int")

        for _, arow in first_authors.iterrows():
            lid = int(arow["lit_id_int"])
            m = meta_lookup.get(lid, {})
            institution = str(arow.get("institution_name", "")) if not pd.isna(arow.get("institution_name", float("nan"))) else ""
            country = str(arow.get("institution_country", "")) if not pd.isna(arow.get("institution_country", float("nan"))) else ""
            gender = str(arow.get("gender", "")) if not pd.isna(arow.get("gender", float("nan"))) else ""
            summary = "; ".join(v for v in [institution, country, gender] if v)
            if summary:
                rows.append(make_extra_row(
                    lit_id=lid,
                    year=m.get("year", ""), authors=m.get("authors", ""),
                    journal=m.get("journal", ""), title=m.get("title", ""),
                    column="openalex_first_author",
                    matched_terms=summary,
                    section="API (OpenAlex)",
                    threshold="", total_freq="", raw_freq="", term_count="",
                    matched_anchors="", binary=1, context="",
                ))

    # Altmetric
    if not altmetric.empty:
        alt_sub = altmetric[altmetric["lit_id_int"].isin(paper_ids)].drop_duplicates(subset="lit_id_int")
        for _, arow in alt_sub.iterrows():
            lid = int(arow["lit_id_int"])
            m = meta_lookup.get(lid, {})
            score = arow.get("alt_score", "")
            summary = f"alt_score={score}"
            rows.append(make_extra_row(
                lit_id=lid,
                year=m.get("year", ""), authors=m.get("authors", ""),
                journal=m.get("journal", ""), title=m.get("title", ""),
                column="alt_score",
                matched_terms=summary,
                section="API (Altmetric)",
                threshold="", total_freq="", raw_freq="", term_count="",
                matched_anchors="", binary=1 if not pd.isna(score) else 0,
                context="",
            ))

    # Unpaywall
    if not unpaywall.empty:
        unp_sub = unpaywall[unpaywall["lit_id_int"].isin(paper_ids)].drop_duplicates(subset="lit_id_int")
        for _, urow in unp_sub.iterrows():
            lid = int(urow["lit_id_int"])
            m = meta_lookup.get(lid, {})
            oa_status = str(urow.get("oa_status", "")) if not pd.isna(urow.get("oa_status", float("nan"))) else ""
            oa_license = str(urow.get("oa_license", "")) if not pd.isna(urow.get("oa_license", float("nan"))) else ""
            summary = "; ".join(v for v in [oa_status, oa_license] if v)
            if summary:
                rows.append(make_extra_row(
                    lit_id=lid,
                    year=m.get("year", ""), authors=m.get("authors", ""),
                    journal=m.get("journal", ""), title=m.get("title", ""),
                    column="oa_status",
                    matched_terms=summary,
                    section="API (Unpaywall)",
                    threshold="", total_freq="", raw_freq="", term_count="",
                    matched_anchors="", binary=1 if oa_status not in ("", "closed") else 0,
                    context="",
                ))

    # Geographic (SQLite)
    if not geo_df.empty:
        geo_sub = geo_df[geo_df["lit_id_int"].isin(paper_ids)]
        for _, grow in geo_sub.iterrows():
            lid = int(grow["lit_id_int"])
            m = meta_lookup.get(lid, {})
            study_country = str(grow.get("study_country", "")) if not pd.isna(grow.get("study_country", float("nan"))) else ""
            auth_country = str(grow.get("first_author_country", "")) if not pd.isna(grow.get("first_author_country", float("nan"))) else ""
            basin = str(grow.get("study_ocean_basin", "")) if not pd.isna(grow.get("study_ocean_basin", float("nan"))) else ""
            summary = "; ".join(v for v in [f"auth={auth_country}", f"study={study_country}", f"basin={basin}"] if v.split("=")[1])
            if summary:
                rows.append(make_extra_row(
                    lit_id=lid,
                    year=m.get("year", ""), authors=m.get("authors", ""),
                    journal=m.get("journal", ""), title=m.get("title", ""),
                    column="geography",
                    matched_terms=summary,
                    section="GEO (geographic)",
                    threshold="", total_freq="", raw_freq="", term_count="",
                    matched_anchors="", binary=1, context="",
                ))

    return rows


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main() -> None:
    print("Loading data...")

    # Load evidence
    evidence = pd.read_csv(EVIDENCE_CSV, dtype={"literature_id": str})
    evidence["lit_id_int"] = evidence["literature_id"].apply(lit_id_to_int)
    print(f"  Evidence: {len(evidence):,} rows")

    # Load author-paper mapping
    authors = pd.read_csv(
        AUTHORS_CSV,
        usecols=["literature_id", "last_name"],
        dtype={"literature_id": str},
    )
    authors["lit_id_int"] = authors["literature_id"].apply(lit_id_to_int)
    authors["last_name_norm"] = authors["last_name"].apply(normalise_name)

    # Load paper metadata from parquet
    meta = pd.read_parquet(PARQUET, columns=["literature_id", "year", "authors", "journal", "title"])
    meta["lit_id_int"] = meta["literature_id"].apply(lit_id_to_int)
    meta = meta.drop_duplicates(subset="lit_id_int")
    print(f"  Papers (parquet): {len(meta):,}")
    print(f"  Author-paper rows: {len(authors):,}")

    # Build a quick lookup dict for metadata
    meta_lookup: dict[int, dict] = {
        int(r["lit_id_int"]): {
            "year": r["year"],
            "authors": r["authors"],
            "journal": r["journal"],
            "title": r["title"],
        }
        for _, r in meta.iterrows()
    }

    # -----------------------------------------------------------------------
    # Load extra parquet columns (sp_, a_, derived)
    # -----------------------------------------------------------------------
    print("  Loading sp_ and a_ columns from parquet (this may take a moment)...")
    all_parquet_cols = pd.read_parquet(PARQUET, columns=["literature_id"]).columns.tolist()
    # Get full column list without loading full parquet
    import pyarrow.parquet as pq_lib
    pq_file = pq_lib.ParquetFile(str(PARQUET))
    all_cols = pq_file.schema_arrow.names

    sp_cols = [c for c in all_cols if c.startswith("sp_")]
    a_cols = [c for c in all_cols if c.startswith("a_")]
    derived_cols = [c for c in all_cols if c in (
        "eco_1_guess", "eco_2_guess", "eco_3_guess",
        "imp_is_bycatch", "imp_quantified", "imp_direction",
        "depth_range", "depth_min_m", "depth_max_m",
        "gear_target_species",
    )]
    print(f"    sp_ columns: {len(sp_cols)}, a_ columns: {len(a_cols)}, derived: {len(derived_cols)}")

    extra_cols_to_load = ["literature_id"] + sp_cols + a_cols + derived_cols
    parquet_extra = pd.read_parquet(PARQUET, columns=extra_cols_to_load)
    parquet_extra["lit_id_int"] = parquet_extra["literature_id"].apply(lit_id_to_int)
    print(f"  Extra parquet rows: {len(parquet_extra):,}")

    # -----------------------------------------------------------------------
    # Load external data sources
    # -----------------------------------------------------------------------
    print("  Loading external data sources...")

    openalex_authors = pd.read_csv(AUTHORS_CSV, dtype={"literature_id": str})
    openalex_authors["lit_id_int"] = openalex_authors["literature_id"].apply(lit_id_to_int)

    altmetric = pd.read_csv(ALTMETRIC_CSV, dtype={"literature_id": str})
    altmetric["lit_id_int"] = altmetric["literature_id"].apply(lit_id_to_int)

    unpaywall = pd.read_csv(UNPAYWALL_CSV, dtype={"literature_id": str})
    unpaywall["lit_id_int"] = unpaywall["literature_id"].apply(lit_id_to_int)

    # Load geographic data from SQLite
    try:
        conn = sqlite3.connect(DB)
        geo_df = pd.read_sql_query(
            "SELECT paper_id, first_author_country, study_country, study_ocean_basin FROM paper_geography",
            conn,
        )
        conn.close()
        geo_df["lit_id_int"] = geo_df["paper_id"].apply(lit_id_to_int)
        print(f"  Geographic rows: {len(geo_df):,}")
    except Exception as exc:
        print(f"  Warning: could not load paper_geography ({exc})")
        geo_df = pd.DataFrame(columns=["paper_id", "first_author_country", "study_country", "study_ocean_basin", "lit_id_int"])

    # -----------------------------------------------------------------------
    # Build per-coauthor paper sets
    # -----------------------------------------------------------------------
    coauthor_paper_ids: dict[str, set[int]] = {}
    for name, variants in TEAM.items():
        norm_variants = [normalise_name(v) for v in variants]
        mask = authors["last_name_norm"].isin(norm_variants)
        paper_ids = set(authors.loc[mask, "lit_id_int"].unique())
        coauthor_paper_ids[name] = paper_ids
        print(f"  {name}: {len(paper_ids)} papers in OpenAlex")

    # -----------------------------------------------------------------------
    # Merge evidence with metadata
    # -----------------------------------------------------------------------
    # Drop the title column from evidence (we'll use parquet's version)
    ev_cols_from_csv = [c for c in evidence.columns if c != "title"]
    merged = evidence[ev_cols_from_csv].merge(
        meta[["lit_id_int", "year", "authors", "journal", "title"]],
        on="lit_id_int",
        how="left",
    )
    # Use integer literature_id for display
    merged["literature_id"] = merged["lit_id_int"]

    # -----------------------------------------------------------------------
    # Create workbook
    # -----------------------------------------------------------------------
    wb = Workbook()

    # -- Summary tab --
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.append(["Coauthor", "Papers in DB", "Evidence rows", "Note"])

    summary_rows: list[tuple[str, int, int, str]] = []

    for name in TEAM:
        paper_ids = coauthor_paper_ids[name]
        if not paper_ids:
            summary_rows.append((name, 0, 0, "No OpenAlex author match"))
            continue

        # Filter evidence to this coauthor's papers
        mask = merged["lit_id_int"].isin(paper_ids)
        coauthor_ev = merged.loc[mask].copy()

        # Build extra evidence rows
        extra_rows = build_extra_evidence(
            paper_ids=paper_ids,
            meta_lookup=meta_lookup,
            parquet_extra=parquet_extra,
            openalex_authors=openalex_authors,
            altmetric=altmetric,
            unpaywall=unpaywall,
            geo_df=geo_df,
            sp_cols=sp_cols,
            a_cols=a_cols,
        )

        n_schema_rows = len(coauthor_ev)
        n_extra_rows = len(extra_rows)
        n_papers = coauthor_ev["lit_id_int"].nunique()
        n_rows = n_schema_rows + n_extra_rows
        note = ""
        if n_rows > 50_000:
            note = f"Large tab: {n_rows:,} rows"

        summary_rows.append((name, n_papers, n_rows, note))

        # Build tab dataframe: schema evidence first
        tab_df = coauthor_ev[EVIDENCE_COLS].copy()

        # Append extra rows
        if extra_rows:
            extra_df = pd.DataFrame(extra_rows, columns=EVIDENCE_COLS)
            # Add separator row
            sep_row = pd.DataFrame(
                [{col: ("--- Additional evidence (species/techniques/lookups) ---" if col == "column" else "") for col in EVIDENCE_COLS}]
            )
            tab_df = pd.concat([tab_df, sep_row, extra_df], ignore_index=True)

        tab_df["Notes"] = ""
        tab_df = tab_df.sort_values(["literature_id", "column"], na_position="last").reset_index(drop=True)

        # Write tab
        ws = wb.create_sheet(title=name)
        write_dataframe(ws, tab_df)
        style_header(ws)
        auto_width(ws)
        ws.auto_filter.ref = ws.dimensions
        print(f"  Wrote tab '{name}': {n_schema_rows:,} schema rows + {n_extra_rows:,} extra rows, {n_papers} papers")

    # Write summary data
    for row in summary_rows:
        ws_summary.append(list(row))

    # -----------------------------------------------------------------------
    # Glossary section on Summary tab
    # -----------------------------------------------------------------------
    ws_summary.append([])
    ws_summary.append([])
    bold_gold = Font(bold=True, color="8B6914")
    section_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    # Glossary header
    gloss_header_row = ws_summary.max_row + 1
    ws_summary.append(["Evidence Table Field Glossary", "Description"])
    for cell in ws_summary[ws_summary.max_row]:
        cell.font = Font(bold=True)
        cell.fill = section_fill

    for field, desc in GLOSSARY_FIELDS:
        ws_summary.append([field, desc])

    ws_summary.append([])

    # Technique codes header
    ws_summary.append(["Extraction Technique Codes", "Description"])
    for cell in ws_summary[ws_summary.max_row]:
        cell.font = Font(bold=True)
        cell.fill = section_fill

    for code, desc in TECHNIQUE_CODES:
        ws_summary.append([code, desc])

    ws_summary.append([])

    # -----------------------------------------------------------------------
    # Acronym Matches tab
    # -----------------------------------------------------------------------
    print("Building Acronym Matches tab...")

    # Build regex pattern: match each acronym as a standalone token in matched_terms
    # matched_terms look like: "CPUE(3); marine(1); BRUVs(2)"
    # We want to find rows where any acronym appears before a "("
    acronym_pattern = re.compile(
        r"\b(" + "|".join(re.escape(a) for a in ACRONYMS) + r")\(",
    )

    def find_acronyms(text: str | float) -> str:
        """Return comma-separated list of acronyms found in matched_terms."""
        if pd.isna(text):
            return ""
        found = acronym_pattern.findall(str(text))
        return ", ".join(sorted(set(found))) if found else ""

    merged["acronym_matched"] = merged["matched_terms"].apply(find_acronyms)
    acr_mask = merged["acronym_matched"] != ""
    acr_df = merged.loc[acr_mask, EVIDENCE_COLS + ["acronym_matched"]].copy()
    acr_df = acr_df.sort_values(["literature_id", "column"]).reset_index(drop=True)

    n_acr = len(acr_df)
    print(f"  Acronym matches: {n_acr:,} evidence rows")

    ws_acr = wb.create_sheet(title="Acronym Matches")
    write_dataframe(ws_acr, acr_df)
    style_header(ws_acr)
    auto_width(ws_acr)
    ws_acr.auto_filter.ref = ws_acr.dimensions

    # Add acronym summary to the Summary tab
    ws_summary.append(["Acronym Matches", "", n_acr, "All evidence rows with acronym term matches"])

    # Breakdown per acronym
    ws_summary.append([])
    ws_summary.append(["Acronym", "Evidence rows"])
    acr_counts: dict[str, int] = {}
    for acrs in merged.loc[acr_mask, "acronym_matched"]:
        for a in acrs.split(", "):
            acr_counts[a] = acr_counts.get(a, 0) + 1
    for acr_name, count in sorted(acr_counts.items(), key=lambda x: -x[1]):
        ws_summary.append([acr_name, count])

    style_header(ws_summary)
    auto_width(ws_summary)
    ws_summary.auto_filter.ref = ws_summary.dimensions

    # -----------------------------------------------------------------------
    # Save
    # -----------------------------------------------------------------------
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    print(f"\nSaved to: {OUTPUT}")
    print(f"  File size: {OUTPUT.stat().st_size / 1024 / 1024:.1f} MB")


if __name__ == "__main__":
    main()
