#!/usr/bin/env python3
"""OCR non-extractable library PDFs in-place.

Reads `outputs/non_extractable_pdfs.xlsx` (produced by
`find_non_extractable_pdfs.py`) and runs ocrmypdf on each flagged PDF,
adding a text layer so future `pdftotext`-based extraction works.

Writes OCR output to a tempfile and atomically renames over the
original; ocrmypdf preserves original content and only adds a text
layer, so existing pages are not destroyed.

Usage:
    python scripts/ocr_library.py                # OCR all flagged PDFs
    python scripts/ocr_library.py --limit 10     # first N PDFs (testing)
    python scripts/ocr_library.py --dry-run      # list what would run
    python scripts/ocr_library.py --language eng+fra+deu  # multilingual
"""

import argparse
import os
import re
import subprocess
import sys
from pathlib import Path
from datetime import datetime

from openpyxl import load_workbook

PROJECT = Path("/media/simon/data/Documents/Si Work/PostDoc Work/EEA/2025/Data Panel")
PDF_ROOT = Path("/media/simon/data/Documents/Si Work/Papers & Books/SharkPapers")
REPORT = PROJECT / "outputs/non_extractable_pdfs.xlsx"
LOG_DIR = PROJECT / "logs"
LOG_FILE = LOG_DIR / "ocr_library_log.txt"

# Language -> tesseract code. Only packs actually installed will work;
# missing packs cause ocrmypdf to error, so we fall back to eng per-file.
LANG_MAP = {
    "English": "eng",
    "French": "fra",
    "German": "deu",
    "Spanish": "spa",
    "Portuguese": "por",
    "Italian": "ita",
    "Dutch": "nld",
    "Czech/Slovak/Croatian": "ces",
    "Russian": "rus",
    # The find_non_extractable detector emits the combined "Japanese/Chinese"
    # label (script is CJK, can't disambiguate), so try all three models.
    # Separate keys cover any detector that does distinguish them.
    "Japanese/Chinese": "jpn+chi_sim+chi_tra",
    "Japanese": "jpn",
    "Chinese": "chi_sim+chi_tra",
    "Korean": "kor",
    "Arabic": "ara",
    "Hindi": "hin",
    "Thai": "tha",
    "Unknown": "eng",
}


def installed_tesseract_langs() -> set[str]:
    try:
        result = subprocess.run(
            ["tesseract", "--list-langs"],
            capture_output=True, text=True, timeout=10,
        )
        # output: "List of available languages...\nlang1\nlang2\n..."
        lines = result.stderr.splitlines() + result.stdout.splitlines()
        return {l.strip() for l in lines if l.strip() and " " not in l.strip()}
    except Exception:
        return set()


def read_flagged_pdfs(report: Path = REPORT) -> list[tuple[str, Path]]:
    """Return list of (language, path) tuples for flagged PDFs.

    Reconstructs path from year + filename cols rather than the hyperlink
    cell, since openpyxl read-only mode does not expose hyperlinks.
    """
    wb = load_workbook(report, read_only=True, data_only=True)
    ws = wb.active
    rows = []
    # Index rather than unpack: the standard report has 4 cols, but a
    # language-resolved report (resolve_pdf_language.py) carries extra
    # columns (resolved_year, lang_source). Cols 0-2 are year, filename, lang.
    for row in ws.iter_rows(min_row=2, values_only=True):
        year, fname = row[0], row[1]
        lang = row[2] if len(row) > 2 else None
        if not year or not fname:
            continue
        path = PDF_ROOT / str(year) / str(fname)
        rows.append((str(lang or "Unknown"), path))
    wb.close()
    return rows


def ocr_one(pdf_path: Path, lang: str, available: set[str],
            redo: bool = False, timeout: int = 600) -> tuple[bool, str]:
    """OCR one PDF in place. Returns (success, message).

    redo=True re-OCRs pages that already carry a text layer (via
    ``--redo-ocr``), used to correct files OCR'd earlier under the wrong
    language model. Default (``--skip-text``) only OCRs image-only pages.
    Both modes fall back to ``--force-ocr`` if the first pass bails.
    """
    if not pdf_path.exists():
        return False, "file missing"

    # Translate language, fall back to eng if pack missing
    tlang = LANG_MAP.get(lang, "eng")
    parts = tlang.split("+")
    usable = [p for p in parts if p in available] or ["eng"]
    # Fraktur (blackletter) routing: many pre-~1940 German scientific texts
    # are set in Fraktur, which the roman-type `deu` model reads poorly. But
    # the era is mixed — by ~1900 many journals had switched to roman
    # (Antiqua) — so we ADD Fraktur as a combined "deu+Fraktur" (deu kept
    # first) rather than replacing deu. A/B tests showed this never hurts
    # roman pages (deu+Fraktur == deu) while giving tesseract the Fraktur
    # option for genuine blackletter. Year comes from the path's year folder.
    if "deu" in usable and "Fraktur" in available:
        try:
            year = int(re.match(r"(\d{4})", pdf_path.parent.name).group(1))
        except (AttributeError, ValueError):
            year = 9999
        if year < 1940:
            usable = [p for p in usable if p != "Fraktur"] + ["Fraktur"]
    tlang_used = "+".join(usable)

    first_mode = "--redo-ocr" if redo else "--skip-text"
    tmp_path = pdf_path.with_suffix(".ocr.tmp.pdf")
    try:
        result = subprocess.run(
            ["ocrmypdf", first_mode, "--output-type", "pdf",
             "--language", tlang_used, "--quiet",
             str(pdf_path), str(tmp_path)],
            capture_output=True, timeout=timeout,
        )
        if result.returncode != 0 or not tmp_path.exists() or tmp_path.stat().st_size == 0:
            # --skip-text may bail when all pages already have text (shouldn't
            # happen on flagged files but be robust); try force mode
            if tmp_path.exists():
                tmp_path.unlink()
            result = subprocess.run(
                ["ocrmypdf", "--force-ocr", "--output-type", "pdf",
                 "--language", tlang_used, "--quiet",
                 str(pdf_path), str(tmp_path)],
                capture_output=True, timeout=600,
            )
            if result.returncode != 0 or not tmp_path.exists() or tmp_path.stat().st_size == 0:
                err = result.stderr.decode("utf-8", errors="replace")[:200]
                if tmp_path.exists():
                    tmp_path.unlink()
                return False, f"ocrmypdf failed ({tlang_used}): {err}"

        # Atomic replace
        os.replace(tmp_path, pdf_path)
        return True, f"OK ({tlang_used})"
    except subprocess.TimeoutExpired:
        if tmp_path.exists():
            tmp_path.unlink()
        return False, "timeout"
    except Exception as e:
        if tmp_path.exists():
            tmp_path.unlink()
        return False, f"exception: {e}"


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--limit", type=int, default=0,
                    help="process only first N PDFs (testing)")
    ap.add_argument("--dry-run", action="store_true",
                    help="list what would run, no OCR")
    ap.add_argument("--skip-already-text", action="store_true", default=True,
                    help="skip PDFs that now have text (from partial runs)")
    ap.add_argument("--languages", default="",
                    help="comma-separated title_language labels to include "
                         "(e.g. 'German,French,Spanish'); default all")
    ap.add_argument("--redo", action="store_true",
                    help="re-OCR pages that already have a text layer "
                         "(--redo-ocr); use to correct wrong-language OCR. "
                         "Bypasses the 'already has text' skip check.")
    ap.add_argument("--report", default="",
                    help="override the input report XLSX (default: the "
                         "standard non_extractable_pdfs.xlsx). Use a filtered "
                         "report to resume a partial run.")
    ap.add_argument("--timeout", type=int, default=600,
                    help="per-file ocrmypdf timeout in seconds (default 600). "
                         "Raise for very large multi-hundred-page volumes.")
    args = ap.parse_args()

    report_path = Path(args.report) if args.report else REPORT
    if not report_path.exists():
        print(f"ERROR: report not found: {report_path}")
        print("Run: python scripts/find_non_extractable_pdfs.py")
        sys.exit(1)

    available = installed_tesseract_langs()
    print(f"Installed tesseract languages: {sorted(available)}")

    flagged = read_flagged_pdfs(report_path)
    print(f"Flagged PDFs in report: {len(flagged)}")
    if args.languages:
        wanted = {s.strip() for s in args.languages.split(",") if s.strip()}
        flagged = [(lang, p) for lang, p in flagged if lang in wanted]
        print(f"Filtered to languages {sorted(wanted)}: {len(flagged)}")
    if args.limit:
        flagged = flagged[: args.limit]
        print(f"Limiting to first {args.limit}")

    if args.dry_run:
        for lang, path in flagged[:20]:
            print(f"  [{lang}] {path.name}")
        if len(flagged) > 20:
            print(f"  ... and {len(flagged) - 20} more")
        return

    LOG_DIR.mkdir(exist_ok=True)
    # Per-report log so concurrent/sequential runs on different reports don't
    # clobber each other's record (the default full-run keeps the base name).
    log_file = (LOG_DIR / f"ocr_library_log_{report_path.stem}.txt"
                if args.report else LOG_FILE)
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ok_count = fail_count = skipped = 0
    # Incremental append-logging: write each result as it happens and flush,
    # so a kill mid-run leaves an accurate record of what completed.
    logf = open(log_file, "w")
    def logline(s: str) -> None:
        logf.write(s + "\n")
        logf.flush()
    logline(f"OCR library run: {ts}")
    logline(f"Tesseract langs: {sorted(available)}")
    logline(f"Report: {report_path}  redo={args.redo}  n={len(flagged)}\n")

    for i, (lang, path) in enumerate(flagged, 1):
        # Optional: re-check extractability (skip if a partial previous run
        # already OCR'd this file). In --redo mode we deliberately reprocess
        # files that already have text, so this skip check is bypassed.
        if not args.redo:
            try:
                result = subprocess.run(
                    ["pdftotext", "-l", "1", str(path), "-"],
                    capture_output=True, timeout=15,
                )
                alpha = sum(1 for c in result.stdout.decode("utf-8", errors="replace")
                            if c.isalpha())
                if alpha >= 50:
                    skipped += 1
                    logline(f"SKIP (now has text): {path.name}")
                    if i % 50 == 0:
                        print(f"  [{i}/{len(flagged)}] skipped {path.name}")
                    continue
            except Exception:
                pass

        ok, msg = ocr_one(path, lang, available, redo=args.redo, timeout=args.timeout)
        status = "OK " if ok else "FAIL"
        print(f"  [{i}/{len(flagged)}] {status} [{lang}] {path.name} — {msg}")
        logline(f"{status}: {path.name} [{lang}] — {msg}")
        if ok:
            ok_count += 1
        else:
            fail_count += 1

    summary = (f"\n{'=' * 60}\nOCR library summary\n"
               f"  Total flagged: {len(flagged)}\n"
               f"  OCR'd:         {ok_count}\n"
               f"  Skipped:       {skipped}\n"
               f"  Failed:        {fail_count}\n")
    print(summary)
    logline(summary)
    logf.close()
    print(f"Log: {log_file}")


if __name__ == "__main__":
    main()
