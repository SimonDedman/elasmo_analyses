#!/usr/bin/env python3
"""Apply decisions from ocr_retry_report.xlsx notes.

Reads the user's annotated notes in `outputs/ocr_retry_report.xlsx` and
executes the appropriate action per row:

  "check for duplicate: ..."      → look up whether SharkPapers holds an
                                    alternate copy of the same paper. If
                                    yes, delete only this one. If no,
                                    delete AND re-queue with a
                                    `source_corrupted_retry` flag.
  "has dupe, delete both, re-add" → delete this and the dupe; re-queue
                                    once as `accidental_duplicate`.
  "is SM doc ..., name & file"    → rename to `<stem>_SM.pdf` (adds the
                                    _SM suffix if not already present).
                                    No DB change — SM files are not
                                    separate DB entries.
  "delete"                        → delete file; mark DB entries as
                                    `unavailable_unrecoverable`.

Defaults to dry run; emits `outputs/ocr_residual_decisions_plan.xlsx`
listing every planned action. Pass ``--apply`` to actually execute.
"""

import argparse
import csv
import json
import os
import shutil
import sqlite3
import sys
import unicodedata
from datetime import datetime
from pathlib import Path
from urllib.parse import quote

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

PROJECT = Path("/media/simon/data/Documents/Si Work/PostDoc Work/EEA/2025/Data Panel")
PDF_ROOT = Path("/media/simon/data/Documents/Si Work/Papers & Books/SharkPapers")
REPORT_IN = PROJECT / "outputs/ocr_retry_report.xlsx"
PLAN_OUT = PROJECT / "outputs/ocr_residual_decisions_plan.xlsx"
LOG_OUT = PROJECT / "logs/apply_ocr_residual_decisions.log"
VIZ_DATA = PROJECT / "outputs/viz_data.csv"
PAPERS_JSON = PROJECT / "docs/papers_data.json"
QUEUE_DB = PROJECT / "outputs/download_queue.db"
TRACKER_DB = PROJECT / "database/download_tracker.db"


# ----- database helpers -------------------------------------------------

def load_viz_data() -> list[dict]:
    with open(VIZ_DATA) as f:
        return list(csv.DictReader(f))


def build_filename(row: dict) -> str:
    """Replicate ingest_pdfs.build_filename (cannot import cleanly)."""
    import re
    authors = str(row.get("authors", "")).strip()
    first = authors.split(" & ")[0].strip()
    first = re.sub(r"\(\d{4}\)", "", first).strip()
    if "," in first:
        surname = first.split(",")[0].strip()
    else:
        parts = first.split()
        surname = parts[-1] if parts else "Unknown"
    surname = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "", surname)[:20] or "Unknown"
    try:
        year = str(int(float(row.get("year", ""))))
    except (ValueError, TypeError):
        year = "Unknown"
    title = str(row.get("title", "")).strip()
    title = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "", title)
    title = " ".join(title.split())
    if len(title) > 60:
        title = title[:60].rsplit(" ", 1)[0]
    multi = "&" in authors
    return (f"{surname}.etal.{year}.{title}.pdf"
            if multi
            else f"{surname}.{year}.{title}.pdf")


def row_for_filename(rows: list[dict], target_path: Path) -> dict | None:
    """Find DB row whose canonical filename matches ``target_path``.

    First tries exact canonical filename match. If none, falls back to
    fuzzy match on filename year + surname prefix + title word overlap.
    Historical ingest runs truncated titles differently, so exact-match
    frequently misses for older files.
    """
    import re
    target_name = target_path.name
    target_year = target_path.parent.name

    # Candidates: DB rows in same year
    year_candidates = []
    for r in rows:
        try:
            ry = str(int(float(r["year"])))
        except (ValueError, TypeError):
            continue
        if ry != target_year:
            continue
        year_candidates.append(r)

    # Pass 1: exact canonical match
    for r in year_candidates:
        if build_filename(r) == target_name:
            return r

    # Pass 2: surname+year prefix + title-word overlap
    # Extract stem components: "Surname[.etal].YYYY.title words..."
    stem = Path(target_name).stem
    # Surname may contain apostrophes (Dell'apa), hyphens (Young-Veenstra),
    # or spaces (de Carvalho, van der X) — match up to the first run of dots.
    m = re.match(r"^([A-Za-zÀ-ÿ][A-Za-zÀ-ÿ'\- ]*?)(\.etal)?\.\d{4}\.(.+)$", stem)
    if not m:
        return None
    # Normalise surname: strip accents, apostrophes, spaces, hyphens, lowercase
    raw_surname = m.group(1)
    surname = re.sub(r"[ '\-]", "", strip_accents_lower(raw_surname))
    is_multi = m.group(2) is not None
    fname_words = set(re.findall(r"[a-z]{4,}", m.group(3).lower()))

    best = None
    best_overlap = 0
    for r in year_candidates:
        auths = str(r.get("authors", "")).strip()
        first = auths.split(" & ")[0]
        first = re.sub(r"\(\d{4}\)", "", first).strip()
        db_surname = first.split(",")[0].strip() if "," in first else first.split()[-1]
        db_surname_norm = re.sub(r"[ '\-]", "", strip_accents_lower(db_surname))
        if db_surname_norm != surname:
            continue
        if is_multi != ("&" in auths):
            continue  # multi-author mismatch
        db_title_words = set(re.findall(r"[a-z]{4,}", str(r.get("title", "")).lower()))
        overlap = len(fname_words & db_title_words)
        if overlap > best_overlap:
            best_overlap = overlap
            best = r
    # Accept match if >=3 title words overlap OR the surname+year is unique
    if best and best_overlap >= 3:
        return best
    if best and best_overlap == 0:
        # fall back to unique surname+year+multiauthor-flag
        same_surname = [
            r for r in year_candidates
            if (lambda auths: (
                (lambda f: re.sub(r"[ '\-]", "", strip_accents_lower(
                    f.split(",")[0].strip() if "," in f else f.split()[-1]
                )))(re.sub(r"\(\d{4}\)", "", auths.split(" & ")[0]).strip()) == surname
                and (("&" in auths) == is_multi)
            ))(str(r.get("authors", "")).strip())
        ]
        if len(same_surname) == 1:
            return same_surname[0]
    return best if best and best_overlap >= 2 else None


def strip_accents_lower(s: str) -> str:
    return "".join(c for c in unicodedata.normalize("NFD", s)
                   if unicodedata.category(c) != "Mn").lower()


def find_alternate_copies(db_row: dict) -> list[Path]:
    """Scan library for PDFs likely to be the same paper under another name.

    Checks (in ``PDF_ROOT/<year>/`` only):
      - files sharing the surname+year prefix (e.g. ``Carlson.2002.*``)
        except the canonical name
    Returns list of matching paths.
    """
    import re
    try:
        year = str(int(float(db_row.get("year", ""))))
    except (ValueError, TypeError):
        return []
    year_dir = PDF_ROOT / year
    if not year_dir.exists():
        return []

    authors = str(db_row.get("authors", "")).strip()
    first = authors.split(" & ")[0].strip()
    first = re.sub(r"\(\d{4}\)", "", first).strip()
    surname = first.split(",")[0].strip() if "," in first else first.split()[-1]
    surname = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "", surname)[:20]

    canonical = build_filename(db_row)
    prefix = f"{surname}."
    candidates = [p for p in year_dir.glob(f"{surname}*{year}*.pdf")
                  if p.name != canonical]
    return candidates


# ----- decision model --------------------------------------------------

ACT_DUP_CHECK = "dedup_check_delete_requeue"
ACT_DUAL_DELETE = "dual_delete_requeue"
ACT_SM_RENAME = "sm_rename"
ACT_DELETE = "delete_only"
ACT_UNKNOWN = "unknown"


def classify_note(note: str) -> str:
    n = (note or "").strip().lower()
    if not n:
        return ACT_UNKNOWN
    if "has dupe" in n and "delete both" in n:
        return ACT_DUAL_DELETE
    if "check for duplicate" in n:
        return ACT_DUP_CHECK
    if "sm doc" in n or "supplementary" in n:
        return ACT_SM_RENAME
    if n == "delete":
        return ACT_DELETE
    return ACT_UNKNOWN


def sm_new_name(old: Path) -> Path:
    """Compute the _SM.pdf filename. Idempotent."""
    stem = old.stem
    if stem.endswith("_SM"):
        return old  # already correct
    return old.with_name(f"{stem}_SM.pdf")


def corruption_reason(path: Path) -> str:
    """Classify the failure type based on file content."""
    if not path.exists():
        return "file missing"
    sz = path.stat().st_size
    try:
        with open(path, "rb") as f:
            head = f.read(256)
    except Exception:
        return f"unreadable ({sz} bytes)"
    if head[:5] == b"%PDF-":
        if sz < 20_000:
            return f"truncated_pdf ({sz} bytes)"
        return f"structurally_damaged_pdf ({sz:,} bytes)"
    if b"<!DOCTYPE" in head[:100] or b"<html" in head[:100].lower() or head.startswith(b"\n\n<"):
        return f"html_instead_of_pdf ({sz:,} bytes)"
    return f"unknown_format ({sz} bytes, head={head[:20]!r})"


# ----- action executors ------------------------------------------------

def delete_file(path: Path, dry: bool) -> str:
    if not path.exists():
        return "skip: file already gone"
    if dry:
        return f"WOULD DELETE: {path.name}"
    path.unlink()
    return f"deleted: {path.name}"


def rename_to_sm(old: Path, dry: bool) -> tuple[Path, str]:
    new = sm_new_name(old)
    if new == old:
        return old, "already _SM suffix"
    if new.exists():
        return old, f"target exists: {new.name}"
    if dry:
        return new, f"WOULD RENAME: {old.name} → {new.name}"
    old.rename(new)
    return new, f"renamed: {old.name} → {new.name}"


def requeue_in_queue_db(db_row: dict, reason: str, dry: bool) -> str:
    """Flip download_queue.db row for this DOI to pending + corrupt flag."""
    if not QUEUE_DB.exists():
        return "queue_db missing"
    doi = (db_row or {}).get("doi", "").strip()
    if not doi:
        return "skip queue_db: no DOI"
    if dry:
        return f"WOULD queue_db: set status=pending, error_message='{reason}' for {doi}"
    con = sqlite3.connect(str(QUEUE_DB))
    try:
        cur = con.execute(
            "SELECT id, status FROM download_queue WHERE LOWER(doi) = LOWER(?)",
            (doi,),
        ).fetchone()
        ts = datetime.now().isoformat(timespec="seconds")
        if cur:
            con.execute(
                "UPDATE download_queue "
                "SET status='pending', error_message=?, pdf_path=NULL, "
                "    download_timestamp=NULL "
                "WHERE id = ?",
                (reason, cur[0]),
            )
            msg = f"queue_db row {cur[0]}: {cur[1]} → pending ({reason})"
        else:
            try:
                yr = int(float(db_row.get("year", "")))
            except (ValueError, TypeError):
                yr = None
            con.execute(
                "INSERT INTO download_queue "
                "(literature_id, doi, year, authors, title, source, "
                " priority, status, added_timestamp, error_message) "
                "VALUES (?, ?, ?, ?, ?, 'ocr_residual_requeue', 5, "
                "        'pending', ?, ?)",
                (db_row.get("literature_id", ""), doi, yr,
                 db_row.get("authors", ""), db_row.get("title", ""),
                 ts, reason),
            )
            msg = f"queue_db: inserted new pending row ({reason})"
        con.commit()
        return msg
    finally:
        con.close()


def requeue_in_tracker_db(db_row: dict, reason: str, dry: bool) -> str:
    """Reset tracker DB download_status for this paper."""
    if not TRACKER_DB.exists():
        return "tracker_db missing"
    lid = str(db_row.get("literature_id", "")).strip()
    if not lid:
        return "skip tracker: no lit_id"
    if dry:
        return f"WOULD tracker_db: reset status=pending with notes for lit_id={lid}"
    con = sqlite3.connect(str(TRACKER_DB))
    try:
        try:
            lid_int = int(float(lid))
        except (ValueError, TypeError):
            return f"skip tracker: bad lit_id {lid!r}"
        row = con.execute(
            "SELECT id FROM papers WHERE literature_id IN (?, ?)",
            (str(lid_int), f"{lid_int}.0"),
        ).fetchone()
        if not row:
            return f"skip tracker: lit_id {lid_int} not in papers"
        paper_id = row[0]
        ts = datetime.now().isoformat(timespec="seconds")
        # Update any downloaded status rows for this paper
        updated = con.execute(
            "UPDATE download_status SET status='pending', notes=?, "
            "    last_attempt=? WHERE paper_id=?",
            (reason, ts, paper_id),
        ).rowcount
        if updated == 0:
            con.execute(
                "INSERT INTO download_status "
                "(paper_id, status, notes, attempts, last_attempt) "
                "VALUES (?, 'pending', ?, 0, ?)",
                (paper_id, reason, ts),
            )
            updated = 1
        con.commit()
        return f"tracker_db: reset {updated} row(s) for paper_id={paper_id}"
    finally:
        con.close()


def readd_to_papers_json(db_row: dict, reason: str, dry: bool) -> str:
    """Ensure this paper is in papers_data.json with corruption note."""
    if not PAPERS_JSON.exists():
        return "papers_json missing"
    doi = (db_row or {}).get("doi", "").strip().lower()
    lid = str(db_row.get("literature_id", "")).strip()
    if dry:
        return f"WOULD papers_json: ensure entry for lit_id={lid} / doi={doi} notes='{reason}'"
    with open(PAPERS_JSON) as f:
        data = json.load(f)
    # Find any existing entry by lit_id or DOI
    for entry in data:
        e_lid = str(entry.get("literature_id", "")).strip()
        e_doi = str(entry.get("doi", "")).strip().lower()
        if (lid and (e_lid == lid or e_lid == f"{lid}.0")) or (doi and e_doi == doi):
            prev_notes = entry.get("notes") or ""
            entry["notes"] = (f"{prev_notes} | {reason}".strip(" |")
                              if prev_notes and reason not in prev_notes
                              else reason or prev_notes)
            entry["last_status"] = "source_corrupted_retry"
            with open(PAPERS_JSON, "w") as f:
                json.dump(data, f, indent=2)
            return f"papers_json: updated existing entry for {doi or lid}"
    # Insert new
    try:
        yr = int(float(db_row.get("year", "")))
    except (ValueError, TypeError):
        yr = None
    new_entry = {
        "id": max(int(e.get("id", 0)) for e in data) + 1 if data else 1,
        "literature_id": lid,
        "year": yr,
        "authors": db_row.get("authors", ""),
        "title": db_row.get("title", ""),
        "journal": db_row.get("journal", ""),
        "doi": db_row.get("doi", ""),
        "priority_group": 5,
        "last_status": "source_corrupted_retry",
        "notes": reason,
    }
    data.append(new_entry)
    with open(PAPERS_JSON, "w") as f:
        json.dump(data, f, indent=2)
    return f"papers_json: inserted new entry for {doi or lid}"


def mark_unrecoverable(db_row: dict, reason: str, dry: bool) -> str:
    """Mark a paper as permanently unavailable in tracker DB."""
    if not TRACKER_DB.exists():
        return "tracker_db missing"
    lid = str(db_row.get("literature_id", "")).strip()
    if not lid:
        return "skip tracker: no lit_id"
    if dry:
        return f"WOULD tracker_db: mark lit_id={lid} unavailable ({reason})"
    con = sqlite3.connect(str(TRACKER_DB))
    try:
        try:
            lid_int = int(float(lid))
        except (ValueError, TypeError):
            return f"bad lit_id {lid!r}"
        row = con.execute(
            "SELECT id FROM papers WHERE literature_id IN (?, ?)",
            (str(lid_int), f"{lid_int}.0"),
        ).fetchone()
        if not row:
            return f"lit_id {lid_int} not in papers"
        paper_id = row[0]
        ts = datetime.now().isoformat(timespec="seconds")
        updated = con.execute(
            "UPDATE download_status SET status='unavailable', notes=?, "
            "    last_attempt=? WHERE paper_id=?",
            (reason, ts, paper_id),
        ).rowcount
        if updated == 0:
            con.execute(
                "INSERT INTO download_status "
                "(paper_id, status, notes, attempts, last_attempt) "
                "VALUES (?, 'unavailable', ?, 1, ?)",
                (paper_id, reason, ts),
            )
            updated = 1
        con.commit()
        return f"tracker_db: marked {updated} row(s) unavailable"
    finally:
        con.close()


# ----- planning & execution -------------------------------------------

def plan(rows: list[dict], db_rows: list[dict]) -> list[dict]:
    plans = []
    for r in rows:
        year = r["year"]
        fname = r["filename"]
        note = r["notes"]
        path = PDF_ROOT / year / fname
        category = classify_note(note)

        db_row = row_for_filename(db_rows, path)
        lit_id = (db_row or {}).get("literature_id", "")
        doi = (db_row or {}).get("doi", "")
        reason_tag = ""
        planned_steps = []
        warnings = []

        if category == ACT_DUP_CHECK:
            reason_tag = "source_corrupted_retry: " + corruption_reason(path)
            if db_row:
                alts = find_alternate_copies(db_row)
            else:
                alts = []
                warnings.append("DB row not found for canonical filename")
            if alts:
                planned_steps.append(f"alternate copies found: {', '.join(p.name for p in alts)}")
                planned_steps.append("delete this PDF only (keep alternate)")
                planned_steps.append("no re-queue (alternate exists)")
            else:
                planned_steps.append("no alternate copy in library")
                planned_steps.append("delete this PDF")
                planned_steps.append(f"re-queue with tag: {reason_tag}")
        elif category == ACT_DUAL_DELETE:
            reason_tag = "accidental_duplicate: reacquire single canonical copy"
            planned_steps.append("delete this PDF")
            planned_steps.append(f"re-queue once with tag: {reason_tag} (dedup of sister file handled separately)")
        elif category == ACT_SM_RENAME:
            new = sm_new_name(path)
            if new == path:
                planned_steps.append("already _SM suffix, no action")
            else:
                planned_steps.append(f"rename to {new.name}")
            if db_row:
                warnings.append("DB row matched — _SM file is same DB entry (unusual)")
        elif category == ACT_DELETE:
            reason_tag = "unavailable_unrecoverable: marked for deletion (OCR unrecoverable)"
            planned_steps.append("delete this PDF")
            planned_steps.append(f"tracker_db: mark unavailable ({reason_tag})")
        else:
            planned_steps.append("UNKNOWN note — no action")
            warnings.append(f"unrecognised note: {note!r}")

        plans.append({
            "year": year, "filename": fname, "path": str(path),
            "note": note, "category": category,
            "literature_id": lit_id, "doi": doi,
            "planned_steps": " | ".join(planned_steps),
            "requeue_tag": reason_tag,
            "warnings": " | ".join(warnings),
        })
    return plans


def execute(plans: list[dict], db_rows: list[dict], dry: bool) -> list[dict]:
    """Execute planned actions; returns list of result dicts."""
    results = []
    for p in plans:
        path = Path(p["path"])
        db_row = row_for_filename(db_rows, path) if path.exists() else None
        # Fall back to lit_id-based DB lookup if row_for_filename missed
        if db_row is None and p["literature_id"]:
            for r in db_rows:
                if r.get("literature_id") == p["literature_id"]:
                    db_row = r
                    break

        logs = []
        try:
            if p["category"] == ACT_DUP_CHECK:
                alts = find_alternate_copies(db_row) if db_row else []
                logs.append(delete_file(path, dry))
                if not alts and db_row:
                    logs.append(requeue_in_queue_db(db_row, p["requeue_tag"], dry))
                    logs.append(requeue_in_tracker_db(db_row, p["requeue_tag"], dry))
                    logs.append(readd_to_papers_json(db_row, p["requeue_tag"], dry))
            elif p["category"] == ACT_DUAL_DELETE:
                logs.append(delete_file(path, dry))
                if db_row:
                    logs.append(requeue_in_queue_db(db_row, p["requeue_tag"], dry))
                    logs.append(requeue_in_tracker_db(db_row, p["requeue_tag"], dry))
                    logs.append(readd_to_papers_json(db_row, p["requeue_tag"], dry))
            elif p["category"] == ACT_SM_RENAME:
                _, msg = rename_to_sm(path, dry)
                logs.append(msg)
            elif p["category"] == ACT_DELETE:
                logs.append(delete_file(path, dry))
                if db_row:
                    logs.append(mark_unrecoverable(db_row, p["requeue_tag"], dry))
            else:
                logs.append("no action (unknown category)")
        except Exception as e:
            logs.append(f"EXCEPTION: {e}")

        results.append({**p, "execution_log": " | ".join(logs)})
    return results


def write_plan_xlsx(plans: list[dict], out: Path, applied: bool) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Apply Residuals Plan" if not applied else "Apply Residuals Results"
    headers = ["year", "filename", "category", "note", "literature_id",
               "doi", "planned_steps", "requeue_tag", "warnings",
               "execution_log", "open_file"]
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")
    link_font = Font(color="0000FF", underline="single")
    for i, p in enumerate(plans, 2):
        ws.cell(row=i, column=1, value=p["year"])
        ws.cell(row=i, column=2, value=p["filename"])
        ws.cell(row=i, column=3, value=p["category"])
        ws.cell(row=i, column=4, value=p["note"])
        ws.cell(row=i, column=5, value=p["literature_id"])
        ws.cell(row=i, column=6, value=p["doi"])
        ws.cell(row=i, column=7, value=p["planned_steps"])
        ws.cell(row=i, column=8, value=p["requeue_tag"])
        ws.cell(row=i, column=9, value=p["warnings"])
        ws.cell(row=i, column=10, value=p.get("execution_log", ""))
        uri = "file://" + quote(p["path"], safe="/:")
        c = ws.cell(row=i, column=11, value="Open")
        c.hyperlink = uri
        c.font = link_font
    for col_idx, h in enumerate(headers, 1):
        max_len = len(h)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=2,
                                max_row=ws.max_row, values_only=True):
            val = str(row[0]) if row[0] is not None else ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 80)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    wb.save(out)


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--apply", action="store_true",
                    help="actually execute actions (default: dry run)")
    args = ap.parse_args()

    # Load annotated report
    if not REPORT_IN.exists():
        print(f"ERROR: {REPORT_IN} not found")
        sys.exit(1)
    wb = load_workbook(REPORT_IN, data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        year, fname, lang, prev, notes, action, b, a, final, outcome, *_ = (
            row + (None,) * 11
        )[:11]
        if not year or not fname:
            continue
        if not (notes and str(notes).strip()):
            continue
        rows.append({
            "year": str(year), "filename": str(fname),
            "prev_status": str(prev or ""), "final_status": str(final or ""),
            "notes": str(notes),
        })
    wb.close()
    print(f"Loaded {len(rows)} annotated rows from {REPORT_IN.name}")

    db_rows = load_viz_data()
    print(f"Loaded {len(db_rows):,} DB rows from viz_data.csv")

    plans = plan(rows, db_rows)

    # Summary by category
    from collections import Counter
    by_cat = Counter(p["category"] for p in plans)
    print("\nCategories:")
    for cat, n in by_cat.most_common():
        print(f"  {cat}: {n}")

    if args.apply:
        print("\n*** APPLYING ACTIONS (not a dry run) ***")
        results = execute(plans, db_rows, dry=False)
    else:
        print("\n[DRY RUN — no changes]")
        results = execute(plans, db_rows, dry=True)

    write_plan_xlsx(results, PLAN_OUT, applied=args.apply)
    LOG_OUT.parent.mkdir(exist_ok=True)
    with open(LOG_OUT, "w") as f:
        f.write(f"apply_ocr_residual_decisions  dry={not args.apply}  "
                f"ts={datetime.now().isoformat(timespec='seconds')}\n\n")
        for r in results:
            f.write(f"[{r['category']}] {r['year']}/{r['filename']}\n")
            for line in r["execution_log"].split(" | "):
                f.write(f"    {line}\n")
            f.write("\n")
    print(f"\nPlan:  {PLAN_OUT}")
    print(f"Log:   {LOG_OUT}")
    if not args.apply:
        print("\nReview the plan XLSX, then re-run with --apply to execute.")


if __name__ == "__main__":
    main()
