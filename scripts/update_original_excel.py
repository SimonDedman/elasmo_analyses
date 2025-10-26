#!/usr/bin/env python3
"""
Update search_query_new column DIRECTLY in the original Excel file.
"""

import pandas as pd
from openpyxl import load_workbook
import re

def parse_synonyms(syn_text):
    """Parse synonyms field into a list of clean terms."""
    if pd.isna(syn_text) or not syn_text or syn_text == "NaN":
        return []
    terms = re.split(r'[,;]', str(syn_text))
    return [s.strip() for s in terms if s.strip()]

def should_quote(term):
    """Determine if a term needs quotes (multi-word phrases)."""
    if term.startswith('"') and term.endswith('"'):
        return False
    if ' ' in term or any(c in term for c in ['/', '-', '&']):
        return True
    return False

def format_term(term):
    """Format a search term with quotes if needed."""
    if should_quote(term):
        clean = term.strip('"')
        return f'"{clean}"'
    return term

def generate_search_query_new(row):
    """Generate corrected search query using OR logic."""
    technique = row['technique_name']
    synonyms_raw = row['synonyms']

    synonyms = parse_synonyms(synonyms_raw)
    all_terms = [technique] + synonyms

    # Remove duplicates (case-insensitive)
    unique_terms = []
    seen_lower = set()
    for term in all_terms:
        term_lower = term.lower()
        if term_lower not in seen_lower:
            unique_terms.append(term)
            seen_lower.add(term_lower)

    # Format and join
    formatted_terms = [format_term(term) for term in unique_terms]
    return " OR ".join(formatted_terms)

# File path
excel_path = "data/Techniques DB for Panel Review.xlsx"

print(f"Reading: {excel_path}")

# Load workbook with openpyxl to preserve formatting
wb = load_workbook(excel_path)

# Read Full_List sheet with pandas
df = pd.read_excel(excel_path, sheet_name='Full_List')
print(f"Processing {len(df)} techniques...")

# Generate new search queries
df['search_query_new'] = df.apply(generate_search_query_new, axis=1)

# Write back to the same sheet
ws = wb['Full_List']

# Find the column index for search_query_new
header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=False))
col_idx = None
for idx, cell in enumerate(header_row, 1):
    if cell.value == 'search_query_new':
        col_idx = idx
        break

if col_idx is None:
    print("ERROR: search_query_new column not found!")
else:
    print(f"Found search_query_new at column {col_idx}")

    # Write the values starting from row 2 (after header)
    for row_idx, value in enumerate(df['search_query_new'], start=2):
        ws.cell(row=row_idx, column=col_idx, value=value)

    # Save the workbook
    wb.save(excel_path)
    print(f"\n✅ Updated {excel_path} successfully!")

    # Show some examples
    print("\nSample updated queries:")
    examples = df[['technique_name', 'synonyms', 'search_query_new']].head(10)
    print(examples.to_string(index=False))

    print(f"\nTotal non-empty queries: {df['search_query_new'].notna().sum()}")
