#!/usr/bin/env python3
"""Run v8 confidence-scored fusion on the duplicates tab.

Loads features from the cache for each pair, runs fusion.evaluate_pair,
writes four new columns to the duplicates tab and preserves all others:

    v8_decision    - one of: same_content, keep_1, keep_2, rename_sm,
                     is_corrigendum, is_book, is_chapter, distinct,
                     manual, file_missing, needs_index
    v8_confidence  - 0.0 - 1.0
    v8_rule        - primary driving rule name
    v8_signals     - semicolon-separated list of signals that fired

Sorts the duplicates tab by (v8_confidence DESC, step) so you review the
trustworthy cases first and the borderline cases together.

Usage:
    python3 scripts/evaluate_duplicates_v8.py              # evaluate + write
    python3 scripts/evaluate_duplicates_v8.py --dry-run    # print only
    python3 scripts/evaluate_duplicates_v8.py --step 2_very_likely
    python3 scripts/evaluate_duplicates_v8.py --compare    # vs user notes
"""
from __future__ import annotations

import argparse
import sys
from collections import Counter
from pathlib import Path

import pandas as pd

SCRIPT_DIR = Path(__file__).parent
sys.path.insert(0, str(SCRIPT_DIR.parent))

from scripts.dedup import FeatureCache  # noqa: E402
from scripts.dedup.fusion import evaluate_pair  # noqa: E402


XLSX = Path("/media/simon/data/Documents/Si Work/PostDoc Work/EEA/2025/Data Panel"
            "/outputs/pdf_library_audit.xlsx")
CACHE_PATH = Path("/media/simon/data/Documents/Si Work/PostDoc Work/EEA/2025/"
                  "Data Panel/outputs/pdf_features.sqlite")
PDF_BASE = Path("/media/simon/data/Documents/Si Work/Papers & Books/SharkPapers")


NEW_COLS = ["v8_decision", "v8_confidence", "v8_rule", "v8_signals", "v8_keep"]


def rel_path(year, fname: str) -> str:
    return f"{int(year)}/{fname}"


def evaluate_rows(df: pd.DataFrame, cache: FeatureCache) -> pd.DataFrame:
    paths = set()
    for _, r in df.iterrows():
        try:
            paths.add(rel_path(r["year_1"], r["file_1"]))
            paths.add(rel_path(r["year_2"], r["file_2"]))
        except Exception:
            pass
    feats = cache.bulk_get(paths)
    print(f"Features loaded: {len(feats)}/{len(paths)} referenced paths")

    rows = []
    for _, r in df.iterrows():
        try:
            p1 = rel_path(r["year_1"], r["file_1"])
            p2 = rel_path(r["year_2"], r["file_2"])
        except Exception:
            rows.append(("file_missing", 0.0, "R_bad_row", "bad year/file", None))
            continue
        f1 = feats.get(p1); f2 = feats.get(p2)
        if f1 is None or f2 is None:
            missing = []
            if f1 is None: missing.append(p1)
            if f2 is None: missing.append(p2)
            rows.append(("needs_index", 0.0, "R_uncached",
                         "run build_features.py on: " + "; ".join(missing), None))
            continue
        d = evaluate_pair(f1, f2)
        rows.append((d.decision, round(d.confidence, 3), d.rule,
                     "; ".join(d.signals), d.keep))

    result = df.copy()
    for i, col in enumerate(NEW_COLS):
        result[col] = [row[i] for row in rows]
    return result


def compare_against_notes(df: pd.DataFrame) -> None:
    """Quick accuracy check against user notes, v8 vs v7."""
    noted = df[df["notes"].notna()].copy()
    print(f"\nNotes-labelled rows: {len(noted)}")
    if noted.empty:
        return

    agree_v8 = 0; agree_v7 = 0; checked = 0
    v8_wrong = []
    for _, r in noted.iterrows():
        note = str(r["notes"]).lower().strip()
        v8 = r.get("v8_decision")
        v7 = r.get("auto_decision")
        expected = None
        if note == "y":
            expected = {"keep_1", "keep_2", "same_content", "rename_sm", "is_corrigendum"}
        elif "distinct" in note or "different" in note:
            expected = {"distinct"}
        elif "book" in note:
            expected = {"is_book", "is_chapter"}
        elif "sm" in note and "is sm" in note:
            expected = {"rename_sm"}
        elif "keep 1" in note or "keep1" in note:
            expected = {"keep_1", "same_content"}
        elif "keep 2" in note or "keep2" in note:
            expected = {"keep_2", "same_content"}
        elif "identical" in note or "smaller" in note:
            expected = {"keep_1", "keep_2", "same_content"}
        if expected is None:
            continue
        checked += 1
        if v8 in expected:
            agree_v8 += 1
        else:
            v8_wrong.append((r["file_1"][:50], r["file_2"][:50], v8, r["v8_confidence"], note[:60]))
        if v7 in expected:
            agree_v7 += 1

    if checked == 0:
        return
    print(f"Parsable notes: {checked}")
    print(f"  v7 agrees: {agree_v7}/{checked} ({agree_v7/checked*100:.0f}%)")
    print(f"  v8 agrees: {agree_v8}/{checked} ({agree_v8/checked*100:.0f}%)")
    if v8_wrong:
        print(f"\nv8 disagreements ({len(v8_wrong)}):")
        for a, b, got, conf, note in v8_wrong[:15]:
            print(f"  {a}")
            print(f"  {b}   →  got={got} conf={conf}  note={note!r}")


def _hyperlink_formula(abs_path: Path, label: str = "Open") -> str:
    safe = str(abs_path).replace('"', "'")
    return f'=HYPERLINK("file://{safe}","{label}")'


def _apply_hyperlinks_and_autofilter(xlsx: Path) -> None:
    """Re-open the pandas-written workbook and (1) convert year+filename
    'Open' columns into HYPERLINK formulas; (2) attach an autofilter to
    every sheet's full range."""
    from openpyxl import load_workbook

    wb = load_workbook(xlsx)

    for ws in wb.worksheets:
        # autofilter over the used range
        if ws.max_row > 1 and ws.max_column > 0:
            ws.auto_filter.ref = ws.dimensions

        # column header -> column index (1-based)
        headers = {cell.value: cell.column for cell in ws[1] if cell.value}

        # duplicates tab: rewrite open_1 / open_2 using (year_N, file_N)
        if ws.title == "duplicates" and {"open_1", "open_2", "year_1", "file_1",
                                          "year_2", "file_2"} <= headers.keys():
            for side in (1, 2):
                open_col = headers[f"open_{side}"]
                year_col = headers[f"year_{side}"]
                file_col = headers[f"file_{side}"]
                for row in range(2, ws.max_row + 1):
                    y = ws.cell(row=row, column=year_col).value
                    f = ws.cell(row=row, column=file_col).value
                    if y is None or not f:
                        continue
                    try:
                        yr = str(int(float(y)))
                    except Exception:
                        yr = str(y)
                    p = PDF_BASE / yr / str(f)
                    ws.cell(row=row, column=open_col).value = _hyperlink_formula(p)

        # other sheets: (year, filename, open_file)
        elif {"year", "filename", "open_file"} <= headers.keys():
            open_col = headers["open_file"]
            year_col = headers["year"]
            file_col = headers["filename"]
            for row in range(2, ws.max_row + 1):
                y = ws.cell(row=row, column=year_col).value
                f = ws.cell(row=row, column=file_col).value
                if y is None or not f:
                    continue
                try:
                    yr = str(int(float(y)))
                except Exception:
                    yr = str(y)
                p = PDF_BASE / yr / str(f)
                ws.cell(row=row, column=open_col).value = _hyperlink_formula(p)

    wb.save(xlsx)


def write_workbook(updated: pd.DataFrame) -> None:
    xl = pd.ExcelFile(XLSX)
    sheets = {name: pd.read_excel(XLSX, sheet_name=name) for name in xl.sheet_names}
    sheets["duplicates"] = updated.sort_values(
        by=["v8_confidence", "step"], ascending=[False, True]
    ).reset_index(drop=True)
    with pd.ExcelWriter(XLSX, engine="openpyxl", mode="w") as w:
        for name, df in sheets.items():
            df.to_excel(w, sheet_name=name, index=False)
    _apply_hyperlinks_and_autofilter(XLSX)
    print(f"Workbook written: {XLSX}")


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--step", help="Restrict to one step label")
    ap.add_argument("--dry-run", action="store_true", help="Don't write workbook")
    ap.add_argument("--compare", action="store_true", help="Accuracy vs user notes")
    args = ap.parse_args()

    cache = FeatureCache(CACHE_PATH)
    df = pd.read_excel(XLSX, sheet_name="duplicates")
    if args.step:
        print(f"Filter: step == {args.step!r}  ({(df['step']==args.step).sum()} rows)")
        subset = df[df["step"] == args.step].copy()
        updated_subset = evaluate_rows(subset, cache)
        # merge back
        full = df.copy()
        for c in NEW_COLS:
            if c not in full.columns: full[c] = pd.NA
        full.loc[updated_subset.index, NEW_COLS] = updated_subset[NEW_COLS].values
        updated = full
    else:
        updated = evaluate_rows(df, cache)

    cnt = Counter(updated["v8_decision"])
    print("\n=== v8 decision counts ===")
    for d, n in sorted(cnt.items(), key=lambda x: -x[1]):
        print(f"  {d:<20} {n}")

    auto = (updated["v8_confidence"] >= 0.85).sum()
    review = ((updated["v8_confidence"] >= 0.55) & (updated["v8_confidence"] < 0.85)).sum()
    below = (updated["v8_confidence"] < 0.55).sum()
    print(f"\nConfidence bands: auto(>=0.85)={auto}  review(0.55-0.85)={review}  low(<0.55)={below}")

    if args.compare:
        compare_against_notes(updated)

    if not args.dry_run:
        write_workbook(updated)
    else:
        print("\n(dry-run: workbook not modified)")

    cache.close()


if __name__ == "__main__":
    main()
