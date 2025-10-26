#!/usr/bin/env python3
"""
Fix ALL boolean columns to use TRUE/FALSE instead of 1/0.
Keep 'software' as text (not boolean).
"""

import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
import numpy as np

# Paths
excel_path = Path("data/Techniques DB for Panel Review.xlsx")

# Boolean columns that should be TRUE/FALSE
BOOLEAN_COLUMNS = [
    'is_parent',
    'data_collection_technology',
    'statistical_model',
    'analytical_algorithm',
    'inference_framework',
    'conceptual_field',
    'data_science'  # Add data_science!
]

# Software should be TEXT (tool names), not boolean!
# It got converted to 0/1 by mistake


def main():
    print("=" * 80)
    print("FIXING ALL BOOLEAN COLUMNS")
    print("=" * 80)

    # Load Excel
    print(f"\nLoading: {excel_path}")
    df = pd.read_excel(excel_path, sheet_name='Full_List')
    print(f"Loaded {len(df):,} techniques")

    # Create backup
    print(f"\n📦 Creating backup...")
    backup_path = excel_path.with_suffix('.backup4.xlsx')
    import shutil
    shutil.copy2(excel_path, backup_path)
    print(f"✓ Backup created: {backup_path}")

    # Save using openpyxl
    print(f"\n💾 Fixing and saving to: {excel_path}")

    # Load workbook
    wb = load_workbook(excel_path)
    ws = wb['Full_List']

    # Get column mapping
    headers = [cell.value for cell in ws[1]]
    col_mapping = {header: idx + 1 for idx, header in enumerate(headers)}

    print(f"\nConverting columns to TRUE/FALSE:")

    # Fix boolean columns
    for col_name in BOOLEAN_COLUMNS:
        if col_name not in col_mapping:
            print(f"  ⚠️  Column '{col_name}' not found")
            continue

        col_idx = col_mapping[col_name]
        true_count = 0
        false_count = 0

        # Update all cells in this column
        for row_idx in range(2, len(df) + 2):  # Start from row 2 (skip header)
            cell = ws.cell(row=row_idx, column=col_idx)
            value = cell.value

            # Convert to boolean
            if value == 1 or value == 1.0 or value == '1' or value is True or value == 'TRUE':
                ws.cell(row=row_idx, column=col_idx, value=True)
                true_count += 1
            else:
                ws.cell(row=row_idx, column=col_idx, value=False)
                false_count += 1

        print(f"  {col_name:30s}: {true_count:>3} TRUE, {false_count:>3} FALSE")

    # Fix 'software' column - should be text, not boolean
    if 'software' in col_mapping:
        print(f"\nFixing 'software' column (text, not boolean):")
        col_idx = col_mapping['software']

        # Clear 0/1 values, leave only actual software names
        cleared_count = 0
        kept_count = 0

        for row_idx in range(2, len(df) + 2):
            cell = ws.cell(row=row_idx, column=col_idx)
            value = cell.value

            # If it's 0 or 1, clear it (should be text)
            if value == 0 or value == 0.0 or value == '0':
                ws.cell(row=row_idx, column=col_idx, value=None)
                cleared_count += 1
            elif value == 1 or value == 1.0 or value == '1':
                ws.cell(row=row_idx, column=col_idx, value=None)
                cleared_count += 1
            elif value is not None and str(value).strip():
                # Keep actual text values
                kept_count += 1

        print(f"  Cleared {cleared_count} numeric values (0/1)")
        print(f"  Kept {kept_count} text values (software names)")

    # Save
    wb.save(excel_path)
    print(f"\n✓ Saved successfully")

    # Verify
    print("\n" + "=" * 80)
    print("VERIFICATION")
    print("=" * 80)

    df_verify = pd.read_excel(excel_path, sheet_name='Full_List')

    print("\nBoolean columns:")
    for col in BOOLEAN_COLUMNS:
        if col in df_verify.columns:
            true_count = (df_verify[col] == True).sum()
            false_count = (df_verify[col] == False).sum()
            na_count = df_verify[col].isna().sum()
            print(f"  {col:30s}: {true_count:>3} TRUE, {false_count:>3} FALSE, {na_count:>3} NA")

    print("\nSoftware column:")
    if 'software' in df_verify.columns:
        non_null = df_verify['software'].notna().sum()
        null = df_verify['software'].isna().sum()
        print(f"  Non-null (has software name): {non_null}")
        print(f"  Null (no software): {null}")

        if non_null > 0:
            print(f"\n  Sample software values:")
            samples = df_verify[df_verify['software'].notna()]['software'].head(10)
            for idx, val in samples.items():
                tech_name = df_verify.loc[idx, 'technique_name']
                print(f"    {tech_name:40s}: {val}")

    print("\n" + "=" * 80)
    print("COMPLETE")
    print("=" * 80)
    print("\n✓ All boolean columns fixed!")


if __name__ == "__main__":
    main()
