#!/usr/bin/env python3
"""
Fix issues in Rachel's technique additions:
1. Convert 1/0 back to TRUE/FALSE for boolean columns
2. Fix conceptual_field (should be TRUE/FALSE not text)
3. Populate remove_reason column (should be empty for active techniques)
4. Add index values as increments from max existing index
"""

import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
import numpy as np

# Paths
excel_path = Path("data/Techniques DB for Panel Review.xlsx")

# Boolean columns that should be TRUE/FALSE
BOOLEAN_COLUMNS = [
    'is_parent',
    'data_collection_technology',
    'statistical_model',
    'analytical_algorithm',
    'inference_framework',
    'conceptual_field'  # This should also be TRUE/FALSE!
]


def convert_to_true_false(df):
    """Convert 1/0 values to TRUE/FALSE in boolean columns."""
    print("Converting boolean columns from 1/0 to TRUE/FALSE...")

    for col in BOOLEAN_COLUMNS:
        if col in df.columns:
            # Convert numeric values to boolean
            df[col] = df[col].apply(lambda x:
                True if pd.notna(x) and (x == 1 or x == '1' or x is True or x == 'TRUE' or x == 'True')
                else False if pd.notna(x) and (x == 0 or x == '0' or x is False or x == 'FALSE' or x == 'False')
                else False  # Default to False for NaN/None
            )

            # Count changes
            true_count = (df[col] == True).sum()
            false_count = (df[col] == False).sum()
            print(f"  {col:30s}: {true_count:>3} TRUE, {false_count:>3} FALSE")

    return df


def populate_remove_reason(df):
    """Populate remove_reason column (should be empty for active techniques)."""
    print("\nPopulating remove_reason column...")

    # For active techniques, remove_reason should be empty/None
    # Only set if there's already a value
    if 'remove_reason' in df.columns:
        empty_count = df['remove_reason'].isna().sum()
        print(f"  Empty remove_reason: {empty_count} (leaving as empty)")

    return df


def add_index_values(df):
    """Add index values for rows missing them."""
    print("\nAdding index values...")

    if 'index' not in df.columns:
        print("  ⚠️  No 'index' column found")
        return df

    # Find max existing index
    max_idx = df['index'].max()
    print(f"  Max existing index: {max_idx}")

    # Find rows with missing index
    missing_idx = df['index'].isna()
    missing_count = missing_idx.sum()

    if missing_count == 0:
        print("  ✓ All rows have index values")
        return df

    print(f"  Adding index to {missing_count} rows...")

    # Assign sequential indices starting from max_idx + 1
    next_idx = int(max_idx) + 1

    for idx, row in df[missing_idx].iterrows():
        df.at[idx, 'index'] = next_idx
        print(f"    Row {idx}: {row['technique_name']:40s} -> index {next_idx}")
        next_idx += 1

    return df


def main():
    print("=" * 80)
    print("FIXING RACHEL'S EXCEL ENTRIES")
    print("=" * 80)

    # Load Excel
    print(f"\nLoading: {excel_path}")
    df = pd.read_excel(excel_path, sheet_name='Full_List')
    print(f"Loaded {len(df):,} techniques")

    # Show current state
    print("\nCurrent state:")
    print(f"  Rows: {len(df)}")
    print(f"  Columns: {len(df.columns)}")

    # Fix boolean columns
    df = convert_to_true_false(df)

    # Populate remove_reason (leave empty)
    df = populate_remove_reason(df)

    # Add index values
    df = add_index_values(df)

    # Create backup
    print(f"\n📦 Creating backup...")
    backup_path = excel_path.with_suffix('.backup3.xlsx')
    import shutil
    shutil.copy2(excel_path, backup_path)
    print(f"✓ Backup created: {backup_path}")

    # Save using openpyxl to preserve formatting
    print(f"\n💾 Saving to: {excel_path}")

    # Load workbook
    wb = load_workbook(excel_path)
    ws = wb['Full_List']

    # Get column mapping
    headers = [cell.value for cell in ws[1]]
    col_mapping = {header: idx + 1 for idx, header in enumerate(headers)}

    print(f"  Updating {len(df)} rows...")

    # Update each cell
    for row_idx, row in df.iterrows():
        excel_row = row_idx + 2  # Excel is 1-indexed, skip header

        # Update boolean columns (write TRUE/FALSE as boolean)
        for col_name in BOOLEAN_COLUMNS:
            if col_name in col_mapping:
                col_idx = col_mapping[col_name]
                value = row[col_name]

                # Convert to Excel boolean (True/False)
                if pd.notna(value):
                    ws.cell(row=excel_row, column=col_idx, value=bool(value))
                else:
                    ws.cell(row=excel_row, column=col_idx, value=False)

        # Update index
        if 'index' in col_mapping:
            col_idx = col_mapping['index']
            value = row['index']
            if pd.notna(value):
                ws.cell(row=excel_row, column=col_idx, value=int(value))

        # Update remove_reason (leave empty unless there's a value)
        if 'remove_reason' in col_mapping:
            col_idx = col_mapping['remove_reason']
            value = row['remove_reason']
            if pd.notna(value) and str(value).strip():
                ws.cell(row=excel_row, column=col_idx, value=str(value))
            else:
                ws.cell(row=excel_row, column=col_idx, value=None)

    # Save
    wb.save(excel_path)
    print(f"✓ Saved successfully")

    # Verify
    print("\n" + "=" * 80)
    print("VERIFICATION")
    print("=" * 80)

    df_verify = pd.read_excel(excel_path, sheet_name='Full_List')

    print("\nBoolean column values:")
    for col in BOOLEAN_COLUMNS:
        if col in df_verify.columns:
            true_count = (df_verify[col] == True).sum()
            false_count = (df_verify[col] == False).sum()
            print(f"  {col:30s}: {true_count:>3} TRUE, {false_count:>3} FALSE")

    print("\nIndex values:")
    print(f"  Min: {df_verify['index'].min()}")
    print(f"  Max: {df_verify['index'].max()}")
    print(f"  Missing: {df_verify['index'].isna().sum()}")

    print("\n" + "=" * 80)
    print("COMPLETE")
    print("=" * 80)
    print("\n✓ All fixes applied successfully!")


if __name__ == "__main__":
    main()
