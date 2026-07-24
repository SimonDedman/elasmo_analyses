#!/usr/bin/env python3
"""
Scan coauthor drop folders under database/others_libraries/, file what matches,
and route the rest to a review sheet.

Every PDF that lands in a coauthor folder was downloaded by a team member
clicking a link we generated from the remaining-todo list
(docs/remaining_downloads.html, backed by docs/papers_data.json). Each delivery
therefore corresponds to a known literature_id by construction. A PDF that fails
to match is overwhelmingly a matching failure, NOT a paper Shark References has
missed, so this script never orphan-stages anything: unresolved files go to an
xlsx for a human to look at.

The download helper logs every click to a Google Apps Script endpoint, so
deliveries can be attributed to whoever fetched them. Pass --clicks to fold that
attribution into the review sheet (needs network; skipped otherwise).

Usage:
    python3 scripts/scan_coauthor_libraries.py --check       # dry run, report only
    python3 scripts/scan_coauthor_libraries.py               # file matches, write review sheet
    python3 scripts/scan_coauthor_libraries.py --person Elena
    python3 scripts/scan_coauthor_libraries.py --clicks      # add click attribution
"""

from __future__ import annotations

import argparse
import json
import sys
import urllib.request
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))

import ingest_pdfs  # noqa: E402
from ingest_pdfs import (  # noqa: E402
    LOG_DIR,
    PDF_BASE,
    PROJECT,
    build_filename,
    extract_doi_from_pdf,
    ingest_source,
    load_database,
    match_pdf,
    normalise_doi,
)

LIBRARIES = PROJECT / "database/others_libraries"
REVIEW_DIR = PROJECT / "outputs"
LOG_FILE = LOG_DIR / "scan_coauthor_libraries_log.txt"
STATE_FILE = PROJECT / "outputs/.coauthor_scan_state.json"

# Folders under others_libraries that hold metadata rather than deliveries.
SKIP_DIRS = {"SharkRefsExport"}

# Subfolders that are not todo-list deliveries and must not be reported as
# unresolved matches. Carylanne's programme folders are whole conference
# abstract books feeding the book-chapter/abstract mining subproject; they have
# no single DOI and were never queue entries.
SKIP_SUBDIRS = {"digitised programs", "undigitised programs"}

CLICKS_URL = (
    "https://script.google.com/macros/s/"
    "AKfycbwCmkL89I8GGK3-IoCZh9x9XAVpvTshOysMlnWiRmqoXAtICFO16TkljEPlxTwXaufR"
    "/exec?action=getAll"
)


def fetch_clicks() -> dict[str, dict]:
    """DOI -> {by, at} from the download helper's shared click log."""
    try:
        with urllib.request.urlopen(CLICKS_URL, timeout=60) as resp:
            payload = json.load(resp)
    except Exception as e:  # noqa: BLE001 - network/JSON, all non-fatal
        print(f"  WARNING: could not fetch click log ({e}); continuing without it")
        return {}
    out: dict[str, dict] = {}
    for row in payload.get("data", []):
        nd = normalise_doi(row.get("doi") or "")
        if nd:
            out[nd] = {"by": row.get("by", ""), "at": row.get("at", "")}
    return out


def verified_in_library(src: Path, all_rows: list[dict], literature_id: str) -> bool:
    """True only if this paper is demonstrably filed in the library already.

    Mirrors acquire_cascade's verified-delete gate: a source file is removed
    only when the corpus row it matched has a real PDF on disk. Deleting a
    coauthor's only copy because a match was recorded but a copy silently
    failed would be unrecoverable.
    """
    row = next((r for r in all_rows
                if str(r.get("literature_id", "")).replace(".0", "") ==
                str(literature_id).replace(".0", "")), None)
    if row is None:
        return False
    try:
        year = str(int(float(row.get("year") or 0)))
    except (ValueError, TypeError):
        return False
    dest = PDF_BASE / year / build_filename(row)
    if dest.exists() and dest.stat().st_size > 1024:
        return True
    # Filed under a different name in the right year folder: accept only an
    # unambiguous single match on size, never a guess.
    folder = PDF_BASE / year
    if folder.is_dir():
        same = [q for q in folder.glob("*.pdf")
                if q.stat().st_size == src.stat().st_size]
        if len(same) == 1:
            return True
    return False


def load_state() -> dict:
    """Files already scanned, so repeat runs only report what's new."""
    if STATE_FILE.exists():
        try:
            return json.loads(STATE_FILE.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            pass
    return {"seen": {}}


def save_state(state: dict) -> None:
    STATE_FILE.write_text(json.dumps(state, indent=1), encoding="utf-8")


def discover(person: str | None) -> dict[str, list[Path]]:
    """Coauthor name -> PDFs found anywhere beneath their folder."""
    found: dict[str, list[Path]] = {}
    if not LIBRARIES.exists():
        return found
    for d in sorted(p for p in LIBRARIES.iterdir() if p.is_dir()):
        if d.name in SKIP_DIRS:
            continue
        if person and d.name.lower() != person.lower():
            continue
        pdfs = sorted(
            p for p in (set(d.rglob("*.pdf")) | set(d.rglob("*.PDF")))
            if not any(part.lower() in SKIP_SUBDIRS for part in p.parts)
        )
        if pdfs:
            found[d.name] = pdfs
    return found


def write_review_sheet(rows: list[dict], path: Path) -> bool:
    """Formatted xlsx (frozen + bold + autofiltered header). False if openpyxl absent."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
    except ImportError:
        return False

    headers = ["person", "filename", "status", "reason", "pdf_doi",
               "clicked_by", "clicked_at", "path"]
    wb = Workbook()
    ws = wb.active
    ws.title = "unresolved"
    ws.append(headers)
    for r in rows:
        ws.append([r.get(h, "") for h in headers])

    for c in ws[1]:
        c.font = Font(bold=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    for i, h in enumerate(headers, start=1):
        width = max(len(h), *(len(str(r.get(h, ""))) for r in rows)) if rows else len(h)
        ws.column_dimensions[get_column_letter(i)].width = min(width + 2, 70)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return True


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--check", action="store_true",
                    help="Dry run: report matches without filing anything")
    ap.add_argument("--person", help="Limit to one coauthor folder")
    ap.add_argument("--clicks", action="store_true",
                    help="Fetch the click log to attribute deliveries")
    ap.add_argument("--all", action="store_true",
                    help="Re-report files seen on a previous run")
    ap.add_argument("--no-ocr", action="store_true", help="Disable OCR fallback")
    ap.add_argument("--delete-ingested", action="store_true",
                    help="Delete source PDFs that were verifiably filed into the "
                         "library (same gate as acquire_cascade: the file must "
                         "have matched AND the library copy must exist and be "
                         ">1KB). Off by default — run a clean pass first.")
    args = ap.parse_args()

    if args.no_ocr:
        ingest_pdfs.OCR_ENABLED = False

    LOG_DIR.mkdir(exist_ok=True)
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    libraries = discover(args.person)
    if not libraries:
        print("No coauthor PDFs found.")
        return 0

    state = load_state()
    seen: dict = state.get("seen", {})

    print(f"{'=' * 70}\n  Coauthor library scan — {timestamp}\n{'=' * 70}")
    for person, pdfs in libraries.items():
        fresh = [p for p in pdfs if args.all or str(p) not in seen]
        print(f"  {person:<12} {len(pdfs):>4} PDFs  ({len(fresh)} new)")

    clicks = fetch_clicks() if args.clicks else {}
    if clicks:
        print(f"  click log: {len(clicks)} DOIs")

    print("\nLoading corpus...")
    all_rows, doi_lookup, author_year_lookup = load_database()
    print(f"  {len(all_rows)} rows, {len(doi_lookup)} DOIs")

    unresolved: list[dict] = []
    deleted: list[str] = []
    matched_total = 0
    log_lines: list[str] = [f"Coauthor library scan {timestamp}"]

    for person, pdfs in libraries.items():
        todo = [p for p in pdfs if args.all or str(p) not in seen]
        if not todo:
            continue

        if args.check:
            print(f"\n{'=' * 70}\n  {person} — {len(todo)} PDFs (dry run)\n{'=' * 70}")
            for i, p in enumerate(todo, 1):
                row, method = match_pdf(p, doi_lookup, author_year_lookup, all_rows)
                if row:
                    matched_total += 1
                    print(f"  [{i:3}] OK   {p.name[:64]}")
                    print(f"        -> {row['literature_id']} {str(row['title'])[:62]}")
                else:
                    nd = normalise_doi(extract_doi_from_pdf(p) or "")
                    click = clicks.get(nd, {})
                    print(f"  [{i:3}] MISS {p.name[:64]}")
                    print(f"        {method[:90]}")
                    unresolved.append({
                        "person": person, "filename": p.name, "status": "unresolved",
                        "reason": method, "pdf_doi": nd,
                        "clicked_by": click.get("by", ""),
                        "clicked_at": click.get("at", ""), "path": str(p),
                    })
        else:
            filed_map: dict = {}
            _ids, _dois, _names, lines = ingest_source(
                person, todo, doi_lookup, author_year_lookup, all_rows,
                filed_map=filed_map)
            log_lines.extend(lines)
            matched_total += len(filed_map)
            for p in todo:
                if str(p) in filed_map:
                    seen[str(p)] = {"person": person, "at": timestamp,
                                    "literature_id": str(filed_map[str(p)])}
                    continue
                row, method = match_pdf(p, doi_lookup, author_year_lookup, all_rows)
                if row:
                    continue
                nd = normalise_doi(extract_doi_from_pdf(p) or "")
                click = clicks.get(nd, {})
                unresolved.append({
                    "person": person, "filename": p.name, "status": "unresolved",
                    "reason": method, "pdf_doi": nd,
                    "clicked_by": click.get("by", ""),
                    "clicked_at": click.get("at", ""), "path": str(p),
                })

    # Deletion sweep. Deliberately separate from the filing loop and run over
    # every discovered PDF, not just this run's new ones: a file filed on an
    # earlier pass is skipped as already-seen, so folding deletion into filing
    # would leave those sitting in the coauthor folder forever.
    if args.delete_ingested:
        for person, pdfs in libraries.items():
            for p in pdfs:
                record = seen.get(str(p))
                if not record or not record.get("literature_id"):
                    continue
                if not p.exists():
                    continue
                if not verified_in_library(p, all_rows, record["literature_id"]):
                    print(f"  KEPT (no verified library copy): {p.name[:56]}")
                    continue
                if args.check:
                    deleted.append(str(p))
                    print(f"  WOULD DELETE (filed as "
                          f"{record['literature_id']}): {p.name[:52]}")
                    continue
                try:
                    p.unlink()
                    deleted.append(str(p))
                    print(f"  DELETED source (filed as "
                          f"{record['literature_id']}): {p.name[:52]}")
                except OSError as e:
                    print(f"  WARNING: could not delete {p.name}: {e}")

    print(f"\n{'=' * 70}\n  SUMMARY\n{'=' * 70}")
    print(f"  Matched:     {matched_total}")
    print(f"  Unresolved:  {len(unresolved)}")
    if args.delete_ingested:
        print(f"  Deleted:     {len(deleted)} source PDFs")

    if unresolved:
        stamp = datetime.now().strftime("%Y-%m-%d")
        sheet = REVIEW_DIR / f"coauthor_unresolved_{stamp}.xlsx"
        if write_review_sheet(unresolved, sheet):
            print(f"  Review sheet: {sheet}")
        else:
            fallback = sheet.with_suffix(".csv")
            import csv as _csv
            with open(fallback, "w", newline="", encoding="utf-8") as f:
                w = _csv.DictWriter(f, fieldnames=list(unresolved[0].keys()))
                w.writeheader()
                w.writerows(unresolved)
            print(f"  openpyxl missing — wrote {fallback}")
        print("  These are almost certainly matcher failures, not new papers.")
        print("  Check the DOI against docs/papers_data.json before staging any of them.")

    if not args.check:
        state["seen"] = seen
        save_state(state)
        with open(LOG_FILE, "a", encoding="utf-8") as f:
            f.write(f"\n{'=' * 70}\n")
            f.write("\n".join(log_lines) + "\n")
            f.write(f"matched={matched_total} unresolved={len(unresolved)}\n")
        print(f"  Log: {LOG_FILE}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
