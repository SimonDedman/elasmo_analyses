#!/usr/bin/env python3
"""
Populate the data_science column in Techniques DB for Panel Review.xlsx

Logic:
- TRUE if discipline = DATA
- TRUE if technique is a data science/ML/statistical modeling technique
- FALSE otherwise
"""

import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Paths
excel_path = Path("data/Techniques DB for Panel Review.xlsx")
output_path = Path("data/Techniques DB for Panel Review.xlsx")

# Known data science techniques (case-insensitive matching)
DATA_SCIENCE_TECHNIQUES = {
    # Machine Learning / Statistical Learning
    'brt', 'boosted regression', 'gbm', 'xgboost', 'gradient boosting',
    'random forest', 'random forests',
    'neural network', 'neural networks', 'deep learning', 'cnn', 'convolutional neural',
    'support vector machine', 'svm',
    'decision tree', 'decision trees',
    'classification tree', 'regression tree',
    'ensemble', 'ensemble learning',
    'bagging', 'bootstrap aggregating',
    'adaboost', 'adaptive boosting',
    'k-nearest neighbor', 'knn', 'k-nn',
    'naive bayes',
    'logistic regression',
    'linear discriminant analysis', 'lda',
    'quadratic discriminant analysis', 'qda',

    # Statistical Modeling
    'generalized linear model', 'glm', 'glms',
    'generalized additive model', 'gam', 'gams',
    'generalized linear mixed model', 'glmm',
    'generalized additive mixed model', 'gamm',
    'mixed model', 'mixed models', 'mixed-effects',
    'hierarchical model', 'hierarchical models',
    'bayesian', 'bayesian inference', 'bayesian modeling',
    'markov chain monte carlo', 'mcmc',
    'state space model', 'state-space',
    'hidden markov model', 'hmm',

    # Multivariate Statistics
    'permanova', 'permdisp',
    'principal component analysis', 'pca',
    'principal components analysis',
    'correspondence analysis',
    'canonical correspondence analysis', 'cca',
    'redundancy analysis', 'rda',
    'non-metric multidimensional scaling', 'nmds', 'nms',
    'multidimensional scaling', 'mds',
    'discriminant analysis',
    'cluster analysis', 'clustering',
    'hierarchical clustering',
    'k-means', 'kmeans',
    'self-organizing map', 'som',
    'factor analysis',
    'structural equation model', 'sem',

    # Time Series / Spatial
    'autoregressive', 'arima', 'arma',
    'time series', 'time-series',
    'spatial', 'geostatistics', 'kriging',
    'geographically weighted regression', 'gwr',
    'spatial autocorrelation',
    'moran', "moran's i",

    # Model Selection / Validation
    'cross-validation', 'cross validation',
    'bootstrap', 'bootstrapping',
    'permutation test', 'permutation tests',
    'information criterion', 'aic', 'bic', 'aicc',
    'model selection',
    'model averaging',

    # Specific Ecological Models (data-driven)
    'maxent', 'maximum entropy',
    'species distribution model', 'sdm', 'species distribution modeling',
    'ecological niche model', 'enm',
    'habitat suitability model',
    'occupancy model', 'occupancy modeling',
    'capture-recapture', 'mark-recapture',
    'distance sampling',

    # Network / Graph Analysis
    'network analysis',
    'social network analysis', 'sna',
    'graph theory',

    # Natural Language Processing (if any)
    'text mining',
    'topic model', 'topic modeling',
    'sentiment analysis',

    # Computer Vision (if any)
    'image recognition',
    'object detection',
    'image classification',
    'computer vision',

    # Simulation / Computational
    'agent-based model', 'agent based model', 'abm',
    'individual-based model', 'ibm',
    'monte carlo',
    'simulation',
}

# Phrases that indicate data science (for partial matching)
DATA_SCIENCE_PHRASES = [
    'machine learning',
    'statistical learning',
    'predictive model',
    'classification',
    'regression',
    'clustering',
    'dimension reduction',
    'feature selection',
    'variable selection',
    'model validation',
    'model comparison',
]


def is_data_science_technique(technique_name, synonyms, discipline):
    """
    Determine if a technique is a data science technique.

    Args:
        technique_name: Name of the technique
        synonyms: Synonyms (comma-separated string or NaN)
        discipline: Discipline code (e.g., DATA, BIO, etc.)

    Returns:
        True if data science technique, False otherwise
    """
    # If discipline is DATA, it's definitely data science
    if pd.notna(discipline) and discipline.strip().upper() == 'DATA':
        return True

    # Check technique name
    if pd.notna(technique_name):
        tech_lower = str(technique_name).lower().strip()

        # Exact match with known techniques
        if tech_lower in DATA_SCIENCE_TECHNIQUES:
            return True

        # Check for phrases in technique name
        for phrase in DATA_SCIENCE_PHRASES:
            if phrase in tech_lower:
                return True

    # Check synonyms
    if pd.notna(synonyms):
        syn_lower = str(synonyms).lower()

        # Split synonyms and check each
        for syn in syn_lower.split(','):
            syn = syn.strip()

            # Exact match
            if syn in DATA_SCIENCE_TECHNIQUES:
                return True

            # Phrase match
            for phrase in DATA_SCIENCE_PHRASES:
                if phrase in syn:
                    return True

    return False


def main():
    print("=" * 80)
    print("POPULATING DATA_SCIENCE COLUMN")
    print("=" * 80)

    # Load Excel file (Full_List sheet)
    print(f"\nLoading: {excel_path} (Full_List sheet)")
    df = pd.read_excel(excel_path, sheet_name='Full_List')
    print(f"Loaded {len(df):,} techniques")

    # Check if data_science column exists, add if not
    if 'data_science' not in df.columns:
        print("\n⚠️  'data_science' column not found - creating it")
        df['data_science'] = None
    else:
        print(f"✓ Found 'data_science' column")

    # Count existing values
    existing_true = (df['data_science'] == True).sum()
    existing_false = (df['data_science'] == False).sum()
    existing_na = df['data_science'].isna().sum()

    print(f"\nExisting values:")
    print(f"  TRUE:  {existing_true:>4}")
    print(f"  FALSE: {existing_false:>4}")
    print(f"  NA:    {existing_na:>4}")

    # Populate data_science column
    print(f"\nProcessing techniques...")

    data_science_count = 0
    not_data_science_count = 0

    for idx, row in df.iterrows():
        technique_name = row.get('technique_name', '')
        synonyms = row.get('synonyms', '')
        discipline = row.get('discipline', '')

        is_ds = is_data_science_technique(technique_name, synonyms, discipline)

        df.at[idx, 'data_science'] = is_ds

        if is_ds:
            data_science_count += 1
        else:
            not_data_science_count += 1

    print(f"\nResults:")
    print(f"  Data science techniques:     {data_science_count:>4} (TRUE)")
    print(f"  Non-data science techniques: {not_data_science_count:>4} (FALSE)")

    # Show breakdown by discipline
    print(f"\nBreakdown by discipline:")
    discipline_counts = df.groupby(['discipline', 'data_science']).size().unstack(fill_value=0)

    if True in discipline_counts.columns:
        discipline_counts = discipline_counts.sort_values(True, ascending=False)

    print(discipline_counts)

    # Show some examples
    print(f"\n" + "=" * 80)
    print("EXAMPLES OF DATA SCIENCE TECHNIQUES:")
    print("=" * 80)

    ds_techniques = df[df['data_science'] == True].head(20)
    for _, row in ds_techniques.iterrows():
        tech = row['technique_name']
        disc = row.get('discipline', 'N/A')
        print(f"  [{disc:5s}] {tech}")

    if len(df[df['data_science'] == True]) > 20:
        print(f"  ... and {len(df[df['data_science'] == True]) - 20} more")

    print(f"\n" + "=" * 80)
    print("EXAMPLES OF NON-DATA SCIENCE TECHNIQUES:")
    print("=" * 80)

    non_ds_techniques = df[df['data_science'] == False].head(20)
    for _, row in non_ds_techniques.iterrows():
        tech = row['technique_name']
        disc = row.get('discipline', 'N/A')
        print(f"  [{disc:5s}] {tech}")

    if len(df[df['data_science'] == False]) > 20:
        print(f"  ... and {len(df[df['data_science'] == False]) - 20} more")

    # Save updated Excel file (preserving other sheets)
    print(f"\n" + "=" * 80)
    print("SAVING UPDATED FILE")
    print("=" * 80)

    print(f"\nSaving to: {output_path}")

    # Create a backup first
    backup_path = output_path.with_suffix('.backup.xlsx')
    import shutil
    shutil.copy2(output_path, backup_path)
    print(f"✓ Backup created: {backup_path}")

    # Use openpyxl to load and modify
    from openpyxl import load_workbook

    wb = load_workbook(excel_path)
    ws = wb['Full_List']

    # Update the header to add data_science column if needed
    headers = [cell.value for cell in ws[1]]
    if 'data_science' not in headers:
        # Add data_science header to next available column
        ws.cell(row=1, column=len(headers) + 1, value='data_science')
        print(f"✓ Added 'data_science' column header")

    # Find column indices
    col_mapping = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}
    data_science_col = col_mapping['data_science']

    # Update data_science column for each row
    for idx, row in df.iterrows():
        excel_row = idx + 2  # +2 because Excel is 1-indexed and we skip header
        value = row['data_science']
        ws.cell(row=excel_row, column=data_science_col, value=value)

    # Save
    wb.save(output_path)
    print(f"✓ Saved successfully (all sheets preserved)")

    print(f"\n" + "=" * 80)
    print("COMPLETE")
    print("=" * 80)

    # Summary
    print(f"\nSummary:")
    print(f"  Total techniques: {len(df):,}")
    print(f"  Data science: {data_science_count:,} ({data_science_count/len(df)*100:.1f}%)")
    print(f"  Non-data science: {not_data_science_count:,} ({not_data_science_count/len(df)*100:.1f}%)")
    print(f"\n✓ data_science column populated successfully!")


if __name__ == "__main__":
    main()
